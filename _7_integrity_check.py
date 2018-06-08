from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText as tkst
from openpyxl import load_workbook
import re
import os


class FileCheck(Toplevel):
    def __init__(self):
        self.d = Data()
        Toplevel.__init__(self)
        self.title("Data Integrity Check")
        self.tk.call("wm", "iconphoto", self._w, PhotoImage(data=self.d.logo))
        self.geometry("450x600")
        self.resizable(width=False, height=False)

        frm = ttk.Frame(self, padding=20)
        frm.pack(fill=BOTH, expand=YES)

        top = ttk.Frame(frm)
        mid = ttk.Frame(frm)
        bot = ttk.Frame(frm)
        top.place(relx=0, rely=0, relwidth=1, relheight=0.57)
        mid.place(relx=0, rely=0.57, relwidth=1, relheight=0.27)
        bot.place(relx=0, rely=0.84, relwidth=1, relheight=0.16)
        bot.rowconfigure(0, weight=1)
        bot.rowconfigure(1, weight=1)
        bot.columnconfigure(0, weight=1)
        bot.columnconfigure(1, weight=1)
        bot.columnconfigure(2, weight=1)

        newbook = ttk.Notebook(top)
        newbook.pack(fill=BOTH, expand=YES)
        newbook.bind("<<NotebookTabChanged>>", lambda event: self.switch_tab(event))
        xlsx_tab = ttk.Frame(newbook)
        eepr_tab = ttk.Frame(newbook)
        newbook.add(xlsx_tab, text=" Excel ")
        newbook.add(eepr_tab, text=" EEPROM ")

        self.xlsx_cv = Canvas(xlsx_tab)
        self.xlsx_frm = ttk.Frame(self.xlsx_cv)
        xlsx_sb = Scrollbar(xlsx_tab, orient=VERTICAL, command=self.xlsx_cv.yview)
        xlsx_ctrl = ttk.Frame(xlsx_tab, relief=SOLID, padding=5)
        self.xlsx_cv.place(relx=0.015, rely=0.017, relwidth=0.94, relheight=0.743)
        self.xlsx_cv.create_window(0, 0, anchor=NW, window=self.xlsx_frm)
        xlsx_sb.place(relx=0.955, rely=0.017, relwidth=0.045, relheight=0.743)
        self.xlsx_cv.config(yscrollcommand=xlsx_sb.set)
        xlsx_tab.bind("<Configure>", self.xlsx_resize)
        xlsx_ctrl.place(relx=0.015, rely=0.783, relwidth=0.97, relheight=0.2)
        xlsx_ctrl.columnconfigure(0, weight=1)
        xlsx_ctrl.columnconfigure(1, weight=1)
        xlsx_ctrl.columnconfigure(2, weight=1)
        self.xlsx_imp = ttk.Label(self.xlsx_frm, textvariable=self.d.xlsxfilename, font=("Consolas 10", 8), foreground="gray50", wraplength=375)
        Checkbutton(xlsx_ctrl, text="All             ", variable=self.d.xlsx_check_all, command=self.check_all0).grid(row=0, column=0, sticky=NW)
        Checkbutton(xlsx_ctrl, text="EDID            ", variable=self.d.xlsx_check[0], command=self.check_one0).grid(row=1, column=0, sticky=NW)
        Checkbutton(xlsx_ctrl, text="Bits in Sequence", variable=self.d.xlsx_check[1], command=self.check_one0).grid(row=0, column=1, padx=(5, 0), sticky=NW)
        Checkbutton(xlsx_ctrl, text="Cell Format     ", variable=self.d.xlsx_check[2], command=self.check_one0).grid(row=1, column=1, padx=(5, 0), sticky=NW)
        Checkbutton(xlsx_ctrl, text="Indicators      ", variable=self.d.xlsx_check[3], command=self.check_one0).grid(row=0, column=2, padx=(5, 0), sticky=NW)
        #Checkbutton(xlsx_ctrl, text="").grid(row=1, column=2, padx=(5, 0), sticky=NW)

        self.eepr_cv = Canvas(eepr_tab)
        self.eepr_frm = ttk.Frame(self.eepr_cv)
        eepr_sb = Scrollbar(eepr_tab, orient=VERTICAL, command=self.eepr_cv.yview)
        eepr_ctrl = ttk.Frame(eepr_tab, relief=SOLID, padding=5)
        self.eepr_cv.place(relx=0.015, rely=0.017, relwidth=0.94, relheight=0.743)
        self.eepr_cv.create_window(0, 0, anchor=NW, window=self.eepr_frm)
        eepr_sb.place(relx=0.955, rely=0.017, relwidth=0.045, relheight=0.743)
        self.eepr_cv.config(yscrollcommand=eepr_sb.set)
        eepr_tab.bind("<Configure>", self.eepr_resize)
        eepr_ctrl.place(relx=0.015, rely=0.783, relwidth=0.97, relheight=0.2)
        eepr_ctrl.columnconfigure(0, weight=1)
        eepr_ctrl.columnconfigure(1, weight=1)
        eepr_ctrl.columnconfigure(2, weight=1)
        self.eepr_imp = ttk.Label(self.eepr_frm, textvariable=self.d.eeprfilename, font=("Consolas 10", 8), foreground="gray50", wraplength=375)
        Checkbutton(eepr_ctrl, text="All             ", variable=self.d.eepr_check_all, command=self.check_all1).grid(row=0, column=0, sticky=NW)
        Checkbutton(eepr_ctrl, text="Data Format     ", variable=self.d.eepr_check[0], command=self.check_one1).grid(row=1, column=0, sticky=NW)
        Checkbutton(eepr_ctrl, text="Main Page       ", variable=self.d.eepr_check[1], command=self.check_one1).grid(row=0, column=1, padx=(5, 0), sticky=NW)
        Checkbutton(eepr_ctrl, text="Pages           ", variable=self.d.eepr_check[2], command=self.check_one1).grid(row=1, column=1, padx=(5, 0), sticky=NW)
        Checkbutton(eepr_ctrl, text="Checksum        ", variable=self.d.eepr_check[3], command=self.check_one1).grid(row=0, column=2, padx=(5, 0), sticky=NW)
        #Checkbutton(eepr_ctrl, text="", variable=self.d.eepr_check[4], command=self.check_one1).grid(row=1, column=2, padx=(5, 0), sticky=NW)

        self.log_box = tkst(mid, state=DISABLED, relief=RIDGE, bd=2)
        self.log_box.pack(side=TOP, fill=BOTH, expand=YES, pady=10)
        self.log_box.tag_config("underline", foreground="blue", underline=1)
        self.log_box.tag_config("red_tag", foreground="red")
        self.log_box.tag_config("blue_tag", foreground="blue")
        self.log_box.tag_bind("blue_tag", "<Button-1>", lambda event: self.hypertext_click(event))
        self.log_box.tag_bind("blue_tag", "<Enter>", lambda event: self.hypertext_enter(event))
        self.log_box.tag_bind("blue_tag", "<Leave>", lambda event: self.hypertext_leave(event))
        self.xlsx_check_btn = ttk.Button(bot, text="Examine Excel File", width=50, command=self.check_excel)
        self.eepr_check_btn = ttk.Button(bot, text="Examine EEPROM File", width=50, command=self.check_eeprom)
        self.xlsx_save_btn = ttk.Button(bot, text="Save", width=15, command=self.xlsx_save)
        self.eepr_save_btn = ttk.Button(bot, text="Save", width=15, command=self.eepr_save)
        ttk.Button(bot, text="Reset", width=15, command=self.reset).grid(row=1, column=1, pady=(0, 5), ipady=3, sticky=S)
        ttk.Button(bot, text="Exit", width=15, command=self.exit).grid(row=1, column=2, pady=(0, 5), ipady=3, sticky=S)
    
    # Text Box Functions
    def write_log(self, text=""):
        self.log_box.configure(state=NORMAL)
        self.log_box.insert(END, text + "\n")
        self.log_box.configure(state=DISABLED)

    def text_tag(self, word, tag_name):
        offset = f"+{len(word)}c"
        pos_start = self.log_box.search(word, END, stopindex=1.0, backwards=True)
        pos_end = pos_start + offset
        if pos_start:
            self.log_box.tag_add(tag_name, pos_start, pos_end)

    # Hypertext Event
    def hypertext_index(self, event):
        count = IntVar()
        pattern = r"[^ ]+"

        start = f"@{event.x}, {event.y} + 1c"
        index1 = self.log_box.search(pattern, start, backwards=True, regexp=True)
        self.log_box.search(pattern, index1, regexp=True, count=count)
        index2 = self.log_box.index(f"{index1} + {count.get()}c")

        return index1, index2

    def hypertext_click(self, event):
        index1, index2 = self.hypertext_index(event)
        filename = self.log_box.get(index1, index2)
        try:
            os.startfile(filename)
        except FileNotFoundError:
            self.write_log(text="[ERROR] No Such File or Directory: " + filename)

    def hypertext_enter(self, event):
        index1, index2 = self.hypertext_index(event)
        self.log_box.tag_add("underline", index1, index2)

    def hypertext_leave(self, event):
        index1, index2 = self.hypertext_index(event)
        self.log_box.tag_remove("underline", index1, index2)

    # Checkbutton Functions
    def check_all0(self):
        if self.d.xlsx_check_all.get() == 0:
            for i in range(5):
                self.d.xlsx_check[i].set(0)
        elif self.d.xlsx_check_all.get() == 1:
            for i in range(5):
                self.d.xlsx_check[i].set(1)

    def check_all1(self):
        if self.d.eepr_check_all.get() == 0:
            for i in range(5):
                self.d.eepr_check[i].set(0)
        elif self.d.eepr_check_all.get() == 1:
            for i in range(5):
                self.d.eepr_check[i].set(1)

    def check_one0(self):
        if (self.d.xlsx_check[0].get() == 0) | (self.d.xlsx_check[1].get() == 0) | (self.d.xlsx_check[2].get() == 0) | (self.d.xlsx_check[3].get() == 0):
            self.d.xlsx_check_all.set(0)
        if (self.d.xlsx_check[0].get() == 1) & (self.d.xlsx_check[1].get() == 1) & (self.d.xlsx_check[2].get() == 1) & (self.d.xlsx_check[3].get() == 1):
            self.d.xlsx_check_all.set(1)

    def check_one1(self):
        if (self.d.eepr_check[0].get() == 0) | (self.d.eepr_check[1].get() == 0) | (self.d.eepr_check[2].get() == 0) | (self.d.eepr_check[3].get() == 0) | (self.d.eepr_check[4].get() == 0):
            self.d.eepr_check_all.set(0)
        if (self.d.eepr_check[0].get() == 1) & (self.d.eepr_check[1].get() == 1) & (self.d.eepr_check[2].get() == 1) & (self.d.eepr_check[3].get() == 1) & (self.d.eepr_check[4].get() == 1):
            self.d.eepr_check_all.set(1)

    # Check Excel File
    def check_excel(self):
        for slave in self.xlsx_frm.pack_slaves()[1:]:
            slave.destroy()
        self.d.xlsxfilename.set(filedialog.askopenfilename(filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))))
        self.d.xlsx_text.set("")
        if self.d.xlsxfilename.get() == "":
            return
        self.write_log(text="Checking Excel File: " + self.d.xlsxfilename.get())
        self.text_tag(self.d.xlsxfilename.get(), "blue_tag")
        self.xlsx_imp.pack(side=TOP, fill=X)
        ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)
        self.after(100, self.read_xlsx)

    def read_xlsx(self):
        wb = load_workbook(filename=self.d.xlsxfilename.get())
        ws = wb["desc"]

        edid = []
        bits = []
        cell = []
        indc = []
        self.d.pages = []
        self.d.groups = []
        self.d.names_w_width = []
        self.d.names_wo_width = []
        self.d.widths = []
        self.d.addr_total = []
        self.d.defaults = []
        self.d.desc = []
        self.d.merge_check = []
        self.d.ronly_check = []

        row = 1
        while True:
            row += 1

            pg = "A" + str(row)
            grp = "B" + str(row)
            name = "C" + str(row)
            addr = "D" + str(row)
            default = "E" + str(row)
            desc = "G" + str(row)
            merge = "H" + str(row)
            ronly = "K" + str(row)

            if ws[pg].value == "end":
                break

            self.d.pages.append(ws[pg].value)
            self.d.groups.append(ws[grp].value)
            self.d.names_w_width.append(ws[name].value)
            self.d.addr_total.append(ws[addr].value)
            self.d.defaults.append(ws[default].value)
            self.d.desc.append(ws[desc].value)
            self.d.merge_check.append(ws[merge].value)
            self.d.ronly_check.append(ws[ronly].value)

        for item in self.d.names_w_width:
            if ("[" not in item) & ("]" not in item):
                raw_name = item
                width = "-"
            elif ("[" in item) & ("]" in item):
                raw_name, width = item.strip("]").split("[")
            else:
                raw_name = None
                width = None
            self.d.names_wo_width.append(raw_name)
            self.d.widths.append(width)

        if self.d.xlsx_check[0].get() == 1:
            edid.extend(self.xlsx_edid())
        if self.d.xlsx_check[1].get() == 1:
            bits.extend(self.xlsx_bits())
        if self.d.xlsx_check[2].get() == 1:
            cell.extend(self.xlsx_cell())
        if self.d.xlsx_check[3].get() == 1:
            indc.extend(self.xlsx_indc())

        self.xlsx_warnings(edid, bits, cell, indc)

        cnt = len(edid) + len(bits) + len(cell[0]) + len(cell[1]) + len(cell[2]) + len(cell[3]) + len(cell[4]) + len(cell[5]) + len(cell[6]) + len(indc[0]) + len(indc[1])
        self.write_log(text="Found " + str(cnt) + " Warnings.\n")
        self.text_tag(str(cnt), "red_tag")

    def xlsx_edid(self):
        edid = []
        for (i, grp) in list(enumerate(self.d.groups)):
            if i == 0:
                continue
            if (self.d.groups[i - 1] == "EDID") & (grp != "EDID"):
                edid.append("B" + str(i + 2))

        return edid

    def xlsx_bits(self):
        bits = []
        idxs = []

        for name in self.d.names_wo_width:
            widths = []
            if self.d.names_wo_width.count(name) == 1:
                continue
            elif self.d.names_wo_width.count(name) > 1:
                for idx in [x for (x, y) in list(enumerate(self.d.names_wo_width)) if y == name]:
                    if idx not in idxs:
                        idxs.append(idx)
                        widths.append((self.d.widths[idx], idx))
            widths = sorted(widths, key=lambda item: self.sort_key(item))

            for (i, tup) in list(enumerate(widths)):
                if i == 0:
                    continue
                curr_width = tup[0]
                prev_width = widths[i - 1][0]
                if (":" in curr_width) & (":" in prev_width):
                    diff = int(curr_width.split(":")[1]) - int(prev_width.split(":")[0])
                elif (":" in curr_width) & (":" not in prev_width):
                    diff = int(curr_width.split(":")[1]) - int(prev_width)
                elif (":" not in curr_width) & (":" in prev_width):
                    diff = int(curr_width) - int(prev_width.split(":")[0])
                elif (":" not in curr_width) & (":" not in prev_width):
                    diff = int(curr_width) - int(prev_width)

                if diff != 1:
                    bits.append(("C" + str(widths[i - 1][1] + 2), "C" + str(tup[1] + 2)))

        return bits

    def xlsx_cell(self):
        A = []
        B = []
        C = []
        D = []
        E = []
        H = []
        K = []

        for (a, pg) in list(enumerate(self.d.pages)):
            try:
                int(pg)
            except ValueError or TypeError:
                A.append("A" + str(a + 2))

        for (b, grp) in list(enumerate(self.d.groups)):
            if grp is None:
                B.append("B" + str(b + 2))
            elif grp.replace(" ", "") == "":
                B.append("B" + str(b + 2))
            elif len(grp.split()) != 1:
                B.append("B" + str(b + 2))

        for (c, name) in list(enumerate(self.d.names_wo_width)):
            if name is None:
                C.append("C" + str(c + 2))
            elif name.replace(" ", "") == "":
                C.append("C" + str(c + 2))
            elif len(name.split()) != 1:
                C.append("C" + str(c + 2))
            elif (("[" not in name) & ("]" in name)) | (("[" in name) & ("]" not in name)):
                C.append("C" + str(c + 2))
            elif ("[" in name) & ("]" in name):
                width = name.strip("]").split("[")[1]
                if ":" in width:
                    try:
                        if int(width.split(":")[0]) <= int(width.split(":")[1]):
                            C.append("C" + str(c + 2))
                    except ValueError:
                        C.append("C" + str(c + 2))
                elif ":" not in width:
                    try:
                        int(width)
                    except ValueError:
                        C.append("C" + str(c + 2))

        for (d, addr) in list(enumerate(self.d.addr_total)):
            if addr is None:
                D.append("D" + str(d + 2))
            elif len(addr.split()) != 1:
                D.append("D" + str(d + 2))
            elif ("(" not in addr) | (")" not in addr):
                D.append("D" + str(d + 2))
            else:
                num, bit = addr.strip(")").split("(")
                try:
                    int(num)
                except ValueError:
                    D.append("D" + str(d + 2))
                if ":" in bit:
                    try:
                        int(bit.split(":")[0])
                        int(bit.split(":")[1])
                    except ValueError:
                        if "D" + str(d + 2) not in D:
                            D.append("D" + str(d + 2))
                elif ":" not in bit:
                    try:
                        int(bit)
                    except ValueError:
                        if "D" + str(d + 2) not in D:
                            D.append("D" + str(d + 2))

        for (e, default) in list(enumerate(self.d.defaults)):
            if default is None:
                E.append("E" + str(e + 2))
            elif len(default.split()) != 1:
                E.append("E" + str(e + 2))
            elif (default != "R") & (default[0] != "b"):
                E.append("E" + str(e + 2))

        for (h, merge) in list(enumerate(self.d.merge_check)):
            if (merge is not None) & (merge != "O"):
                H.append("H" + str(h + 2))

        for (k, ronly) in list(enumerate(self.d.ronly_check)):
            if (ronly is not None) & (ronly != "O"):
                K.append("K" + str(k + 2))

        return [A, B, C, D, E, H, K]

    def xlsx_indc(self):
        m_indc = []
        r_indc = []

        for (h, merge) in list(enumerate(self.d.merge_check)):
            if (merge == "O") & (self.d.names_wo_width.count(self.d.names_wo_width[h]) == 1):
                m_indc.append("U H" + str(h + 2))
            if (merge is None) & (self.d.names_wo_width.count(self.d.names_wo_width[h]) > 1):
                m_indc.append("M H" + str(h + 2))

        for (k, ronly) in list(enumerate(self.d.ronly_check)):
            if (ronly == "O") & (self.d.defaults[k] != "R"):
                r_indc.append("U K" + str(k + 2))
            elif (ronly is None) & (self.d.defaults[k] == "R"):
                r_indc.append("M K" + str(k + 2))

        return [m_indc, r_indc]

    def xlsx_warnings(self, edid, bits, cell, indc):
        length = 0
        length += len(edid)
        for item in edid:
            ttk.Label(self.xlsx_frm, text="Group Below EDID: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Group Below EDID: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(bits)
        for item in bits:
            ttk.Label(self.xlsx_frm, text="Bits Not in Sequence: " + item[0] + " & " + item[1]).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Bits Not in Sequence: " + item[0] + " & " + item[1] + "\n")
        if (length != 0) & (len(cell[0]) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[0])
        for item in cell[0]:
            ttk.Label(self.xlsx_frm, text="Wrong Page Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Page Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[1])
        for item in cell[1]:
            ttk.Label(self.xlsx_frm, text="Wrong Group Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Group Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[2])
        for item in cell[2]:
            ttk.Label(self.xlsx_frm, text="Wrong Name Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Name Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[3])
        for item in cell[3]:
            ttk.Label(self.xlsx_frm, text="Wrong Address Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Address Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[4])
        for item in cell[4]:
            ttk.Label(self.xlsx_frm, text="Wrong Default Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Default Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[5])
        for item in cell[5]:
            ttk.Label(self.xlsx_frm, text="Wrong Merge Indicator Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Merge Indicator Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(cell[6])
        for item in cell[6]:
            ttk.Label(self.xlsx_frm, text="Wrong Read-only Indicator Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.xlsx_text.set(self.d.xlsx_text.get() + "Wrong Read-only Indicator Format: " + item + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        length += len(indc[0])
        for item in sorted(indc[0]):
            if item.split()[0] == "M":
                ttk.Label(self.xlsx_frm, text="Missing Merge Indicator: " + item.split()[1]).pack(side=TOP, fill=X, pady=(5, 0))
                self.d.xlsx_text.set(self.d.xlsx_text.get() + "Missing Merge Indicator: " + item.split()[1] + "\n")
            elif item.split()[0] == "U":
                ttk.Label(self.xlsx_frm, text="Unnecessary Merge Indicator: " + item.split()[1]).pack(side=TOP, fill=X, pady=(5, 0))
                self.d.xlsx_text.set(self.d.xlsx_text.get() + "Unnecessary Merge Indicator: " + item.split()[1] + "\n")
        if (length != 0) & (len(bits) != 0):
            ttk.Label(self.xlsx_frm, text="").pack(side=TOP, fill=X)

        for item in sorted(indc[1]):
            if item.split()[0] == "M":
                ttk.Label(self.xlsx_frm, text="Missing Read-only Indicator: " + item.split()[1]).pack(side=TOP, fill=X, pady=(5, 0))
                self.d.xlsx_text.set(self.d.xlsx_text.get() + "Missing Read-only Indicator: " + item.split()[1] + "\n")
            elif item.split()[0] == "U":
                ttk.Label(self.xlsx_frm, text="Unnecessary Read-only Indicator: " + item.split()[1]).pack(side=TOP, fill=X, pady=(5, 0))
                self.d.xlsx_text.set(self.d.xlsx_text.get() + "Unnecessary Read-only Indicator: " + item.split()[1] + "\n")

    # Check EEPROM File
    def check_eeprom(self):
        for slave in self.eepr_frm.pack_slaves()[1:]:
            slave.destroy()
        self.d.eeprfilename.set(filedialog.askopenfilename(filetypes=(("Text Files", "*.txt"), ("All Files", "*.*"))))
        self.d.eepr_text.set("")
        if self.d.eeprfilename.get() == "":
            return
        self.write_log(text="Checking EEPROM File: " + self.d.eeprfilename.get())
        self.text_tag(self.d.eeprfilename.get(), "blue_tag")
        self.eepr_imp.pack(side=TOP, fill=X)
        ttk.Label(self.eepr_frm, text="").pack(side=TOP, fill=X)
        self.after(100, self.read_eepr)

    def read_eepr(self):
        data = []
        main = []
        page = []
        csum = []
        self.d.byte_list = []
        self.d.mod_pg = []
        self.d.dict = {}

        txt = open(self.d.eeprfilename.get(), "r")
        byte_text = txt.readlines()
        txt.close()
        self.d.byte_list = list(map(str.strip, byte_text))

        if len(self.d.byte_list) == 0:
            return

        k = 0
        while k < len(self.d.byte_list):
            self.d.mod_pg.append((self.d.byte_list[k], k + 1))
            n = int(self.d.byte_list[k + 1], 16) + 4
            self.d.dict[self.d.byte_list[k]] = self.d.byte_list[k + 1:k + n]
            k += n

        if self.d.eepr_check[0].get() == 1:
            data.extend(self.eepr_data())
        if self.d.eepr_check[1].get() == 1:
            main.extend(self.eepr_main())
        if self.d.eepr_check[2].get() == 1:
            page.extend(self.eepr_page())
        if self.d.eepr_check[3].get() == 1:
            csum.extend(self.eepr_csum())

        self.eepr_warnings(data, main, page, csum)

        cnt = len(data) + len(main) + len(page) + len(csum)
        self.write_log(text="Found " + str(cnt) + " Warnings.\n")
        self.text_tag(str(cnt), "red_tag")

    def eepr_data(self):
        data = []
        for (i, byte) in list(enumerate(self.d.byte_list)):
            if (re.compile(r"[^0-9A-F]").search(byte) is not None) | (len(byte) != 2):
                data.append("Line " + str(i + 1))

        return data

    def eepr_main(self):
        if self.d.byte_list[0] != "4A":
            return [True]
        else:
            return []

    def eepr_page(self):
        page = []
        for (pg, i) in self.d.mod_pg:
            try:
                if (int(pg, 16) < 64) | (int(pg, 16) > 237):
                    page.append("Line " + str(i))
            except ValueError:
                page.append("Line " + str(i))

        return page

    def eepr_csum(self):
        csum = []
        for (pg, i) in self.d.mod_pg:
            sum = 0
            try:
                sum += int(pg, 16)
            except ValueError:
                pass
            bytes = self.d.dict[pg]
            for b in bytes:
                try:
                    sum += int(b, 16)
                except ValueError:
                    pass
            if int("FE", 16) - (sum % 256) != 0:
                csum.append((pg, "Line " + str(i + len(bytes))))

        return csum

    def eepr_warnings(self, data, main, page, csum):
        length = 0
        length += len(data)
        for item in data:
            ttk.Label(self.eepr_frm, text="Wrong Data Format: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.eepr_text.set(self.d.eepr_text.get() + "Wrong Data Format: " + item + "\n")
        if (length != 0) & (len(main) != 0):
            ttk.Label(self.eepr_frm, text="").pack(side=TOP, fill=X)

        length += len(main)
        if True in main:
            ttk.Label(self.eepr_frm, text="Initial Page Not Set To Main Page").pack(side=TOP, fill=X, pady=(5, 0))
            self.d.eepr_text.set(self.d.eepr_text.get() + "Initial Page Not Set To Main Page\n")
        if (length != 0) & (len(page) != 0):
            ttk.Label(self.eepr_frm, text="").pack(side=TOP, fill=X)

        length += len(page)
        for item in page:
            ttk.Label(self.eepr_frm, text="Invalid Page: " + item).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.eepr_text.set(self.d.eepr_text.get() + "Invalid Page: " + item + "\n")
        if (length != 0) & (len(csum) != 0):
            ttk.Label(self.eepr_frm, text="").pack(side=TOP, fill=X)

        for item in csum:
            ttk.Label(self.eepr_frm, text="Checksum for Page " + item[0] + " Incorrect: " + item[1]).pack(side=TOP, fill=X, pady=(5, 0))
            self.d.eepr_text.set(self.d.eepr_text.get() + "Checksum for Page " + item[0] + " Incorrect: " + item[1] + "\n")

    # Button Functions
    def xlsx_save(self):
        if self.d.xlsx_text.get() == "":
            self.write_log(text="Nothing to Save.")
            return

        txtfilename = filedialog.asksaveasfilename(filetypes=(("Text Files", "*.txt"), ("All Files", "*.*")))
        if not txtfilename.lower().endswith(".txt"):
            txtfilename += ".txt"
        txt = open(txtfilename, "w")
        txt.write(self.d.xlsxfilename.get() + "\n\n" + self.d.xlsx_text.get())
        txt.close()

        self.write_log(text="Warnings for " + self.d.xlsxfilename.get() + " Saved.\n")
        self.text_tag(self.d.xlsxfilename.get(), "blue_tag")

    def eepr_save(self):
        if self.d.eepr_text.get() == "":
            self.write_log(text="Nothing to Save.")
            return

        txtfilename = filedialog.asksaveasfilename(filetypes=(("Text Files", "*.txt"), ("All Files", "*.*")))
        if not txtfilename.lower().endswith(".txt"):
            txtfilename += ".txt"
        txt = open(txtfilename, "w")
        txt.write(self.d.eeprfilename.get() + "\n\n" + self.d.eepr_text.get())
        txt.close()

        self.write_log(text="Warnings for " + self.d.eeprfilename.get() + " Saved.\n")
        self.text_tag(self.d.eeprfilename.get(), "blue_tag")

    def reset(self):
        self.log_box.configure(state=NORMAL)
        self.log_box.delete("1.0", END)
        self.log_box.configure(state=DISABLED)

        self.d.xlsxfilename.set("")
        self.d.eeprfilename.set("")
        self.d.xlsx_text.set("")
        self.d.eepr_text.set("")

        self.d.xlsx_check_all.set(1)
        self.d.eepr_check_all.set(1)
        for i in range(5):
            self.d.xlsx_check[i].set(1)
            self.d.eepr_check[i].set(1)

        for slave in self.xlsx_frm.pack_slaves()[1:]:
            slave.destroy()
        for slave in self.eepr_frm.pack_slaves()[1:]:
            slave.destroy()

        self.d.pages = []
        self.d.groups = []
        self.d.names_w_width = []
        self.d.names_wo_width = []
        self.d.widths = []
        self.d.addr_total = []
        self.d.defaults = []
        self.d.desc = []
        self.d.merge_check = []
        self.d.ronly_check = []
        self.d.byte_list = []
        self.d.mod_pg = []
        self.d.dict = {}

    def exit(self):
        self.destroy()

    # etc
    def switch_tab(self, event):
        if self.d.tab.get() == "excel":
            self.xlsx_check_btn.grid_forget()
            self.xlsx_save_btn.grid_forget()
            self.eepr_check_btn.grid(row=0, column=0, columnspan=3, pady=(0, 5), ipady=3, sticky=S)
            self.eepr_save_btn.grid(row=1, column=0, pady=(0, 5), ipady=3, sticky=S)
            self.d.tab.set("eeprom")
        elif self.d.tab.get() == "eeprom":
            self.eepr_check_btn.grid_forget()
            self.eepr_save_btn.grid_forget()
            self.xlsx_check_btn.grid(row=0, column=0, columnspan=3, pady=(0, 5), ipady=3, sticky=S)
            self.xlsx_save_btn.grid(row=1, column=0, pady=(0, 5), ipady=3, sticky=S)
            self.d.tab.set("excel")

    def sort_key(self, tup):
        if ":" in tup[0]:
            return int(tup[0].split(":")[0])
        elif ":" not in tup[0]:
            return int(tup[0])

    def xlsx_resize(self, event):
        self.xlsx_cv.configure(scrollregion=self.xlsx_cv.bbox(ALL))

    def eepr_resize(self, event):
        self.eepr_cv.configure(scrollregion=self.eepr_cv.bbox(ALL))


class Data:
    def __init__(self):
        self.pages = []
        self.groups = []
        self.names_w_width = []
        self.names_wo_width = []
        self.widths = []
        self.addr_total = []
        self.defaults = []
        self.desc = []
        self.merge_check = []
        self.ronly_check = []
        self.byte_list = []
        self.mod_pg = []
        self.dict = {}
        self.tab = StringVar(None, "eeprom")
        self.xlsxfilename = StringVar()
        self.eeprfilename = StringVar()
        self.xlsx_text = StringVar()
        self.eepr_text = StringVar()
        self.xlsx_check_all = IntVar(None, 1)
        self.xlsx_check = (IntVar(None, 1), IntVar(None, 1), IntVar(None, 1), IntVar(None, 1), IntVar(None, 1))
        self.eepr_check_all = IntVar(None, 1)
        self.eepr_check = (IntVar(None, 1), IntVar(None, 1), IntVar(None, 1), IntVar(None, 1), IntVar(None, 1))
        self.logo = b"iVBORw0KGgoAAAANSUhEUgAAAwkAAALQCAYAAADb3TzfAADHE0lEQVR42uzdeZxddX3/8df5fr/33tmXTDKZSQJWa1tbVCJKUaqiVqvVulSruIAsglkEFRGyQVhCFnYRnCSAgBVRtNa6L7Wt9qfWBSnYWtv+/HUBsiezb/fe8/2e3x/3TgjIkmS+N5nl/fRBHzW0Azkzuee8z/ezgIiIiIiIzCxvxwLJAb+S6KKIiIiIiMxGt2M4G7v/v9/MWdzILfv/nsKCiIiIiMiskfDLA8LBjZzMJr7LtWTcTMYGPgzAVtzBfTEREREREZm+lmDZigfgWo4lYxWecwDHKCkOqCPgeSmr+RmnYrm3+n+vkCAiIiIiMoPcjsEDSwicSx2/x3JKrMDSyQiQ4TFYPIF6DPArcrwYGOYiMhKyJ/vSVldXRERERGQayUiYi2MZnq+RcR1vpoV7SHgvRRoZJ8WQkGAAMCSU8dTTyRgLWMOXaMXxbcKT/SN0kiAiIiIiMj0kPIBhcbVU6GpeQMLlwJsoA+N4wGCe5Bk/IaUFR5GzWcWdbMaxjFQhQURERERkOjoXy23VcLCO+RS4GM9yEuoYqZ4ImOrJwZMJBByQYwTLi1nBv3Ejhgt+80RBIUFEREREZKq6DcMfknA8nndiOYlzGOcSHIsYAXy17+BgeQINGFLuo46X8WPKfJHw+P4E9SSIiIiIiEw9CbfjOBfPFjI28Sd0cjewlBItjJGSHNB3cLAq/QkpjSyiRAuf4Bu0/WZ/gk4SRERERESmkvdjubVaWnQ9z8FzGRnvxANjT9N3cHAyMgJNWDLexgr+mi04lj7an6CQICIiIiIyFdyGIVAZaXoZjTRwIYELMbQwTAZkT9t3cLACgRzg6MXzIi7lf9mMYVnlREEhQURERETk6Er4JYbjqqcHG3gjsB7H8xjl0PsODj4oeBqxlPkH8ryaBuD8Shgx+p6IiIiIiBwlS7BAxnF4ruPZXMtncXyFjOcxRIonq0lAADBYRvDkeSVFLuN8MjZXDhF0kiAiIiIicqTdgeEHJNyB50bylLmAEqvJ0cLwQY40jSHgyWPJSCnwTFbwCJdhFBJERERERI6chNuxnFNtEr6e1zLGJnIsZozalRY9VUCwjAHvoZkv44BlBIUEEREREZEjYSmWLfunFv0WKVcA7yXl6bcl1yIgFLA49hF4G6v5PpUqowy0J0FEREREpLYyEv4Lyx14PkiOt/JBUj5NwksYJiOt9h0kRywgpNTjsDyE5w2s4cfcjuMrj+5KcPquiYiIiIjUyANYEjzguZZXkLKRhBdTAlJSDI4j2SccSGnE4fkVgTezhv/LVtz+8qcqlRuJiIiIiMR2F4Yzq7sNrmMBGZeTci4ZMEYK2CNWWnRgQGjGUeanZLyFS9jB2dUTjsdRSBARERERiSUj4T4MJ+KBhBtYyiiXkWM+I2RkEReiHWpAaMFR5tvs5R3cwOCTBQRQuZGIiIiISBynHVBadAMvocw1wEvxQLE6tSg54i/pMwKeOTjKfI5+zuQGityK4f1PHBBAJwkiIiIiIpNzG4YALCFwFa041pDxYTJyjB3hqUWPDwgZgTYsRbawiuVAxs0Yzn+0SVkhQUREREQkpp9hq6VFsIE3AleT4/cZArIjuPPg8QIZloxmDONsYA1ruAnDB8lIKmNOFRJERERERGLqwbCs+sB9LYuA9YT9Ow/Sajg4Os/aGYGEhEYSAheykht4AMtiAjx9QFBIEBERERE5NAm/xHBc9fRgPediuBJDF8PVEp6j0Zg8IRBwGPIELOdwMXfSg2M5/mADgkKCiIiIiMjBOhfLbdVwsInn4bgWeC2jQPkolhZN8AQKGHKMEXgXq/kym3Ese+wOBIUEEREREZHJykj4BJbzSLmRPGVWkLIKQz0jR7Ux+VEBTx0Wx17KvJ1L+B5bcCw99IAAGoEqIiIiIvLkfrZ/rGnKJl7FGNdiOYES4KfA6QHVf48GLPAwRd7CWu5nBTmWUj7cL6mQICIiIiLyeLdjKAEn4rmWucDlpCwnI2EYj8FMiYAQSGnCEfgVCW9mLf+X23Cce/gBAVRuJCIiIiLy+IDgOKdapnMNp5KyEcszj+rG5KcKCJ4fU+QtXMEuXoXj7w+vxEghQURERETk8e7CcCYZkHE9zyZlE4a3MQaUSTFTqAonkNKCo8y3sJzKCgb5BZbnP/kW5UOhciMRERERkV9iq2NNE67ng5RYi6GDwf1jTafKc3NGIDAHR4l76ecMrqFIDyZWQACdJIiIiIjIbHY7hnOqQeA6nkvGjVhezRCQTpHG5AMDQkagHcsQW1jLMgC2YFha/T1EYvWTISIiIiKzUMIvsbyOwDuwnM5H8XyajOcwSEogmTK9BwCBDAi0YCmxnku4kIyEVgznxw0IlYsjIiIiIjKb3Ibh3OqD9VUcT46PYXgFI0ydsaaPDwiWjEYMnhWs5Bq+j+XlBJKD36J8KNSTICIiIiKzRcIvMRyH53ocZS4mYw2eBoZJATvlAkJGhgUaMKQsZxWb6cFxCh5qExCAKXSEIiIiIiJSK+digYzj8GzkRZT4HgnrKdLAGB6DO+pbk38zIAQsCXUEypzBKjazGcfy2gaESpoSEREREZm5Ev4Nwx/geSd5XsRKSqwB8oztPz2Yes/EnkAOQ4Eigfewmi+yBcfSye9AUEgQERERkdnr/VhurY4F3cBLsNyI4SSGgGwK9h5MCHjyWHIMUeYvuJTvsBnHsiMTEBQSRERERGQmSrgDy9mkrKDAPC6lzAoyHGOk1XAwNZ+DA54Clhx7KfHnXMoPjuQJgkKCiIiIiMw852K5bf/pwcsw3IjhhQxTqfE3U7gnN+Cpw5KwA3gjq/n50QgICgkiIiIiMlM82nvwQRo4lrWUuIgMwzi+Gg6m7rNvwNOAJeO/KPIGruDf6cGx/MgHBIUEEREREZn+DtyafA2vxPMxHM9niIyMbEqfHlQCQkojDvhXRngj6/gf7sdyQvVERCFBREREROQQnmVvx3IOKddTh+cKUi4iI6meHtgp/zuYCAie+8h4E5ewgwewLD56AQG0J0FEREREpqPbq8+x55CygZdQ5IfAxYwB44RpExBacHi+T8ZruIQdvPvoB4RK+hIRERERmU4mmnmvxxBYTZm1JOQYnSanBwAZZZrJUeYbDPEONjJChiGplk0dZU4/ZSIiIiIyLdyOoQwsJeUqjqfIx7G8nHEgELDTJCAEUuaQo8QX6Od0rqHIzVMnIIBOEkRERERkOjhwmdgmPkxgPQkNjE7xvQe/+fSd0opjmDu5lPcBGVswLJ06AQGYJmlLRERERGanjIQODOfhuZ5n88fcTZ7zGCVHCY/BTZOAkJHgacVR5mZWs5SMhFYM50+tgAAqNxIRERGRqeonWBI8kHEN76PENSTMoQ8PmGnTfxDISAg04yizkYtZzduwQOAjUy8ggMqNRERERGSqyUj4KYaT8KxnIXluJOHtjADpNGpOrvxeMgyBZixjrOESNnAqlnsJQDZV/7UVEkRERERk6vgJlpOqI0A3cSqBG7AsYARfXYs2fZ5fMwIJCY0kBD7ESj7Oj7GcRCCZugEBVG4kIiIiIlNDwr9geB6eK+mgnmvIOJsSMFY9PZhOr7cDAYehjoyUs1nNndyC48XV5usp/80QERERETma3o/l1urpwVW8AcNNWH6boeqbeDPNnlk9gQKGPEVS3ssaPj8VtigfCp0kiIiIiMjRcwuO80i5jEYa2UDGBykDY6TYafis+mhAGMbzTtbwdXpwLJ4eJwgTdJIgIiIiIkdeD4ZlZCRkbOCPMPRgeT6D1Wk/BjPtfk8BTwGLYQDPm7iUf3zMfgeFBBERERGRJ/EgluOrpTfrWI3lcjJyjJNW9x5MPxMBwbGHEm9iLT9mC46l0y8ggMqNRERERORIyUh4D4bj8dzEMymzGcNrGaBSpmOncUCow5Kwi8AbWMvPp3NAUEgQERERkSPjtP2L0TybeA+j3AB0MkqKwWJrWF4UoGZffSIgWB7B83rW8C9snt4BAVRuJCIiIiK1ft68Bct5pNxIGyWux3A2o9R+MVqoPO3m6vKkpTJZlsX++p4GLBn/TcYbWMOvpvsJwgSjn1sRERERqYlbMGTAeaRcy8sY54fA2QwSSMlqGhBSsDlLY0cbwQcyHz0gpNRjyfg14/wJa/gVPTMjIFSSnYiIiIhIbFtxLCHljVheyBoKXErA1bw5OQPKUJjTRF1rI8O79uFLadxX44GURhyBX5LyBtbyv9yP5YTpswdBIUFEREREjpyMhNMx3I1nA78HbCbPKxkEMkJNR5umgIHmBXPJ5wv0PrSDLA1xu3ADKQ04Mn7BOG/gSh6ZbovSDoYal0VEREQkjp8d0Jx8De/FcwMJHQxUm5OTGgWErBIQXEOe9mO6SFPPvv/eVnkdHjsgNOHw3EfCG7iS3bx35gUEhQQRERERiSHhNiwnknIlHdRxAwnvpcREc3LtnjlDJSQ0zG+lZV4HxZEx+v53RyUgxIwknpRmHJ4fYXkzK9j7mH0PM+4bKiIiIiJyuO7AcHZ1S/K1/DFltmB5NsN4KnuTa/e8WQZb72ieP5f65iZGB4YYeGRX/IAwcYKQ8o/08maup58eDMurv+8ZSCcJIiIiInJ4enCcTco7sPwhl1FmDQHDWI2bk6unB/VzW2jqaMc4y2jfAAPb99QmIFROEP4Wx1u5nmG2YFg6cwOCQoKIiIiIHI6EX2I4jpQN/D6GLcDLGSOrNifX7hkzBVtwNHfNJV9fAAzjQ0O1CwgtOEp8A8fbWcEom2d+QKh8g0VEREREDtb7sdxarcNfz+kkfBxoOyKnBwHq2xpp6pxLpVvZUBobYeCRGgaEMn/Dw7yDzZTZjGHZzA8ICgkiIiIicvAmdh98kAYWcT0JS4/I5uQUcIaW+R3UNTfiU4+xltJoDQNCM44yn+e/OY3bKXMzhvNnR0AAlRuJiIiIyNPJSLiZhCWkrOcFWG4j4YUM4ckwNQsIAfCQb66jpWsuiXX4UhmTy1EcHWGwVgFhDo5e7qbAGdxOICMhmT0BQSFBRERERJ7aaft3H2RsYimB6/A0MlItL6pVXUp1MVpjVxuNbW1kIRDKRyggFPkk6ziHrPpPmGUBAVRuJCIiIiJPZqK86FLm0MzNWN7NEOBrXF5UAtvoaJnfiavPE0opZGByjtLwSK2alD1zsJToYQUf4GoMF5ORkM3Gb71CgoiIiIg81m0YUmAZgQ38EXAblt9nZH95UW2eISeak+c00zR3DmQZIVRe4hvnKA6PMBg7IATAktKKo8jHWMkFXIfho2QwOwOCQoKIiIiIPNYDWBZXpxddw4V41hMoMF7j6UUp4BJa58+l0NyIT9P9j+g1CwhUR7a2YRnnalazkh9g+SPCbD1BmKCeBBERERGpuAnHYlLW002OT+D4c4aBQMDW6Lmx2pyca6mndX5HpTm5nO7/28Y5SsPDDG7fGz8gJPjqCcI6VrOW+7C8aGJV2+ymkwQRERGR2a6yIKxSf7+B1wJbsTyDYTzUsLxoojm5s53Gtja89xDCYwJCzU4QEnx1D8KVrOCy6gmKAoJCgoiIiIjQg2M5lVf3G7ichLV4Eoo1bk4ug22oNicX8oQ0fczfrmEPQoatBoQiV7CKy3kFju9VJziJQoKIiIjIrH4O/CWG4/Cs47eoZwuG19JPRlZtT66FiebkjmaaOh7bnPyYgDA0zOCOvbULCGUuZwVX8Essx+kE4fGMLoGIiIjILHM7Bsg4Ds9V/Dk5/omU19KHJyGpWUBIwThD6zGdNM/rIHh/5AJCpcQo0IJjjLWs4ApegVNAeLIEKSIiIiKzx8T0ovPJsZBNGD5CESjVsLwoqwSEfGs9LZ2V5uTHlxfVNCCEakBow1JkLatYxy+wPF8B4cloupGIiIjI7JBwE5bFpGziWWTcieXlDFa3CdcqIFSbk5u659DQ2kwWsqcOCNv3VsJB7BOENixjXMIa1nMLjuerB+Gpf1hEREREZGbrwbBs//SiN5JwKwldjNR+94Gtc7R0zcMVCk8YDh4TEGrZg1DkElaxXlOMFBJERERE5MDlaBu5AsNaStS2vKjanFzX1khz59wnbE7+jYCwfS/YqE+njy5KG2UNl7CBW3CcpxOEg6FyIxEREZGZqmf/crSF5LkVx+vpPQLlRRZaFsylrqUZn5af9JH8N0qMYgWEQIbD04ijyGouYWM1LKX6oTg4VpdAREREZIbJSOjAcB6eDbyGhC8DL2KwenqQ1KiapAyuIUfbMd3k6uuftLzoNwKCJX6TcjOOMqtYxSZuwfEmgn4wDp5OEkRERERmktOwJNWSmk2sJGM9KYax2pcX1c9tpqmjA7LwtAGhNDxS6UGo3QnCStZwNfdjOUEnCIdKPQkiIiIiM8VtOM4lZSPzsGzB8VYGyAg1XI6WArmEls651DU34stP/TxunKM4OsLgI3tqM+a0BUvKClZyTTUgqEn5MOgkQURERGS6y0g4HcO5pFzNyQTuJON36cNjqv+phTK45gJtnXNJcrmDCgil0dHaBARTDQglLmYV19KD0wmCQoKIiIjI7PRoeZHnaj4AXEegjtEajjetlhc1zGuhsWMOhKcuL9ofEMZGGdi2u3YBocxFrOI6NuNYVp3oJIdF5UYiIiIi01XlYTjlSpqp52YMZzAEeEJNy4ucobV7LoXGhqc9PZgICOWxcfof2Vn9hYgBwRJoxJLyUVZx/f5rIgoJIiIiIrPuGe6XGI7Ds54XYLmDhMUM4wGDqcEzXlYJCPmWOlrmzyWx7mlPD/YHhPFx+h+uUUBoxlLkQlZzwwEnCOpBmCSjSyAiIiIyjdyGATKOw7ORszH8I57FDJNisDUJCNXyosbONtoWdJEk5hACwlj8gJCRARnNWMpcoIAQn3oSRERERKaLie3Ja6inlRtIWMoI4PE16z9IwRQsLV3zyDXU4UsHV8ljnKNcLDLwyK64AeHAEqMyF7CSjykgKCSIiIiIzE4T25PX8rsU+EsSTmJwf3lR/P0HE+VFrQ20zO8gISEcQkBIi0X6H95R+TpxA0JGI5YSH2YNN9GjHoRa0MZlERERkanssduT34jjyxh+t1pe5GqyPbm6WaBx/hxaOjsIWUYWDm5hcSUglOh7eGctAkLlBMHzIVbzcR7A8kZNMaoFnSSIiIiITFWP3Z68moT1lIBSjcuL8oaWrk5y9XX4cvmg/1+NM/hSib5HdkCWxR9zOhEQVvLx/ScropAgIiIiMmtsxbGElKtopcCtON7BAKFaYFSz8qJcSx2t8+eSmINrTn40IDh8qVQpMfJZvHqVA/cgVKYYfVxjThUSRERERGabhI9jWULK1Twfz90Enkcvvjq9KL6J5WidrTTNaScLgeDDQf+/G2fwaZnebTvjBoRKsdXEJuUVrOYG9SAcqR9CEREREZkatmJ4PxkJGZt4JxlbyGhlrIbbk1Mgl9DaNY98YwOhfGjP38YYfPD0P7yDUArxXkFXehA8LTiKXMIq1u+f7iQ1pz0JIiIiIlPBmViWEEiAdWzC8llKtDJKqFlAKINrKDDnmIXkGw4zIPhA/8M74wYEDggIY1zOKtZzC47FBP2gHBk6SRARERE52iZq7G+ki5RPYnk9/XiyGm1PrpYX1Xe00DR3DoRACIf2/G2MIcsCvQ/vIBR93CL2hJQWHJ71XMQl/ALL8ydmLolCgoiIiMhMfxb7JYbj8GziZAJ/ieG3GalxeZGF1q65FJqb8OVDL+83xuCzQP8jOwjjkQNCIGUOjjGuZjUreQWO72lR2pGmciMRERGRo+E2DJBxHJ6NvJ+M7xL4bUZqON60DK4+x5xjF5JvOryAkFRPEAYermFAKHIDq1nJT7EKCEeHphuJiIiIHGkTDbgXU6CDG3EsYwhICdgajDfdX17USNPcDsg4pPGmE4wxZGT0PbITX6xBQGjHUeZmVnIhp2I5USVGR4tOEkRERESOpM04FuO5kmfSzncwLGMAjyeryYDTahZo6e6gubOTcBj9B9WEAGT0P7ITP57W5gShxGYu5oO8E8u9BBIFhKNFPQkiIiIiR0JGwukY7sazgdcCd2LornX/ga1ztHR34vL5wzo9qOQDA0lC3yM7SUeLtSoxuo2VvJ/1GFZXx8DKUaNyIxEREZFauwlDQgA8V3MRgU2kGMZq1H9Q3Z5cN6eB5nnzIMsOOyCQJJBA//YaBoRx7mIV7+cqBYSpQicJIiIiIrU00X+wlkaa2ILjNAbICNUBp7FVq/gb57XROKcdn6aQHeYzd5JgraV/+05KQ+PxA0IzjiKf4VJO5yoS1iggKCSIiIiIzHQT+w828mwC9+A4kWFSDLYmz2EpmLyhpauTXH3d4Z8eHBgQduymNDAKuRoEhD7uJc+7AbhcAWEqUeOyiIiISHwJ52JZRsom/oSMH5DsDwiuJgFhYnvysQsmHxAA6yyDu2oUEFpw9PNFPKdxORk5UECYaj/AIiIiIhLPJzG8j8r4oKs5n8DHSDGU8NUThLgywEP9nGaa5rVD4PCmFx0YEHI5hnbvZmzfSG0CwgBf5j94O58nZT0JlxD0gzO1qHFZREREJJazsLwPz5fI8R98HMdSBsnICDUJCClgoKm7g4bWFny5POkvaXM5hvfsix8QMsq0kKPMN9jLqXyelA0KCFOV1SUQERERiWAzjqvwXMlC9vLXWN7OIJ6MpCYNyimYgqV1YReFxoZJlxdVAoJjeF8fo3sG4wYET0oTOcp8m0Hexk2M04phhQLCVKVyIxEREZHJPk/9DMOJeK7mZDyfwfBbNd1/UIZcSx2tXfNISCZdXjQREEb6+hnZ1U/UtupAShOOlL9nhDexgRG2YFiqgDCVqXFZRERE5HBtxZABJ+LZyJkE/o5Qw4AQgBQa5rXSvqArWkAwOcdo/yAjOyMHBF8NCJ7vMcZbFBCmU/IVERERkUN3FpY78UDCOq6mjosYATyhVuVFiYXmrrnUNzeRltMoX9Y4R3FwiMEd+2pzglDmRzj+lJUMKiAoJIiIiIjMXFtxLCFlAx047iTHG+nDV9ej1WT/ga1ztHV3YvL5KP0H+wPC8AiD2/bUJiB4fkorr+MD9LEVwxIFhOlC041EREREDsXmakBYz/HAPWT8Ab3V8qLY8SCrBIR8WwPN8+dhsixqQCiNjjK4fU+lAD1eQPA04vD8gpQ/4wP0cTaWJXj98Ewfmm4kIiIicnAP7AkdGM7Ds4G3kfAlMo5htIb9Bxk0ds6hpbODLASyEOdFvHGO8tg4A9t2VcKBifbv7KnHkvGfpLyOtew4oCxLphGdJIiIiIg8nZswJNXH9k1cQsI6xoEUX5OAkAIuoW3BPHL1DVH2HzwmIIwX6X9kZ/UXIgaEOizwv6S8nrU8zANYFisgKCSIiIiIzDQTD7qX00wjW7G8i0ECnqRWC9JcfY6W7k5sLhetvGgiIKTFEgOP7IgfEApYEnaS8QbW8v+4XwFBIUFERERkJurBsZiUjTybwD0ETmSQFION3qBc7T+oa2+guXMeROw/ADDGkJZK9D2yo/LPihUQPIE6LI59BP6MNfySHhwnkOoHSCFBREREZCZJ2IxlGSkbeCUZnyGhm+Ea9h8EaJzfRuOcdrxPIWRRA4IPgYFtOyHN4j0BZgQKGAxDlHkTl/JzNuNYpoAw3alxWUREROSxD74Jc0n4AJ6NnIXhs6S0Uaxh/4FJaF3USV1bM6GcVt70RwwIWRboe2QHoRTiBQRPwGHIM0bKW7iU77NFAWGm0EmCiIiIyIQDG5Q3sh7HakaAlICtTf+BLThaF8zH5nOEUtzna2MMGRl9j+wkFH3ME4QMR0KBMp53cCnfZQuOpQoICgkiIiIiM8lEg/L1NBG4Dcs7GajuT7Y12KBchnxLPa1dnWDi9h9UEwIkGX0P78CPp/Ge+gIZjowCCZ7TWMPXFBAUEkRERERmnkcblH+LEp/F8mL6qv0HseNBtf+gfl4LTR3thJBBGnkRcZJgDPRt240fixwQLIF6LGXOYg2fV0BQSBARERGZaQ5sUP4j4LPAMQzVqEE5BQy0LJhLXUsTvlyDZ+skwVpL//ZdpMPFmE97GQmBRiwpy1nDXQoIM5cal0VERGR2OrBB+RreQ8YX8HQwXrsGZZM3tC7qIt/QEL+8aOLhLmcZ3L2HYv8Y5CJ90UCGwdOCo8zFrOKm6hQj7UGYoXSSICIiIrPPJw5oUN7A5SRcRhH2dyDEVoZcY4HWhfNIjKthQMgxtHsv472j8QJCRSUgjHMZq7n2gICQ6YdpZkp0CURERGRWOQvLnXjWUE8bt2F5D30EIIm+IC0AHuo7mmma10EIAUKoyW/L5hzDe/sY3TMQNyAEUubgGOUa1rCCHhzLFRAUEkRERERmiolFX+s5BriXPC+pblCuzYK0DJrmt9PQ1oYvl2v22zI5x2j/ICM7euMGhIwy7eQocQsrOJ8fYTl54ncmM5nKjURERGR2eBDL8aRs4CXA5zAcW7OAkAIuoW3BfHL1dbUNCM5RHBhiZGdv3Ce7yglCjnHuYBXncx2GlyggzBZGl0BERERmtIyEd2M5Hs863o3huwSOZaRGDcolcPU55vzWQnL1dTXrP5gICKXREQZ37qs81cWqEQmktOAY43Os4n3cgOGjZCQKCLOFyo1ERERk5roJw4eoNAFcwyXAOkaZaFCO+7I0A1LItzXQOn8uZFR6EGoWECzlsSL9j+ys/kKkL+yrAaHIl3mEt/PbeC5SQJhtVG4kIiIiM9P9WE7Y36C8GcsZ9BFISKIHhOqCtIbOVpo65uDTFLLaPVMbZ0iLJfq37YobEAIpzTiK/C3HciqXUuZGjAKCQoKIiIjIzAkIV9BNgc8Cp+zfoBxb9bDg0QVp5Zr+1owx+NTT98jOShCJGRCacKT8gHHexhkU2YxhGUE/UAoJIiIiItNbD44TSFnH87B8gYTfY4AUW6MG5byhrXs+ubpCbTYoPz4gZIH+h3eAz+KtxQ14GnF4HiDlLaxniK0YliggzFbqSRAREZGZ81yzFcsSUjbwGgz34Jlb3aAcf0FaWmlQbls4n8QaQlrb5+nEGCCj7+Ed+PE03qvegKceC/yaMq/gMrZxBpZPaZuyQoKIiIjIdH+mORfDbXg2chawlZQcRQK2lg3K8yDLatqgDIAxmCSh75HtpKPluAGhgMWwA88ruZT/4AEsixUQZjuNQBUREZHp7Y7q88xteK7mChx3UMRRrlFA8NAwr5W27nmEEGofEJIEY6B/x664AcETyGPJ0Yvlz7iU/+B+BQSpUE+CiIiITF9nYTkbz/XkCWzBchZ9BKjRBKPswAbltPa/vwSstfTv2EV5aDzek1tGRo6EHGN43spq7t/fyyGikCAiIiLT1lYcS0i5knmMcw85Xl2dYGSJXVI9sUF5YdcRaVCeYF2OwZ27KfWPQT5a2MlwZBTIKPMuLuX7bMGxVAFBFBJERERkOttcDQhX8BwcX8DyXAZrN8HI1jtau+djc7mablB+TEDI5Rjeu4/xvpG4AcESqMdS5gwu5cusIMdSyvqhEoUEERERma4enWB0LaeQ8jkyuhiuUUAoQ761ntauzkqD8pEKCHnHyL4+RvcMQi7mF8bTjKPM+azhL6snCAoIopAgIiIi0zggnIthCSmbeDcpt+Opp4iPviQtAB7q57bQPG8OPvU13aB8IOMco/1DjOzuj/uklpDSgqPIWlZxC5tVYiRP8XOoSyAiIiJT3oETjK5lFYbPUKSeIiH6DoT9DcodlYBQTo9oQCgOjzC8Yx9ROysyyrTiGOMGVrGO63As0xQjeapMKSIiIjKV9WBYTuBULC/kEziW0Ecgqf4nphTIJZUNyvV1R6y8aCIglMZGGXh4d+UJLdar3EDKHBxFPslKzuFHWF5CICHTD5c8GZUbiYiIyNQ1sdjrCtrJ8WngDbWcYGTrHK0L52Nd7sgGBGtIi0UGHqlBQGjG0ccXuZJzuQ7DyRNnJSJP8TOpSyAiIiJTUg+OxXiu5Lcp8F0sb2CItNp/EDcglCHXXMecY7qxxh7ZgGAM3nv6HtkZ9+kskNKEo8x3yXMaGRNTkhQQ5GnpJEFERESmZkBYTsomTibjXgKLGNsfEOLJgBTqO5ppnteBDwFqvUH5cQEhyzIGHtkJPiNad0UgpRGH58eU+QsuZ5xGDBcQ9MMlB0M9CSIiIjK1nk0ewLAYzyb+goQ7KdNUnWAUv0E5QOP8Nhrb2/HlIzwJ1CSYxND38HbSsXK8V7cBTz0W+Dc8r+BS9nA2ljvUqCyH8OOpSyAiIiJTJiDcRsJiPOv5EAlfYJymmk0wAloWzqOhve3IB4QkwRpL/45d8QNCHRb4H1L+jEvZw1kKCHI4fxhFREREjrY7MJxNBmRsYhM5VjBIRgY1mWDkDG0L55Grqz+i/QcTbM4yuGMP4/2j8ZaleQJ1GBy78byaNfwLD2I5XgFBDp16EkREROToOgPL2XiuJ0/gVhxn0Icnw2BmzgSjRwNCjuE9++IGhECggMExRIk3s5Z/oQfH8VqWJodHJwkiIiJy9PwCy/PxXEMbJe4hz58eMMEorjLkWupo65oHJIRw5Ht4Tc4x2tfPyM7+uAHBklBPmRJv5FK+wxZtU5bJ0UmCiIiIHB09OJ5PyiUcS8oXyfGimgSE6gSjuvZGmud3QODoBATnKA4OM7KrP2YPQoYjox7wnKaAIAoJIiIiMn1txrGMlPUcT8LfYPgthmsUEDw0drbS2DEHn5aPypaAiW3Kgzv2VsbGJJF+dw5PA46Uc1jFF9isgCAKCSIiIjId9VQDwgZeA9xLoJ0xfPSAUN0r3LJgLoWWpiM/weiAgFCOv005IyPQjKPIxazik/uDl4hCgoiIiEwjEzsQUtZzGpZPUiJPqQY7EFLAJrQtmk+uvo5QPjrPzsYYfFqmP/42Zc8cHEWuYxXXshXHEgUEicfqEoiIiEjNZSTMIeEtBNZzITm2MIYlrcEOhBRMwdK+qBtXVzgqE4wmAkKWQd8jO8jKIf425XHuYg0f4Cws67RJWeLSSYKIiIjU1s0YkmrxzyauI8+FDBLwJJjIi11TcA0F2hZ2kpjkqAUEjIEkof+RbYSij9monNKEo8jXOYVz+DiGD04UVokoJIiIiMh00INhOYGLKTCXO8jxbnrxmMg7ECYmGLU10Dx/LmQQ0qP0cj1JMMbQv30n6Vgad5tyI46UHzPCqZxCoIdEAUFq8mOsSyAiIiI18QCWxXgup4MCn8PxagZrN8GoYW4LTR3teB8gO3rPzdZZ+nfuoRR3WZqnDkvCf5DwClaxc38AE6kBnSSIiIhIfD04FpOylmeR54sYFjNAiq3NBKOm7g4aW5tJy0e3d9fmcgzt2Rs3IHhCNSDswPEmLmYnD2I5Hq8fNFFIEBERkelhYpnXBl6I4YtkPIPhGgWEBFoXdpJvaJgCAcEx0tfH2N6heAEhI5AjwTFI4M1czH9yhgKC1J7RJRAREZFoeqoBYR2vw/BdUp7BaA12IKSAMbQf002+oeHoNShPPFA5x+jAUPxtyhYo4ClzKqv5GVtxfEoBQWpPJwkiIiISQ8L9GE4gZT1nknA7xdqNOLV1ltYFXdhcbkoEhNLoKMM791XGnMbapmwJ1GPxnMmlfIst2oUgR/DnWpdAREREJvk4m/APGE7As56LyHMnJQyeUJsRpznaj1mAtfboBwRrSItFBrZVtynHGgmT4GnBknIhq/g0m6snNCJHiE4SRERE5PDdgAEyXolnExtxrGS4RjsQypBvrae1qxOyjBCO7mAfYwzeewZib1POKNNGjlE2sYYb6MGxTAFBjiyNQBUREZHDcyuG91fbh6/kVpo4h348WW12INR3NNM8rwPv/VEdcVpNCBgyeh/egR9P4y5La8ZR5g7W8L7qGFktSxOFBBEREZkGzsDyKTzn0cCxfJocb60uSYvbf1DdgdA4v43G9nZ8Wj76j8tJgrEJ/dt2UR4uxg0IjTgG+Co53kI9sIpMAUGOSg7WJRAREZFD8otqQLiWOSzkqyS8lV7S6AEhVP5qWTCXpvY2fLk8JR6XrXMM7NxLeagGASHlR8zjXVxBRmF/TBI54tSTICIiIgevB8fzSVnDIsp8GcsJNdmiPLEDYVEX+fq6o74DYX9AyDmG9uyLv025AUfg38jz53yEERIMH9E2ZTl6VG4kIiIiB2dztYF2PceR8BUMz2KkBgEhBZM3tC7owhXyR32C0YTKLoRBRnb2xnvNWtmmbDBsY5xXcCW/3l/KJXIU6SRBREREDj4gbOJkAn9NxnxGarMkzdY5Whd1Ya2ZUgGhNDzCyK5eohVVBQI5DJZBMt7MlfyaX2B5vgKCKCSIiIjIVNezPyC8noTPEmihWIMm5RRcU4G27k4SEkI6NaptjDOUi0UGduyJtwshkOGAOkrA21nBz/eXcokoJIiIiMiU9iCW40nZyHvIuJMyOYoEbOSAUIZ8WwOt8+dCxlHfgbA/IBiDL3v6Y+5CCAdsUy5xFpfwnf0nNSJThKYbiYiIyBNJOAPL8XjW8yEsd1MkR5mAjfj8kFUCQv3cFtq65xFCNqUCAmQMbN8JaRbvqcniacbi+TCXcA9bFBBk6rG6BCIiIvK4B/eEeSRsJHANV5JnIyP46FuUqyNOG7vaaOqYgy9PoefkJCGxlr5HduHHynFHnbbjGOVq1rBe25RlqlK5kYiIiDzq4xiS6obfa7kFywdqskW5ukO4ZcFcCi1NlR0IU4i1hv6du0lHi/G3KQ9xN2tZyY+wnKwmZZmaNAJVREREKnowLCdwPXlSPkWedzJAWm2xjWdiB8LCTvL1DVNmgtH+gJDLMbxnH6N7B+PtQvDVgFDm78jxegIpK8lItCxNpib1JIiIiAjcj2U5gU20McZXSXgnvTUICB4wCW3HdE/JgGByjrHegbgBYWJZWsq/MMqprKCEAQUEmcp0kiAiIqKAYDkBz0YWAH+N4SSGSLE1WJJWsLQu7MLlclMvIDhHcXiEwW17Kl2bMZ6SDlyW5ng5F/NfZPtLukSmLPUkiIiIzGY9OE4g5Up+G/ga8ByGaxMQbL2jfWEXiXVTMiCUx8cY3LGH6lv+ycsI5EjIMUyRt7Ka/+I0LIn6EGTqU7mRiIjIbLUZx3JS1nM8jn8g4zmMktZii7JrKjBnUTdJYqZeQDAGXyrTv213vKejQIYhIw8Y3s1afspWHHcrIIhCgoiIiEzlgLCMlKs5GcN3yDiGMXz0gFCGfEs97QvmE0imzA6EAwNCRkb/9p3go+1CyPbvQoAPcBFfZTOOJRp1KgoJIiIiMtUDwgZeQ+AbeDoZx2Pib1Gu62iktXs+IWQwxQICSQIJ9G/bTSj6eNujAp5WHCXWs4rN2oUg05Eal0VERGZnQHgbhnsokifgiblgNQNSaJjXQtPcDnxaZirO8bHO0b9jN6XB0XhdmhOjTse5i8s4q7oLoboVQkQhQURERKbaPX8rliWkbOQMDJ9kHEuZgI1YWZABHho759A4p2VqbVE+MCDkcgzt2cvY3qGYuxAqJUbjfIff58/4NZ5V2oUg05PKjURERGZDQDgbwxJSNnEelrsYJyGNHBBC5a+WBXNpmMIBwTjHWP8AY/uGYu9CsHgeoI5TeQdl6tAuBFFIEBERkSkcEO7As4k15LiZUQIpCSZyQABaF3VSaG4iTOGAUBodYWhnb7wCK0+ggAUeIeMtrKCfs7F8RLsQZPqyugQiIiIzVEbCPBLWE7iGTeRYywCegMFELDlOAQdti7rI19dPuRGnBwaEtFik/5FdlYJrE+UaBxwJjhHgdazhV5yN5Q6NOpXpTcvUREREZqKPY6qlLoFN1fk6/XgSDEncgGDyprJFOZ+bugHBGDKfMrB9V/UXInzRQIYD8mRkvJPV/JytGnUqCgkiIiIyFW3GsIzA9TjK3EWB99BbmyVptt7RurALmxhCOkVfnicJJAl9j+wilEKcp59Q3YXQhKPE+1nN17ULQRQSREREZGraimEJgatpYJx7yPFm+ihjorXo7g8IrqFA28JOkim4JO1A1lr6d+zCj5VjPvl4WnCMciWXcNv+0bIiM4RGoIqIiMwUZ2D5FJ61tFLgizj+mOEanCCUIddSR1v3fMiyqR0Qco6hvb2M7RmMOckopQlHkU+ylnPYgmMpHu1CEIUEERERmVIexHI8nmvoIuXLWP6wVgEh39ZAa1cnIXgIU/e52DjH2OAQwzv2VUa1xHjqmQgIJb7By3kTD5DxITIFBJlpNAJVRERkuuvBcTyeK3kmKX+L4Q8Zqk1AqGtvpK2rk+CnfkAojY0yvDNqQPDU40i5H8e7OIVQPZ1QQJAZRycJIiIi09lELfx6fh/L1wg8i7HIASEDUqif20LzvHZ8GiCbwgHBGNI0pe+h7ZV/9zivRD15LAkPkfAyVvEQPRiWaxeCzExqXBYREZnuAWEDJwJfxTOfcXz0gOChobOVpo45+HJ5Sl+SxBgyQmXUqSfWJKOAw2AZpMSfs5aHeADLYu1CEIUEERERmZoB4Y8wfJmUDop4TMRFqdWA0Nw1h/q2likfEDCGxEDfw3sIRR9z1GlGPVDmdNZyPz04FmuSkcxs6kkQERGZbrbsDwivxPB1ynQwTogaEELlr6YFHdS1teDLU/+Z2FrLwM69pKPF2KNOLZ4Ps4avsAXHcgUEUUgQERGRqWQzjqWkrOf11YDQSpGAjXhPr1bZty6YR31zM2E6BIScY3jvPkr9o/ECQiClDscgN7KKm7mxeu1FZgE1LouIiEwXW6oPqRv4cyz3ME4d5doEhLZFXeTq6wjp1H8mNs5RHBxiMOaoU09KM469fIlreBuvxvBdAokmGcnsoJ4EERGR6WAFOZZSZiPvIuHTjGNJaxAQkoT2RV24usK0CQjlsXEGY486bcSRch/jnAHAy8gUEGQ2sboEIiIiU9xmHKtIWc/ZWD7FOAmBDBMxIKSANbQf0z19AoIx+HJK3yM7qr8Q4Yt6AnVYEh6hzGu5hj20Y1ipUacyu+gkQUREZCqrjNpM2cByHJ9gjIAniR0QTMHStqgL63LTIiBURp1mDGzfWTkBifHaMxDIkeAYpcTbuJyHOAvLhzXqVGYfNS6LiIhM0edgbsGxGM9GPkKOTzBKICXBROwpTMHWOeYc0401dloEBJKExED/jt34oo9VFzEx6jQh8F7W8lO24rhTAUFmJ50kiIiITMWAcD+GE0jZyBryXMUQHo+JHRBcQ462hV0kJIQwPSpqrHMM7txFOlSEXLQr7mnBMc5FrOaLbMaxRJOMZPbSSYKIiMjUDAie9ayjwFUM1CYg5JrqaF+0AKZTQMg5hvv6GO8bjRcQAikNOProYTXXcXN1D4XILKaTBBERkakiI2ETCSfg2cS15Pkog6QkWJKIAaEM+dZ6WrvnEUIG0yQgGOcYHxphdFd/3F0ITTh6+QYJ55NhSFRiJKI9CSIiIlMlIFTuzBlXczN5zqOXtLpFOWpAqGtvpGX+PLz3kE2PqZ7GONJyib6Htld/IUpA8NRjCTxIgVP4KINcQ8IKTTIS0UmCiIjI0XYLZv8M/nXcTp73VQNC3Pt0Geo7mmnunINPp09ASExClqWVSUaxAsLEqFPYiedtXMQACUYBQUQhQURE5OjbimEJgXdieRF3YTmtZgFhbgvN89rx5WlUTZMkJElC37adhFKI8+QSyMgBjjHKvJ3L+H/VUbMqMxIhXhYXERGRw7G5GhDeRp7n8XngNPooRw0IWSUgNMxroXlex/QKCFQmGQ3t6SUdLcd6tZlh8TRg8JzDpfyAnuqoWRHZTycJIiIiRysgLCNwDQ2U+TwF3sAAKSbazJ5KQEihobONpo5WfLk8rS6RyTlG9vUy3jscc5KRpx3HOGtYwz1swbFUk4xEHk+NyyIiIkfaRInR1TSR8iUcr2Y4colRBnhonN9GY3sbvjy9noONc5SGRxjYtodorduBlGYcJW7nEs5lK44l+OrVEhGFBBERkaPkNCx341lLKwW+guPltQoIzV1zqG9rmYYBwZCWUvr+N/IkoyYsRb5LPX/Kg2TcTVBAEHmSP4e6BCIiIkdIhuFuPNfSRoGv4ng5QzUICAGaujuom4YBAWPIQsbA9l0QiDfJqB5Lyq8Y5VQuIuUUMgUEEYUEERGRo+smDAmBjzCHIt/A8TKGSbERA0Ko/NXSPZf6lmbCdAsICdjE0L9jN6HoY00yChQwwF4Mb2ETvZyB5VyNOhV5KmpcFhERqbVfYHk+npXMpZWvYngxQ/joASGD1gXzyDc1EtLp14trXY7BnbtJh4pxGpUDGQ7IU8bzDlbxnzyI5XhNMhJ5OjpJEBERqaXTqwHhYubSxjdIeHH1BMFGDQhA66LOaRsQjHOM9PUx3jcSa5JRZdRpE4Yyy1nNP9CDU0AQOTg6SRAREamVf8HyPDzX0kWZrwMn1KLEKAFaF3WRq6+btgGhNDrCyK7+eE8mAc8cHCNcyyXcTg+O5Rp1KqKQICIicjSdXg0IlzGfMt/EsJiRGvQgJNC6qJtcXWGaBgRDWipVRp0a4o06bcLRx5e5kov5EZaTdYIgckh/NnUJREREIvsXLJ/GcxULyfO3+wOCiR0QEtqPWTBtA0JNJhkFPPU4Un5BnveSkfAzTTISOVQ6SRAREYlp4gRhI4vI+CaG50YPCB6who5FXZhCfnoGBMAaQ9+2nXEnGeWxwB7gL7iEQRowfESTjEQUEkRERI6WB7AsxrOeZ5DxLeA50QNCCiZvaDumG2Pt9A0IuRxDu3dTHhqPN8nIklUnGb2LNfxftuBYqj4EkcOhciMREZGYAeFKnknCtzE8h7EaBISCpW3RAqyxBD89X5Ab5xjpH2BsX7RJRgCeZiyB81nN37FZAUFkMnSSICIiMlmnVwPCWp6F49vAs2txgmDrHO2LukgSQwjTNyCUR0cZ2dkbc5JRSh2OfdzIBrbSg2OZAoLIZCS6BCIiIpPwr1iei2cjv0PCt8h4FqN4TMQ9CBMB4ZgFJDB9A4Ix+NTT+9B2yLJYjcqVSUa7+RrX8Sb+DsOrJ1bLicjh0kmCiIjI4XoljueSspbfJePbZPwWYzUICPWVEwSmeUCAhIHtO8FnsRqVJyYZ/St1nA7AzzXJSEQhQURE5Gj5BZbnk7KOPyDHN/Ecy3itAkI3Ccm0DQgAiTH0bd+JjzXJyBMoYEjYS+AvWE8/v8Zyr/YhiEQJ9roEIiIih+gWHM/Hs47jyPGtWgUEV5+jfVE3TPOAYHM5hvfuozw4HusEIcOR4QgY3sMl/AdbcQoIIvHoJEFERORQ3I/lBFKu4PnVE4QFtQoIrYu6pv0JgnGO8cFBRvcNxZ1k1IJjnA+wgu+wGccSNSqLRP2zq0sgIiJykO7DcgKey3h+9QShNgGhoVDtQZj+AaE8XmRw5754TxwTjcoDfJzV9HCdJhmJ1IKmG4mIiBxsQHgRnnU8D8t3yOiqaUDIsmkdEDCGJAv0PrSdUArxGpWbsBT5Ji/mjQC8hkCiRmWR2FRuJCIi8nR+juWF1R4EyzdrHRBClsF0DgiANYbebbviBQRPoAGL59+wnMafEMhAAUGkRjlfl0BEROQp3FcNCJfxB1i+TcbCWgSEXFOB9kWdlf0B0z0g5HIM7dlDOlSM16hcmWTUS8LbWUMvP8EoIIjUjk4SREREnszECcJGnkPCN/G1Cgh1tC2YPyNOEIxzjA4MMrZvJE6jciDDEsgDntNYzb9Vm8c1yUikln+WdQlERESewMQJwsSitBqNOc0119G2oHOGBARDeWyc4Z37iHiVPM1YAhewmm/Sg1NAEKk9nSSIiIg83qM9CL+D5VskHFuLTcr7A0Jg+gcEY/A+0L99V2UsSozXkIGUZhzjfII13MyNOJZrkpHIkaDpRiIiIgeamGK0lmdTx9+S8VuMk2IivlgrQ76ljtbuLkLwlT6Eaf00kWCMoe/h7aRj5biTjMp8hzxv4MdkfFGTjESOWPDXJRAREXlcQNjIs2jgmyT8VrXEKG5AaK2jbUEXIZsBAQGwzlQalWMFBE+gDovn/2E5nYvxvJRMAUHkyFG5kYiIyIEB4UqeSeA7JPw2YzU4QWitp7V7Pt7PjIBgnGOkt5/x3tF4jco5wDJCiXeylt08gGWx+hBEjuifbV0CERGZ9X6+PyA8A8e3ahkQ2rrnE2ZQQCiPjjKypz/ea0eLpwFDmSWs5T56cAoIIkeeThJERGR2m5hidCnH4vgO8Lu1CAh1bQ20dHXOnBMEY/Bpmf7teyodjjG6HAMpBRx9bGQDn6FHjcoiR4sal0VEZPaaKGO5kmNwfBvD7zNSgxOEtgbaZlBAmGhU7n1oG76Yxhl3OtGovIevcC1v4e8wvJoA6kMQOSovAnQJRERkVjq9GhAuZyGWb0PtAkLrTAoIgLWGgZ278eORAsJEo3LKrwicCcDPyRQQRI4elRuJiMjs8y9YnofnGhaQ8q3qCULcKUYlyLfVz6wTBMDkHMP7+igNjMVqVA7kSDAMYngHN9DHNiz3qg9B5Kj+WdclEBGRWeX0akC4nG7KfBvDc6snCPEWpZUrAWEmTTGCSqNyaWSE0T0DsV4zZlgC9SRknMkK/pWtOAUEkaNPJwkiIjJ7PIjleDzX0EXKNzE8l2E8thZjTucRQpg5AcEY0lKZge17Kq8Y4zQqe+bgGONSVvMlNuNYokZlkSnxZ16XQEREZoUeDMfj2UhnNSAczwgpNvIJQnUPQvAZhDBDnhYMkDG4fSf4SE8PgZQmHL18ntVcxdU4likgiCgkiIiIHCmbMSwnsIK5BL6BYTHDNWhSbqmrBIQZdIIAYI2lf9cefNHHqUEIeOpxpDxAgfeRkXCxSoxEptSfe10CERGZ0bZgWEbgGtpwfAPHidUThKgBIddaT9uCrhmzKG2CyTlG9vVR7B2J16icx2Dow/AnrGEndRhepklGIlOJehJERGTm2oxhKYHzaaLM35DjJIYiB4R05p4gGOcoDY0wuncgTkCoNCpnOBIy3s0qfr1/V4WIKCSIiIgckYCwjMDVNJDyJXKcUouAkGuuo7W7c+adIBiDL5UY2Lk3XnFywNOOY5wLWc236cGxWH0IIlPyM0CXQEREZpwbMCwjYxN5Uu7F8WqGIvcgVANCW3cXIctmVEAgSSBJGNi5G3wWq1HZ04xjiLtYzQ1sxrFcJwgiU5V6EkREZOYFhAvIcFgyPkeeN9fqBKGtu3PGlRgB2FyOgZ27SYeLMRuVLSk/5lj+gleRcR4BbVQWmbJ0kiAiIjNJwgVkJGR4Pk2et9YkIDRVA8JMO0Gg0ocw0t9HqX80VkAIFLBk7CDHqZxJkTwoIIgoJIiIiNReRsJbMSRkrOMuWnlnTUqMmgq0LegkZFD5HzPoocAaymPjjOzqjxUQKo3KljIJ72IFD3EmlqUE/cCKTG0qNxIRkZkgoZWEjxHYxFYaeB+9tQoIXdWAMLOec40x+CzQ9/CO6i9ECW6eVhye81nNX7EVx1XqQxCZDjTdSEREpreMhB9ieCmeTXycAu+vbUDIZlxAAEiMYfDhHZBmsU4RUlpwjNDDWjZzI44lmmQkMm0+E3QJRERkWt/HflANCFdxHQ1cyCBlskhT/asBwTUUaF80cwOCzeUY2r2XsX1DsRameRqxeP6BbbyWXQT+ikCiPgSR6UI9CSIiMn0DQg+Wl+LZyFU0cCF9pPEDQo72RV1kMzQgGOcYHRiMFxB8tVHZ8zAp76GHMi+tNpOLyLShciMREZmeNmNZRspGLqXAGgZJMRF77VKw9TnaF87cEwRjDOl4keFd++J0KQYyHBk5SpQ4jcvYoY3KItP080GXQEREpp0tOJaRsomLyXMl/XgyLLHKaFOwdY72RV2VLzkDAwKm8mp/YOeumE8EnhYsngu5lH+sblRWQBBRSBAREamxFeRYSspGPkSOqxnAk2FqFRBCmJnTOq2xDOzcTSiGWBuVU5pxDHMHq7mFG3AsV6OyyHSlxmUREZk+tlYn5FzNUgybGcHjMZh4AcEULHOO6SZJzMwNCDnH8N4+RvcMxGtUrseScR+Wl5FQ4mL1IYhMZzpJEBGR6eEiciwhZT1nkrCZEUL0gJA3tC3smtEBwTjH+NAIo3ujBYSAwwD7yHMqKxmnAAoIItObThJERGTqexDL8Xg28G4sdzNGhieJHhCO6cYaO3MDgjGE1LPvoW2QEeNVYUZCqJ4ivIGVfINX4/iuyoxEpv3nhS6BiIhMaT04jsezjrdi+TRjQICYAQGX0LawC+tmbkAgMZAk9O/YBZ5YfQiVRuXAJazkG9yqgCAyYz4ydAlERGTK2lydYrSO15HnbxgjTyAjifSSKwA2oX1hN66QJ6Qz9/nW5nL079hFaWA03kblJhwlvsil/AU34PgIHlRmJKKQICIiUitbcCwlZQMvw/BNijTiCTEDQpJA6zHd5AqFGR0QjHOM9vUzsqs/3sK0egzwKxwnA4NqVBaZYZ8bugQiIjLl9FQDwpWcAHyZMo2UIwcEoHXR7AgI5bFRRvb0xzpByCgAlmHgXaygnxxqVBaZYbRxWUREppZKk3LKeo7D8nVS2ikSsBFLjIDWRV3k6mZ2iVFiDN6n9G/fU0lFk68fyLB46nCUWMIaHtx/4iMiM+sFgy6BiIhMGWdWpxht5FkkfB1PF0V81ICQQduiTnL1dYR0Zi8DNsYwuH0PpFm8RuUmHEWuZw33sFUBQWTGvmTQJRARkSnhbCx34LmGBaT8HfAcxvAYbMyA0Lqok3xDw4w+QYDKwrShPfsY2zsUb2FaI5aU7/LPvI4G4NMTV1VEZtxLBl0CERE56m7FcAeeW+igxNcwPIfRiAEhq/zVsmDerAgIlYVpw4ztG4q3MK2AxfMwcBpfxHPKxFUVEYUEERGR2K7G8H4C59PEAF8hxwsYIcVGPEHw0NI9l0JT44wPCIkxhFKZoZ17Y5UYZTjAUibhXVzCLs7Ccg5BP7wiM5cal0VE5OjJSICMceqw/DU5TmaQFBvp/pRVQkJLdwf1zU2kMzwgkECSJPTt3EXmo93lJ/oQzmc1P2QrjiXqQxCZ6XSSICIiRy8gXErCqzGU+Sx5XsNQ5IDgoal7DoXW5pkfEADrcgzt3osfS+MtTGvBMcrtrKaHGxQQRGYLNS6LiMjRuf9cR8JHCaznbgq8hwFSTMSAkEJjdxuNbW348sx/rjXOMTY4xPD2ffEaleuxZPyEAV5JB0UtTBOZPXSSICIiR1ZGwg8wfJTAJrbQEDkgUA0InbMoIBhDWiwxvGtfrBOEQA4D7MHwbq5hDIsWponMIupJEBGRI+uHGF6KZz3XU2AJvZEDQhka5rXQ2NE6KwICxgAZgzt2Vf57jIVphowcCY4zuIj/4gEsi/H64RWZPXSSICIiR04PjpfiuYrLqecj1YBgo339MtTPbaZpbge+PDueaa21DOzaiy/6eAvTWrF4LuEivsktOAUEkdlHPQkiInJkbMaxjJQNfJQ6rmWQlKxaxBIpINS1N9IyvxOflmfFJTXOMdrXz8iu/nh9CE1YhvkS63grP8JyshamicxGOkkQEZHa21INCJtYSo5rGcDj4waEfGs9LfPn4f1sCQiG8vg4I3v64/Uh1GFJ+TV1nE1Gwn1amCYyW6knQUREamszjqWkXMVpGDYzgidgMJECQgq5ljpau+fjvZ8dj7TGkHkY3LG7ErMmeyUnFqYZxkl4Dyvp5xdY7lWZkchspZMEERGpnS37S4z+nBx3MUbAxw0IrqFAW3cnIQTIZsdLb2sMAzt3E0oh1p3c04wh4QJW8lO24hQQRBQSRERE4ru1eoKwjlcBn2UcQwpRA0J9jvZFXYQsgxBmx4075xje10d5aDxOPYDH04JjmLtYyRYtTBMRUOOyiIjUwqtw/D0pG3ghhu9Spo0ygSTSy6kUbJ2l/ZgFJCSVU4TZEBCcozQ6ysDDu2P1IXgasHgeZJQ/op0xPqqFaSKikwQREYntX7H8PSkb+T0SvkJKG8W4AcEULO2LumEWBYTEGHy5zMDOPbFGnWbkScgYJuNdbGCEn5IoIIgIqHFZRERiOgPLc/GsYREZXwMWUMRjI+1C8GDyhraFXSTWENIway6tMdC/bTeUszh3b4unDkeJJVzKr3gQy/HqQxCR6meOLoGIiERxLYZP4bmSdhr4CoZnM4aPtiwtADahdWEXNpebVQHB5nIM7e0jHSnHKjNKacBR5OOs4R5uwCkgiMiB1JMgIiKTl1XvJx+mgU6+juMUhiMHhATaj+nGFQqEdPb01RrnKA4PM7htb7w+hEYsZX5IgVfxYzxfJKjMSEQOpHIjERGZfEC4lISAoYl7qgEhxUS6x1QPDNoWds26gJAYQyiVGdy5N1YfQiCPIbCHjNO5mBI3YBQQROTxrC6BiIhM5jmWf8ewmcBruJMm3hE1IFT3/bYumke+oWFWBQQSMMbQt30HWSnEuGNXphbVYQi8k0v5KQ9g+XOCfoxF5PHUkyAiIodva3Ur7zV8jGbey1DkgOChpXsu+YbG2RUQAOtyDO3dix9N45UZtWEpcRVr+Bq34FisPgQReWLqSRARkcOzpbosbSOXUeBy+moREDootDTPuoBgnKM4NMzg9sh9CCN8m3X8KT/CcDKheqVFRH7zc0iXQEREDtnmakC4ig+R53L6IzYpZ0AKjV1zKLTOwoBgDL5UYmjX3jhFwYFAAYvnYXKcSQb8ZKKQS0TkiekkQUREDs2jJwinY/lLRvFkmGj3lDI0dLbS1NGOL8+ugDDRh9D78A78WIRxp4EMS6CejIzXsJLv8SMsJ6vMSESemk4SRETk4E2cIKznTRjuYIyAjxsQ6juaaeqYM/sCApU+hOG9ffjRcqz5g54WLJ5VrOR79OAUEETkYOgkQUREDk4PjuWkbOAULN9gnHrKZNhIL5zKUDenkZb5nfhyedZd3prsQ2jCMsxfsY63cwOOj5DqB1lEDuozSZdARESe1oNYlpNyJYuBL1GigTRiQEgh11JHc+c8fDoLA4Ix+DTyPoQ6LCX+L3M5h4yEvEadisjB0zI1ERF5amdhOR7POn6HHF/F004pYqNyCq6hQFt3F8H7WdlOmxjD4LYd4InTh+AAwxiWd/MhBvghls+rzEhEDp5OEkRE5MmdjeVOPJeyAMNXCSxiPG5AsHWO9kVdkAXIZl9CsDnH0N59pCOR+hAsniYMGR9mJfexFaeAICIKCSIiEsctGO7AczmtFPgKlt9jjDRmQDAFS/uiLrIsI4TZVw1jnGV8aISxfUOQi/AFPZ4GHMPcyWpu5WYcS9SHICKHTo3LIiLyhM+vZGSsoUAjX8fxKoYjLksLgDHMeUY31tjZGRCMwXtP78PbwWeTf21X6UMwBH5JgReTY5QGYIl6EUTkcG4CIiIiB8pIeF31JVLKp8nxKobiBoQkgfZj5mNtblYGhGpKYHDnbihHCQiP9iHU8R4uZpifkSggiMjhsroEIiJygIQ2LJ/Ak9BDA2cwSoqNeIKQQdsxXeQKdQQ/OythbM4xvLeX0sBYnDKjrLoPIeU8VvINHsTyAfUhiMjh00mCiIg8ajOWC0jZwOW0sIzxyCVGGbQumkeuro6Qzs6AYKylODLC2L7BOAEhkNKMY4S7Wc2tXIfjeAUEEZkc9SSIiMhEQHAsI2UDyynwCYaqmxBi3CsywEPLgrkUmptmb0CI3YcAnjwW+HdynAQMcxEZyWwcJCsiUT+vdAlERIQt1YCwkbdjuYUhPD5uQGjsmkOhZfYGBKjuQ9i9N04fAmQYEgxFLKdzMYP8nEQBQUQUEkREZPJ6cCwlZR2vwvCXlMjwGEyk0+YUGua10tjWQijP3oBgco7hfb2kQ8U4+xACnmYMGR/dvw/hXpUZiUgcKjcSEZnNHqxuU76SxTj+vrpNOWAivUQqQ31HM82dc/Hl8uwNCM5QGh1n4OHdcQKCx9OMZZx7uYx3cjOO87UPQUQifm7pEoiIzFJn7A8Iz8DxZTLaKcYNCPm2BprnzcGnszggGEPmAwM79sS56wYC9VhS/i/1LOVGDHmNOhURhQQREZmsWzB8Cs/ldOD4CnAs43hspPtCCrmWOtq6OvE+MJur5BNj6N+5F9JIfQiWDEOZMu/lYvq1D0FEavLZpUsgIjLL3IhhFGiiwCDfJMcpDOOxkXbnpOAacrQvWlBZlJbN3oRgco6RfX2M7h6IN+60HUeZC1jBx/aXi4mIxP780iUQEZlVEr5PwmoydvJpCpzCCGnMgGDrHK0LuyrhYDYHBGsoj44zuncgVqNyShOOYf6aFXyMm3AcrxMEEakNbVwWEZlNAeFBDB/GA5up572Mxl2WZnKGOcd0YxJTOUWYtQnBABn923aSZRHKjAKBApaM/6aNN/EnFKkn42sadyoiNfoY0yUQEZkleqqlKeu5glaWRt+mnCS0LuwisbM8IADWWAZ37SWUQrw+hBwpgfdyPvu4T30IIqKQICIik7UZx3JSNnI+daxlkBQT6TS5+qjatqgTl88T0tn97GqcY2Sgn9LAWLx9CK1YPGu4lB+wFcdn1IcgIrWlxmURkdkQECrblE/F8DnG8NGWpVW3KbcsnEehqXFWb1OGyrjTtJzS99D2yh12slc44GnCMsxXWMebuQHHR/CgMiMRqfHnmS6BiMgMtrUaEDbwx8BfMk7Ak8QMCE3dHQoIVANBkjC4a3e1/GrSAaHSh+B5iDrOJSOp7kNQQBCRmnO6BCIiM9RpWJaQciWLMfwVZfKkEZelpdDY2UpDazO+rGW/1uUY3L0bP5rGGHc60YcQKHEGl7KbX2C5V2VGIqKQICIih/+IaUjwXMGzcHyFlDZK+Gh9CGWo72imsWMOvlye9ZfbOMv40DDjvSOx9iEE2rGMs4ZL+R5bcSxBSUxEjtznmi6BiMgMczWGhMAlzMPxZRKOoRgxIKSQb22gubNDAYFKH4Ivpwzu3BvnrjrRhzDAN1nNBm7AsUQnCCKikCAiIocrIyEAV1NHgS/ieG51F0K0gJBrKtDaNRef6rkVIDGGod17wUfYh+AJ5DF4tpHyPjISrPoQREQhQUREJvG8yj9hWEWgyKco8DKGI+5CSMHWO9q6uwgZs3qb8v6baN4x3NtHeagYo4A3w5FVv85ZrGcH/4ThQ9qHICIKCSIicrh6sJyMZz030sw7GIobEEze0L6wC8gg6LnVGEN5ZJzRvQPx9iE0YCmxkdX8LdfgOFllRiJydGhPgojITDCxC2ETHyXPtfTF36Y859hurHOzfptyNSGQkNH70DZCOcJW5YCnEUuJH5LjFZxMxssJJCozEpGj9DGnSyAiMs1t2b8s7V0YrmUQD7G3Kc/H5vMKCFXWWgZ37SUUowSEQI6EjH7ynMkqUn5GpoAgIgoJIiJyeLbiWErKOl5Fwp2MEaJuUw7QunAeubo6LUubuHE6x+hAP6WBsTh9CAkZjRjgPFbyazbj+Kj6EEREIUFERA7HWdVlaVfwXCyfp0yBFGJvU843aJvy/pumMaSlEsO7+uL1ITRhGeYOVvEZbqieComIHGXqSRARmY5uxHABgXV0Y/kBGc9iPO6ytMauNhrb27RN+cCQYC29D2/Dj6WTDwmeQD2GwK8o8GJgmItUZiQiU4M2LouITDfrMXyYjIxGRvkShmcxEnmb8txmGtvbtSztADbnGNqzDz+axtiqnJEjw1Ai4QwuZpBTsdUtFyIiCgkiInJIj5YJryWhDBT5DK2cxDApNtLneRnybQ00z5urgHAA4xzFoRHG9g3FCAiVMqNWHCUuZhU/YyuOJSozEpEp9LmnSyAiMo38BMN38Dg+QTtvZiRiQEjBNedp7erEpwoIj94pDZlPGdi1N85dszLu1FHka6zkBjbjWKJ9CCKikCAiIofjInK8GM96LqWBpbGXpdl6R3t3d2XMqari97PGVAJCmsUZd5rH4NlBHecCCQUCuuIiMsWo3EhEZDqolKOUuYazSbiSgYg9CKG6Tblb25QfzzjHcG8f5cHxOGVGhow8hpRzuICdnInlbJ0iiMgU/PzTJRARmRYBIWUdryNwKyMEAoYYE+pC5au0LugicUbL0h4XEMrjRUb3DsQad5rShGWE61jDN7gBx10KCCIyNWkEqojIVHYWljvxrOcFWP6BIi14MpIIL3mqRS5tx3SSq2/QLoTH3R2NsfQ+9Ah+3McZd9qIwfMTLC/nZDwvJ2jcqYhMVTpJEBGZqm7BVAPCM7D8DSmtpIQoAaG6Tbmlu0MB4QlYl2No9944ASGQkQMyhilzJqso8RPtQxCRqU09CSIiU1FGQkLG1bSR8iUCx1KM2IeQQmNnG4WWZoKWpT2GcY7xoSHG+0di3SU9zTiKfJAr+HcewLJYZUYiMrWp3EhEZCoGhPUkVNpcv0ae1zIccZJRGeo7Gmnu7NQuhN+4K1ay2d7/2QYhwjQjj6cZyzB3s47T2YhjlfYhiMjUp3IjEZEp9pjKjVguIeC5nULkgJBCvrWe5nnztAvhCVhr4447rcOS8mvmch4ZCSt1giAiCgkiInKoNmP5CCmbuIoWzoi9C8E15CrL0rzXZP7H3xCdY2SgvzLudPJXPMOSYUkJnMGHGOCtGPUhiIhCgoiIHGpAcCwjZSPLyLGGgbgBwRQsbQu6IMsqf8mjN0NjSIslRnb3E6XrI+DJYRlhLZfyI7bg+JJOEURk+lBPgojIVNCDYzkpG3gDlq8wRobHYGLtQkiYc+wCrLPahfBEIcEaeh/ejh9PJx8SAp4mLMN8hyt5Hf+E4WRtVRaRafa5qEsgInKUPYBlOSnreQGGzzBOUh10GicgAG2L5mPzOQWEJ2BzjuF9/fjRKAEhkMNQZg+NnE1Cxk/IFBBEZLrRCFQRkaPpOgyL8VzBQhK+iKeVMgEbcRfCwrnk6uq0C+EJGGsojowy1jsIuRiJg0A9jjJLWMk2/hnLC1RmJCIKCSIicrBuwvBBMnbSiOOvMDyTETw24i6E+XMoNDdpF8ITJgRDlsHQrr1xim99dR/CMFu4gi/Rg+MFGncqItP0I1KXQETkqEg4sbowrYm7cLyYEdJoy9LKUD+3hcY5LQoIT8LahMHdewilEGfcaT2WMr8ix4XcgGG5ThBERCFBREQORQ+Wk/Gs53pa+YvouxDa6mme14FXQHjim59zjA4MUxoYm/yZeqiOOzWUMZzFJYxyAaA+BBFRSBARkYO2pTrJaCPnU89H6I14gpCCayrQOn8eXj0IT3zjMwaflhnevY9IVz3QgsVzOav5CTfhSFCHuIhMaxqBKiJyJG3FsYSU9bwJw5cp4skwUT6PU7B1jvZjuoGETJOMnpB1jt5tO0iHizFOETyNWDK+xypexQMYFmvcqYhMfzpJEBE5Us7CsoSUDbwQy92UyAiRAkIA4xJaF86vNDooIDzxTc85hnv7SIeKccqMciQE+snxPtC4UxGZOTTdSETkSLgWw0V41nMMCX9NSnO0UafVPNC6qAtrLcErIDxhQDCG8vg4o3sHYt39Ao1YxvkgH+W/uB/LCWpWFhGFBBERORgZCZCxl2bgiyQcSzHSqNPqLoTWY+bhCgXtQngySeVbMLBrT+WaTfbsZmKr8gj3sJZP8wMFBBGZWVRuJCJS64Dwreru5Dx/iePE6iSjeLsQuuaQb2hUQHgK1jmG9vWSjfvJvx7zBAoYUv6HOs7nagx/pEZlEVFIEBGRg3U9lj/Fs56PMYe3MEyKjXSKW4b6jmYa21oVEJ7qRucc40PDjPeOxDg/r4w7dSRknMPF9HJfdd+FiMgMonIjEZFa2YxjGSkbuYACH2Qg8i6E1oldCGVd6ydNCAm+nDK4a2+c12IBTwuOYa7hCv6OH1X3XYiIzDAagSoiUgs91V0Im3grCV9kDI/HVAuPJh0QXH2O9mMWELIMNMnoSVln6du+i/LQeJwyo0YMKfexj5fyFlJOIegUQURmIp0kiIjE9gCWxaRczR8S+DRFMjxJlIAQwOQNbQvnkykgPKXKVuUhyoPjkJv0l8vIARljeM7iYxRZiOEVCggiopAgIiJPpwfDYjxX8gxSvkRGAykBE3HU6YIuEmsIqQLCkwaEia3KeyJtVQ4EWrGMczGX86/VIKgyIxGZuZ+jugQiIpFsxbCcjDU0Y/kSlgUUq0VGk1Vd0dW6sBNXyCsgPI3EGAZ37AUf4U43Me60xFdZzS304BQQRGSm00mCiEik51L+DwnXAGXuwfGC6qjTOJ+zHpq72sk3NGiS0dOobFXuJx0pTr7MKBDIYUjZDSwBEnIadyois+CzVJdARCSCzVjuxlPm4+T4s6gBoQz1HS3Ut7UpIDzdTc0Y0mKR0X39cV6DJWTUk+BYwhp28ACGcxUSRGTm00mCiMhk7cIxf/+o0w/QF3fUaa61nua5czTq9KAe6hMGd+2p9G9MfpqRpwXLGFtYy99Uy4yU0kRkdnyc6hKIiEzCfVhehGcDb8LyN4wRoo46bcjRvmgBIQTINEjnqdicY2hvL2N7BuOUGdVhCPyKNk6kxBgXTnSGiIgoJIiIyJO5H8sJeNZxPI7vU6IFDyQRPlt9ZdTpnGMWQGLINOr0KRnnKI+N0//wzkoh7WS+A4EMS6AeT8YprOTHZBgSlRmJyCz6XNUlEBE5DNdjOAHP5czH8lcEWimTRQkIARIzMerUKSAcjCxjaNeeSjiIEdHyWAa4gpX8mJtwCggiMtvoJEFE5JAfSEm4hoSd5Ojg2zhOYQSPiTCRP7B/1Gm+UZOMDobN5RjcvZvx3pHJ9yEEPI1YRvg+V/JK/gnDyRPfFRGR2UMnCSIihybhRiwrCHSwmQKnVCcZ2Shf3UNTVwf5xnoFhIO5iTlHcWQ4VkDIql9jgJT3k5DxE/UhiMjspOlGIiKHYjOWZaRcxRoaOYteUmy8UacN81poaG3GlxUQnj6uJWTBM7BrX6xXXoFmLOOsYCP/ub/nRERkNn7E6hKIiBykHhzLSdnIqVg+x8j+fb5RJhnlWxto656vUacHyeZy9O/YRWlgNE6ZUROWwFdZxZsUEERkttNJgojIwXgAy2JSNnAScAdjBELcUaet8+fiVWJ0UIxzjA8OxQoIgRwJnt00sQxI+JlKjEREIUFERJ7KtRgW49nAIuALeBpICVECQqiMOm1b0F2pfNcuhKcPCMbgyymDscqMLIFGHGN8iA+yjX/G8gKdIojILP+s1SUQEXkKGQkeuI16An+F4RiK1XVpEQJCkkDrwi6ShMrCNHlaiU0Y2rMXfDb5u1jAU4+jj89wCZ9jI04BQUREIUFE5CmfR7kRy0oC27gTx0mMRJpklFVCQsuCTlw+r4BwsDct5xjtH6Q8NB6nzKiAocTD5PkgGQkrFRBERBQSRESeymYsHyFlA1fRyqnVgBCnTDOFxq528g3ahXDQNyxj8KUyw7v7iDJwNiEjT4JlCWvp5R8xJOpFEBFRSBAReTJbcSwjZSNnUGANAxEDQhnqO5ppbGtTQDiUZ3pjGNy1u7JwbrLdIL46zWiUT7Cab9KD4xSdIoiI7P/M1SUQEXmcs7HcgWcdL8fxt5RwpCSxJhnlmutoX9ClSUaHwDjHSG8fo7sHIDfJLxYI1GEI/AfzeSElxjhPS9NERB7zuatLICJygHdWA8J6nkmOe/HkSSFWQLB1lraueXj1IBxSQEjHi4zuGyBKN4glw+BJeB/vZ4SfkiggiIg8lkagiohM+DiG8wk8jxbgrwl0VScZTf7RNAC5hNbuLiABhYRDeKwPDO7eU00Mk/4+eFpwjLOeS/khP8FyksqMREQeTycJIiIVCS8iISGjyF/iWMxYpElGAcigrXs+Np/TJKNDuUnlHCO9A/jRdPJ3LE+gHkeJ+xhgHW/D8ofomyEiopAgIvIkerCcjOcqrqOdNzNCOVqjcoCmrg5y9XVqVD6UG5QxlMeqZUa5SX+5jByQME7GOVxHEQeaZiQi8sRUbiQishXHElI2sJQCF9JLionwWAqVSUbzWmhobcaXFRAOSZIwtHtPnG6BgKcZxyhrWcuDPIBlscqMRESejE4SRGR2OwvLElLW8SosH2cIT5wp/JBCvrWe5rlzFBAO9eaUq0wz8qNpjKVpnkYco3yP+7mBn2FZrDIjERGFBBGRJ3IjhjvxrOe3sXyWEjl8vFGnrj5Ha1cnPtUL60O6Me0vMxqMs1XZkeAZosC5/A2en2ncqYiIQoKIyBPJSCgBV9MEfB7opITHRPhcDGCcoa17PmRZ5S85eEnC0K49lU0+SYTvdDOGjI+ygl9zP5blOkUQEVFIEBF5ooCwHssKAkU+SY4Tok4yAloWzidxRpOMDvWmNFFmNJbGGXfaiKXIV7iEW7kfywnqQxARUUgQEXkiN2K5hJSNXEEz72CINMoko6wSElq655KrKxC8AsIh3ZCMoTw2Hq/MqEBCYDeB5UBSLTMSEZGDoOlGIjK79OBYTso63oVlLQORThAAUmjobKWuuUmNyocjZplRQkY9lhLns5ptOkUQEVFIEBF5YpWxlykbOBHDbYwR8NhYjcr5tgaaOubgy2Vd60Nkc46hPfvwY37yOxECniYsfdzLVXyef1BAEBFRSBAReSLXYFiMZyNdZHyelEbKBGykSUYNOVrnz8VrWdohM8ZQGhtnrHcoRkColBml7CDwITISUKOyiMihsroEIjLjZSR4DL9HjpQv41jMGB4bp1HZ5Aztx3STkJBpktEhS4xhYPtOsjQjymypBiyB93E5P+UVWJ6pkCAicqjUuCwiM9/1WE7BU+YWCpzCSLxJRkkCrQvmk6BJRofD5hzD+3orZUZxlqZZBrmH1XyBH2B5pcqMREQUEkREHq8Hx0dJ2cgFNHFu7ElGzV3zcIWCAsLh3ICci1tmlMOQsh3Ph8k0zUhEZDISXQIRmbEexHI8nvW8HsdXGSOrrkub/GdfGRo722joaCNoktHhhQRr2fu/j5CV/eRfWWV4mrGUeSur+RI/wnKyThFERA77M1qXQERmpLdXA8I6fp+EuxknIZDEmmRU19ZAY0e7AsJhsjnH8N5esnEfb2naKJ9mNV/iJpwCgojIJD+ndQlEZMb5OIbryMjTRsK3MBzLOBkmwouRiUlGC7oJXgHhcBhjKI+PMbyrL87StBwG2EaBt/J6xqkn42sqNRIRmdRntS6BiMwwCS8iIQGK/CWO4xglxUaZm4PJG9oWzIcsoMfQw0oIkCQM7toXp+C1sjQtIeE8LmYvPydhiaYZiYgoJIiIHKinWot+FVfTzhsZjtSoXH3sbOmeT2KdGpUPk7WGob17CbHKjJqwjHAXq/kyN+K4W2VGIiIxqHFZRGZSQHAsJ2U9Z1PHJxkiJYu0NDKFlgVzKTQ3EbQw7bAYayiPj9P/8O7JF7tOLE2Dh2niBMboYyUZic53RERi0MZlEZkZHsCymJSreCmWHoYJ+ChFRlCGhrkt1LU04dWofHiSBLIsXpmRISOPocR5fJh9/ABLojIjEZFYVG4kItPfJzAsxnMJx2L4HCkFUog1ySjfWk/T3A4FhEmwzjG0r5dQjFBm5PE0YBnjk1zCV/kJlpeqzEhEJCaVG4nI9P8c+zgJ9RTYyT+Q4yRG8FE2KnuwBcucYxcRsgBBlSyHwzhDeSxqmZEh8BB1HI9jkHpgmU4RRESifnbrEojItA4Im7F8kMB2bqXASYyQRgkIAXDQ2t0FWaaAMBlZwsDuKGVGlZ6DPGBYysX0cx+JAoKISHzqSRCR6WszlmWkbOBi6jiNvkiTjLLKX63dndh8To3Kk2BzjqE9vWRjHnKTDm6eJhxj3MplfJMfY3mxyoxERGpB5UYiMj1txbGElE28noSvMUYgw0T5XCtDY/ccGlpbFBAmwRhDuVSm/6Ed8aYZJfw3OV5IxiAXa5qRiEjNPsN1CURk2jkNyxJSrubZZHyaIgmeJFZAqOtopLGtVQFhshIY2rUnxtK5DENGjoSM5fvLjBQQRERqRuVGIjK9bMYwRsaxNFLmrzDMoRypUTkF11yged48fLmsaz0JJucY2duHH0vjlBm14BhhC5fzbZUZiYgcgc9xXQIRmVZeRMJHCBT4JDmOZyxeo7LJW9q65oG2KU/uxmIM6XiR0X0Dk38VVSkzcozz/3Cs4EoMJ6lRWUSk1nSSICLTxy04TiRlHWto4lR6IzUqVx85Wxd0kiSGoJAwKYkxDO3eW/0vk/1iZOTICCxlFYNkGJUZiYgoJIiIVExsVF7Pm3BcxUCkEqOsEhJaFs7FFQrqQ5gk4xwj/f2kI+U4ZUbNWEb5BJfzXb6PJVGZkYjIkaDpRiIy9d2E4UME1vH75PghRdpIyTARSibL0NDZSlNHuzYqTzYgGENIU/b973Ym3UY+Mc0I/ps5nMAgQ6zQNCMRkSP2ma5LICJT2mYMHriaFgz3ktFOSogSEFLIt9XT1NGmgBBBYgwDu/ZVyrcm9wpqYmlaApzHcgb4uaYZiYhM5ZBQeTf0ISw6hRCRIyHF8BECJe7A8TxG4k0ysvWO1vmd+FQ9CJO+mTjHyMAg6UgxTrNyI5Zx7uISvsl9WL6gMiMRkSNpcg/6Z2K5Sx/cIlIjPTiWk7KRy6jj8qiNyiZhzrELsNaqUXnSCcFA8Oz7n22VHg8zqe9NqJ4gbKeZ4ynRRwNwniYaiYgcSQd7szVAYDNNFPkCZb6L42YuoERGwmYSlusDXEQi+hmWE0nZwNuwXE5fxEblDNoWzMPmcmpUjsCahL6d+8Az+VOEhIw6DEU+xIfZx0+wnKSXUSIiR9rBvO9J+GX1xKGfT+F5HSnXMc6P2MAbq/svAw9g2aoeBxGJIMNwIp51PA/DJxkjI8MQo8zRQ+O8OeTqGxQQYtxEnGN0cJjy4HiMMiNPI5YB7uUSvsgPFBBERKZuSOjBchye9aynjrfST5lRUgIvxPEVruWrXMuLWIxnCYFfql9BRCYVEBIg4yLasXwOTyvl6nT8ySpDXVsDje3NCghR7iAG71OGd+9j0mc8gYwCCSl7qOcjZCTcr0ZlEZGpGRK27q8Hfg95VleP+3MYHOMEhgh4/oyUH3INn+BajuU4PJDxQISyABGZbRL+sbosq4E7sfwB46TYOJOMXGOO5s55eK/qyBistZWA4IkxKy9Qj8FxIWvYzrswfFBlrCIiR+0z/kn/zllY1uHZwEkkfJFxLAFDUn2bl1T/U8KTkiPHiXjew2tJOIkHeTdFMhIexvLPehskIgehB8fb8VzFVbRwLgMRG5VtQtsx3ZUEkukjabKMcxSHhhndMxhnaVoTlpSvspJVnI3l0woIIiJH9XP+CX91K4Y78VxFN3AvnnrK8ITH/QZLQsYwnhKdZFxDAz9lPaeSkHEHnrOx3K5+BRF5Co+eXJ5KgTXRAsL+RuX5WGs0ySiGJCHzKYNxyowCORJS+sk4H0g4eeK7JiIiR+2j/gl/7WYSRnGk/C2Wlx/0XPJABnjqcOSBwN9Tx+V8iP8DwC+xHEfQh7+IPMbEOOUrWUye/0ORRtIneTFxqMrQ1N1OfWur+hAisbkc/Tt2URoYjdGsnNKOo8xSVrCVB7Ecr2ZlEZGjzfxGQNiM5XwC4/SQ4+WMkB702EFDsr9fYZBAxqsY5Xtcyx1s4nf29yv8p/oVRKTqdgx3EbiROVg+j6cpaqNyRyMNbW0KCLFuGtYwPjQcKyB4GnEU+S4ruJUzFBBERKZmSNiMZRkpG/gIzbyPocM87jfV/4zgGcMQOIvAz7iWy7mCVn63ehNYrrAgMsslvKg6zWicT+P4nbiNygWa53bg07KudJQ7hiHLMob37Jt8o3Igw1WLVfN8AMg4RafMIiJTLyRsxlUDwp+ScB2DERYXVfoVYBhPmVbgMhz/zI28D0jowbMVw23qVxCZlW7BshjPOjbSwOsZjtio7AwtXfPIVN0ejbWG4X29hFKIM82ogKHMWi7kP9mK42w1K4uITBWV4/ybMZxP4Gqeg+dHlGnDk5FEfHiv9CsEclgagYwfknAlF/IdoLJd9UUEEt3ORWaFidrzdbydAp9nlJQswp6VyicNbcfOJ1dXrzKjSIxzlMdG6X94d4xmZU8DliI/5FWcwhbgHvWriYhMqc99wHAeGR+ijZQvAO2UIweEyj8pwWApE+gjUOKPKPNtruNzXMsfcCKeZP9+BS1jE5nJ3r4/IByH5XaKBHykP/spNHV1KCDElmUM7d4X4zuUVUPGOAnLeDmel2qakYjI1AsJl1Vvy/O4izzPZTRSPfCT/xMNFkMRzwiBlFMp83Ou41qupZPFWsYmMrMfNkk4kYxraMFyD9BCEeI1KjfT0NqigBDzY9s5Rvr68GM+Ri+Cpx5Liau4jH/hASzLVWYkIjLVWF5BwvfJeD3tZLyBMhlZdVFarQPKxDK2jDyWk/G8i9eScjwP8t7qecZ8DF/XGyaRGaMVx0fxvIy7yPPHjOCxEV4KVBuV27o78QoI8T6ojSEtlRnauZcIZz2VEdllfs7vcDYvBt6mz3cRkamo8nF/K4b3E7iay8hzOb2RmgcPXlZdqGOpAwI/p8gVXMZXAfUriMwUW3EsIWUDF1PH1fTFa1Q2zjDn2AUkiRamxWSdo/fhbaSj5cmOPM1ICBQIGE5mBfdVB93qmyUiMgVVDo7fT+DVOFZwBYPcQROOwJF8FTfRr5AxhMfzQur5CtfxFW7khfv7FTbjUL+CyPR0JrYaEF6DZWOUCWqVR0/IoKW7k8QoIES9QTjH2MAg6Ug5zk6EBixlrmMF9/E2rAKCiMjUlTzmf7+FhPlY/p2vk+c1DEe6iR/6zaRy42jEECiT43YMm7iQhwD4Dyy/p4U7ItPGxGnlDTyDcX6Cp5MSGSbOPoSm7g7qW5rVhxA1IRiy4On9n22VIDaZ71QgUIcB/pUCJ/ETinxep8MiIlP6NnDA/57hgLdTxnMqKb+kAUs4Cg/jBy5jGydHxjJK/IxNXMy1NFUDQsKntV9BZFq8jPgHEv6KHMN8FsN8SoQoAaEM9XPUqFwL1liG9vSCZ7LNyhmWDEcgz3I+ymi1t0EBQURkmoQEWErgDCyX04fnzWTsJofd/2b/yP/bWZJqCVKJThKupsxPuIl3AhmnEzgTy+0KCyJT1oMYPovnl9xMgZcwQhrlhDIF11ygeV4HvqyNylE/ep1lfGiI8sBYjDKjQD2WMT7BRfwfbsTxOZ0Ei4hMdcmT3NQrM8yv5qXA3zJOnhSivPk7/BtNBngKOOoAz9+T5zIu4AeAmptFpqJbcJxHygbOpcCtDEVamDbRqPyMRZU/8OpDiHhXSEiShN7/fYSQhsmXGeVISHiINo5nmCFWkulzWkRk6nvij//j8fTgWMEPCJxNAwZ7lJfdVJaxOYoEBghkvIpxvs8m7uJGfudxy9hE5Gh7AFsNCC8h4WaG8LECQqVReX7lCykgRGWdY3jPPkIpTH4nQkJGPQmWD3I+A7yMRAFBRGSa3A+e9O98ncBmHBfwC15JkUZewxg++ibmQ7/pJPv3K6RY8iymzOn8KXW8hgd5G2NAwnlYfqqbkchRcQ2Gt5GxkU4yvklgLiWI1qjc1UGhqYHgVbUSk3GG8liR4d19caYZNWEZ5F4uYz0/wPJSlRn9//buO/6yqr73/2uttU/9tmkMU2jx3pv8fjEJaNQEroigUdMRRAQLyMVQFHGoM0OvQxOjkhkm2BN/qXZGwVhiAwTEwXJNTK6J4DD1208/e631++OcL+iN1LO//f3kAX+E+GW+++yz9vrsz/p8PiIi8z9ImAoU3kPCRXydl7I/JX6HBumsHjt6IljozGdt4gmUcbyclBP5QyrczUPcT+Q0HG/H8FkFCyIzJmLwWH4NaPMP5Hkx9ewGphWX9tG/YpkKladjWTWW8cd2E0PoLd8TiOQAGMFyLF+hwieBu7UWi4jMm2fCM3rgvxrLecA/8xmW8odMZvTAz0qnXiFQwFECUu7FcRUXcDegegWRmXQLCReQciWbKLOeyYwGpqWQlHIsPXB1ZxaCvs2ZsklCdXiE2t4Juhv8XtZkz1IcDd7GRj7AvTgOVxZBRGRePReeQRgROYrIqwkUOJmU79I3S61Rn/y36AxjaxIYwxM4HM9d3MzfcjO/8X/VK2gYm8h02doNEK7j9fSxnlpGnYwC4AxL1qxklqujFuaDwFrSZpPayEQ2x4z6cDT5Mhv5AKcoQBARWZhBAsAlBE7GchUTpBxHZAcFHH6OTcvsTFdwNPBUiAROpMn93MLNXMt+HIYHFTeLTIsTuxOVr+P5OO6gRsRjew7Mu0HB0Jr9MElOE5WnJ0pgcs++TjDW6zGjBEOkhuEcAI5USCciMh89883yD4i8ioQPMsJRfJMcb8SQkNIpJp5rwY/B0MQDeRxHEHkjryblMLbzFtpEDPtj2aYHmEjP3oflZiJ5+gncDRxIk4jLplC5b78lFIc0UXlaFsskoTE+QWOk2vsxo9itEPNcyUY+xX04/gRFdSIi8/H58Kz+v79Iyl+ScDn3E3gTeSxuqiHhnPztHJFIBY9nDYb3soJvsYk/xRA5m8ADOKKOIIn0wPCibmvLBh8kz/NpkGYVIOSHSvQtGyK0FSBkvkRai2+3O5OVe82vTg1Na/FdCtzCcTh+RwGCiMjiCBIA/oyULSRs5FN41jGIw3SP8czVDYzF0e5Obvb8No5Pcwuf5UZ++/F6hS0kqF5B5Nm7DccReK7jYpbyeioZFSp7cMWEof1X4r32mtOyODpHZd8weHqfieC6fa0S3sEFtMiBmkWIiMzjZ8Rz/l9uJuFsUq7jPZR5FyMZbQymW+i+2erDEkjJcQcFNvFOHgXgX3H8morsRJ6R7TgOw3M9r8bxeWrE7naz94FpwNKD15AkieoQpoFNEpqTFSYe25dNsfIgjirv40rO5WEch2odFRFZnEECGO7BcgSeq/kEBY7L7A3izAQLHoOjD/DsxXIrCbdxARXA8FcY3qxUuciTejeW8wlcwiH08W08+9EiZjUwbXDNCgoD/apDmKaV32DZ98ijkMbesgiBQAED/CcpL2A5k5yjHlQiIvNdL4+GyOEENmEpcAotvk0fyZxqjfrUv7nDdOsV2uyHYRMtvs27eQMQeTOBU3F8YA4MjhOZayKGb2L4R3IU+RiGlbQImQQIbSgvH6Q4qABhurgkR2VkGFqx92NGhtgNEs7lSsZ5EKMAQURkcQcJnYdDEbiYCobjCfwnxTnYGvWpfgOLI+0GC4FfJ/I33MKXuI2X8hE8p6u4WeS/+HMcn8LzQ24mz5GZzUNIITdYpH/FMrwKladn0beWVq1OY7Sa3UyEFn/LpXyOh3F8VMeMREQUJACsI/BWHJeyg8hxOMbJYYjz6KjOzw9jm8CT8gom+Dqb+Cg38D/+r2FsIovbFhLWkXIdJ1PmXKr4TI4ZBrB5y9Cq/fCqQZg+Bib37uv9XX8gkMPQYpilnAcYHlAGQURkochm07udyFYS1vEYx/B98pyEJxIwc3CGwlM9PA0GSwuPx1LgUCJv5tWUOZKHOZEaAGfj9DCURendWN6J5wZ+HfgMTRJCdy5JL6YGph2wCuccUUHCtLBJQnV8nNZoLZuZCEM4PO/kfL7OKTiuUR2XiMhCke0GfgsJZ5FyPWdSYAuTpMR5Usj8ywQ8rlvc3OYn9HEd5/JhILIVSw44TQ9FWSQihhsxLKHEPu7B8lvUs6tDGFiznOKgBqZNX4RgIfUMP7KDx3tQ9bI29uGIfIkN/F63m9HcnZkjIiLP/rGR6U876/EZCrfT5EaGSAik8/jqdIaxTZASeB5tPsiNfIt38yrOIHAagTepXkEWiffgWE9gN5vJ81vU8Fl1Miou7aM0NKgAYRo565jYN5LFTITOq59AlRznAHCfuhmJiCw0Zlp+5q04ziPlGv4/CpzE5DS0Rp16f+9m6NEUum/JSjgcYPl74Fou5PvAVL94vUmThWkrCWeQsom3U+A2RjP6TqeQlHMsPWB1ZxaCvj3T877DOZrVKhM7MpqJ0I+jyUYuZxP3dIfpiYiIgoSnFTHchAHypHwRx5Hd4sZsaiAC5Ep5vE8J9dD72dpnHywY+jEEauTYTORmLmQPAD/E8Xw9MGUBORHH3+G5nsMxfJUmSTeH0PvANGtYdtAarOoQpnWVN8aw76c7IA29ZhE8BRwpD1HkcCIpFxE1WVlEZOGZnhkAhogDLqZBixMI/DjT1qgWfMvTv99+lFcMQgozdqjJdjdHFTwNygQuoMl2buIcbiHP8/FsxvIhzVeQBeB9WP6WwKUsJ/JxPAU8pucAoXs4Zcma/XC5nAKEaeSShOrwGLRC78eMDOBIyXMOF9HiIYwCBBERBQnPzgUEtuO4it20OR7HMAVs9018z4L3TO7aQ9/ypSz7lTUk5UInUAgzduU6B50m8QRWE3kfDe7jBo7lbNUryIJgeFF3E2j4MAm/QpM0qzqEvhVLyJXKqkOYzmXKWlr1JvXRySyOGQXKOOpsYT338B4S/k5ZUxGRhbsJmG6bSTiblBt5JYHP08KRZvAmsrvRsAXLikMOAguN8QqTw6PEhqdbNzAzAhEIFHHkAMvngKu4gO8AqleQ+ek2Et5ByjVcRj9XM5JdHUJ+sMSSNSvxbe0xp5NLEoYf3YGvtXsLEqZmIlgeZYDDqDHOeh0zEhFRkNCrqdaom3grCR+immFr1McLH9dgrCF6T21ikuq+sU4Xj5lswDqVJenD4vHk+UvabGIjjwLwrzh+TW/eZB7oBLaea3gNBbZRfbxppun1++oKjmUHHdCJrXXMaNrYJKExPsbkzrHe67YCnkEcbU5kI3/PN3G8VGuZiMhCNjMThLcRulNaH+JoDP0cQ62bT+j5SQihGUjTNsX+MiFE8uUyxYF+iJ603u4WSM5IyNX5qzOMzZHjxUTezGtwvIrt/AlNwPAOHPfrDZzMUbdh+RMCmzgIy5206SfNYDBi7IQYSw5YjXVWdQjTGiEYQuoZf2xvJ6zr5ZObmonguZONXMpDOH5HAYKIiIKE7AKFyBYSzuMr/E8OoY8X0swuUPC1NtFEiv19hHYb4xzFwX7yfaVOF6RGOrWRn4lgwWKINAlAPwmvJOVY/oARvsgPuJ/IaTjOwfAZBQsypxiuwvI8HJ5P43g+jYwGpnnoX7WcQl+Z4LXHnNaFPckxsWcvvp72tsqHbs7XUiXPa/kiI7wQw51at0REFv6GYKb/e+/HUMPS4i7yvIJKRuecY3cTsnoZpcEnhjJZazHW0KzVqOwdfeKhOXPzFTr1CgUcJSDwFRKuYB3fBOD9JLwDr7O9MidM1SFcza0UWZfZjJM2FJf3MbhyJb7d1nWeRjaxtGp1xh/dm0WxcsoACU3Wcxk3dicrK8ITEVGQMA22YjmDwFUsJ8fXMfx6d3Jr71mNbqAwdOB+5Mt9v9A1xSYJxEBjosLkvlFIYydQmKkrEAgYIsXu75nwMSLXcjH/BjxxBlxktkzdg5s4kYS/ZRKPyeB7mUJSyrH0QA1Mm5EgwTn2/fRRYrvHlqeBQBFLynaex+9yFykfVwMGEZFF8zyZ8f/iGQROwXEFwzR4LZHd5HCZtEY1nd9oYsde0mYLmzzx64U0JfhAcWiQFYespbR8sPOom8n5CgZHHU+DSOQtBO7nBq7mZpY9HiCcMYNHwESm3ILlMDxX8OsY/pI6IZP2vQFwhsE1K4lR28tpX2aShOrIKLERel/dDRFLJMe5vJlmt/hZn6CIyCIxez38v4fjt/DcyMuAu2mQz6w1auik3JcddADG0Hl7OaXbo8UmCWmzRXV4lNZEvfPvZrYTksd1jyB5fkKe63B8hHUETsNxBJHTUWWnTL8nJqSXSfkGlsOoEnDZzEMYWrsf+f4+zUOY7gDBWtJ2yugjj029mOhtfRrAUecOruDPeBDHi5TpFBFRkDBTpmYoXMvJFPg4FXy3RDKTGQqunGPZAWu6RxziL32oGudoVqtU942S1tqzMV/BkyOhALS5lz6uZB1fBOABHC/qHlMSmS63knAeKVfzYYqcmmUdQnm/QfpWLCO0FSBMN5ckjDy6o7OOZTMTYRd9HEqbYcrAO/TSQkREQcJMup2EM0m5nvWU2JTZwKZuoJAfLLFk9f74p3iLOXUsqTlRpTI8SmiGzkN2JusVIFLqvruN/CNFrmYd3wc0jE2mz1YSziBlE2dRYDOj2Q1MS/oLLFu7+im/e5INmyTUx8ep7BzNZibCEhxNTmEjH1MWQUREQcLs/Rm24jiDlGvZQp4zM3uTCdCG0opBBvZb/rRdVWySEIOnNjpBbXR8toaxGfoxBGrk2EyDW7iU3QD8EMfz9bCWjJyI4+/w3MBLiHydFglpBpm8AFjLikPWEI3VwLRpjxAsMQZGfroDfOz9mFEfjsCX2MjvcQqOj2rNERFZjOZGkeydRN6EYw3baPBiyvxaZjMUHKSVJjhLvr/0lAOcYrfzSqG/j2J/HzF60sYsDGNr4vEUyHMEkZN5DW1+le9yKimbsRyL4XPKKkgPbsNyE5E8S4jcTWA/PLHn71y3OHnJAfvj8nmi5iFM/yKeJEzs3oefOi753AOESELE0sRyPF9iL+dg+KzWGhGRxcjOkT9H5EgiFxHJcxIpD1MiIWT0BiuB6p4RWpPVTivUp3tWpikucQyu3p8lB60iKRegDTN2Itd2pzhM4ElZQ+R9rOF+buRYzibwNgIP4DLpPiOLkaHVHfjX4A4S/htN0kxeGqTQt3IJuVJRhcozsVQklma1Smu8lkXWM1DG0eYm1vNDtuPUPEFEZDFvFuaSzVjOJrCJQ4h8i8AamhlNe+0+6pYcuJpcsfCMNzCdYWyWxmSF6vAovuFnZxhbEUcOsHwOzzVczAOA6hXk2XuiDmEdBW7Nsg4hP1hiyZqV+LYyCDMSJDjHyCM/66xLvXyCnkAJi+dHFHkRJRqcg5rWiogoSJhDpgY63cDvEvkyLUqkxMwChcSy4qA1YOxTHj36Lw/jJIEQqY2NUx0Zm616BejDEkjJcQeG67mAn/3CdRN5Kq/H8fd4ruEI8nyVOg6fTR2CyTlWHLSGEPml3cQk4wAhSaiNjVHdNdZ7sTJ4yjjgVVzMP6kWQURE5t7grtuJbCZhHY/wCv6VPCfiu/OKTY8bGQO0I81ajeKSQTDmGW9mYgjEGMn3lSkP9RECpPXWzNcrtPB4EnK8mBan8Bocr2Y7x9EkYvgpju16+ye/xPuw3ExkKUtpcTeBFbSh5wC8e7ctPWAV1jmiAoQZiBAMIfVM7NzbWdd6WRmnipUn+Ssu593ci+McBQgiIov+UTMn/1Rnk7KFhI18gsB5DOIgu/oE30iZ2LkbZ5/9rx/SFGMcg6uWs/Tg1eQGip2pzTM3udlhiFTweJYTuZ4W3+YaTsYQ+QierVg+MEc/W5k9L8JgiAzzAfL8Cs1uDqFXKfTvv5ykUPjFwYUybZxLqAyPdFbF3roZdWYiePbRz8VEDPfpJYOIiDDHC1+fGLb25/RxbqYzFNpQWt7HwMqVT9sa9Un36916hWatSmXvKL6ezsYwtkABRxGAL2O4kvP5JgDvI+EcvIaxCbeR8A5SruU8yrw7yzqE4pI+Blc99++RPPt1p9VoMP7ont6PPAY8gzjanMFG/pJ7cRyuLIKIiDDnu+MYHsDyYgJX8An6eC0TpLjsAoW+VUsoL13S00RYm1iIhsb4JJPDo5DGma9XMESK3eNjOT6G5RrO598BFTcvdvfgOALP9fxPHF+ljs2kDiEFV0pYduBaQogQlUWYkSBhqli55XufiVDCEfgG3+VoSsBfa50QEZHu82aO//kiLyYSgb28hZT76c+4NeruMZoTlWfUGvVJn7VpIHhPcckgKw5ZS2n5QKdWYeaOIFkMjjqeBpHAW2jyIDdyDTezrFvQHDljDtagyPR6D5bDCWxgOfAx2uTwmEwKlR0MrlrZqetRgDAzX/UkoTo6hq97eq4kcXSa4Sa8k0/gOVLdjEREZP4ECZ3tyClYNlMhcByBRyji8Bn07zadKzCxax/tRrOTEejlD9pOiUD/fstZevAa8oOl2ahXMFTwtBjCcCkp93Mrp/EGHFu79Qp3qF5h0Xhxtw4h4Q4cz8usDiHAwKr9SPJ51SHM1NfbWkKrTW14PItjRoE+HC3ew3q2cw+OMzQTQUREfnGbPD+chuNDeK7jBRi+Rpv+TFujWsOyg9fgrMtk02OtxThHs1aluneUdGoa6sxtz2O3KLFTrxC4lwJXso4vAvAAjhd1jynJwvQgjhfhuZYLKHNzZnUIbSitGGRgv2X4tgamzRSXc4w+tpv2RKO3ICEQKGCAn9DkBfRT5SKi1gIREZmfQQI8MQTqOv6IhM9QJxKxmfweHlw+YelBqwHzrGYoPGWwkFjAUp+YpDY8SmiGzgN+pq586J4xLuG6V+ofyXM16/g+oHqFheqbOF6K51qOJMeXs6xDSMo5lh64VhOVZ5BNElqVKuM79vaeRYh4BnB4jmU9n+EkHH+jYmUREflF8+uM+p0EbifhPP6FY9hHmT+i1XMTwO5TGGI70Go1KQ0NEjM6Yx1DJIZAvlSiNNQPFtr1Jhn9qZ9JGGgwWNoE2kCO55NyCq9hGa/iexzPJAD/G8dfKFBYEN6D5fVEHCuwfB7P8kzmIQQgMSxZuwrzLGaMSBZf48jYzt2dGRS9zkTox+H5DBu4itNw/JUCBBER+aVb43nmTFI2k7CRzTS4iSESQkan/hNIJ5tM7t6DS3KZ/rFDmhJDpG/5Mpb9ylqKQ+VOoDCTxc0WQxVPkzKR82nzHW7mHE4gz6/j2YzlQ6pXmPf7yd/p1iG0+QBJhvMQAgytXIHL5TLLtMkz+OomCZXRcUIj9NrNKJJg8FQZ5ALAcIReDIiIyJNtKObrn/seLEfguZq/ocAbmMy2NWp55RD9y5dOy5nrqXqFVq1Bdd8I7WpzNuoVPDkSSkDKdhxXcRGfBuBNOP5K9Qrz0u0knEnKdVxIiZtUhzDPAwRrSdOU0Z8+ls1k5UEcdS7jcq7VTAQREVmIQQJEDDdiGKHAEF/E8VKqeGwGR6gi4KF/zQpKA/3TdvZ6ahhbY7JCdWS009bQ0fl7JrbnU8PYijg6iZNt5LiaddwPqF5hvjkOxyfxXMPLyPNl6phManZSSMoFlh64WnUIM8wl3WLlyQyKlYsYAv9CnhdxD00+rRcBIiKyEIMEgD/H8i4C17A/jm8Q+R80MgwUIiw5YBW5UnFaN0c2SSB6aqOTVEfGwc/CMDaAPiyBNnnuoMUNbORRAH6M41f1xnHOB80At7KCJvcTOIQGoVuu3su9Ac6w7MC1WGd1zGgG2SShWakwsWNfNsXKnVqEP2ADX+DbOH5H32kREXly83u41l1ETsHxHiZ5KV8l4SQsZVICpscAyHSChOZklXx/H865TtHgdOzvQiBGyPeVKQ/0E2Ikrbe6rVlnJFTs/NXC40lIeDGet/AHWF7CdzmBJmA4G8cDevM4J4P9Io4jCfwOf0OR36GOx2Xw/fYwtHY/csUi0WtPOYMhAhAYf2xPdsXKFf6By9nECTjWKUAQEZGFHCQAPEzkYRxvYDev4CESTiZi8N3Nb6+BgodWvUZpaABj7LR2dIkhgLOUBvvI95cIPsU30pkMFiyGSJOAoY+EV2I5jj9mjLv5Hg8Q2YrlTzF8TsHCHPoOOF6H51ouZpC3M5FtHUJ5yZCOGc30wpxzVIbHaFcava3SgUgOMEyQ8jq+xji7gC/o+ysiIk+/DV4YNpNwNinXcRp5PkiVlIjL5Hec6g1/wGpCBGbgyEWnuNnQrNao7B3F19OZLW6eqlcodIexwZcJXMVFfAPQMLa54gQc/4DnWo4i4cs0ILs6BM1DmJVF2Vp8u83oT3f2/hondGciNLiYy7mJe3AcoSyCiIg8PbdgfpNtj89Q+A5H4+jnaGp4TDYzFEIzkKZtyoMDM3IuO8ZIDJGkUKQ01I/JO9r1Rqdl6kzOV0gJtAgY/huBt/L7/Aq/xw84gWGuAraQsE2Bwqy4DcvNRPKsxPAFAktpkc08BGdYcuDqTqSheQgzuyg7x/juvYSW7/W77ingSPkew5zOSURO0HdVRESe8fZ3ATkTz60kbORyJvkYA9nOUGiN1ajsG8blZq6qOKQpwQfKQ4OsOGQt5eWDnU3cTM5XMDjqeBpEPKfQ5jvcwLW8m2WcRQpEzlhAAef8YGh1j4c1+AAJB9HE91yo3A0ShlavwLlEhcozvSAnCY2JSdLJZhbFyp2fYTmP99LkqKlKKxERkae38DZ2dxEpY3FsI+UICvx3WhllFBy0J5uYvCNfKs3oBiqGANZS6C9T7C/jQ9o5gjSz9QqGJp5IiQIvo8UJvIYqy/keH8WzFcuxGD6rjci020rCOaRcz3oGOZvxDOsQ9lMdwuxECBZiYHTH7mxmIvTjqPJxruDdfFPdjERE5Fk+lhbcb2SI5IH1tMjxBjw/oojDk82O3kFl1zCtaq3TunQmhUBop7gkYcmaVSw5aCVJKdfJKoQZu2McnsgEnpTnAR/gBXyDm3gVZxA4ncD7SR5vySnZexOOM0i5kZdiuJbxjNr+ppAMFOhfroFps7IYO0t1ZAzS2Ptk5RwGzwhlLiZi1JVMRESe/ZZ6oToFx0fxXMOvkePrtFlJM4O+8Z2HMMbAkoPWkCQJYZaOZNjEApb6xCS14VFCM8x0cXNn0FqxO+va8AkiV3Ex3wc0jG06bMVSA9osoc39RP4bTUImdQjWsOKgtaB5CDP/XbaWtNVm9JEMi5WbvIvLeK+KlUVE5Dk9mxbsb/ZRPA/juIx/JfB6EprkmOra0/NVixHGd+wixoC1s3MZQxoIaUppcIBlB6+lvN9Q51/MZL2CxdEgUCESOZ7AfWzi3VzHag7DA5Efql4hMy0s6wg02ELCf6NJ2nOA0A0ShlavwORUhzAbjLVM7t3XezgdCJRxNHmInWzm/ViOQB+oiIgoSPgFh+LZTMJGvgb8L8pYbEZvti2E1tT5YdP5e5aENCWGSN/ypSz7lbXkh8rgZzxYMFTwNCljOQ/DA9zEOVxAgefj2YzlQwv8fptutz9eh3A2/byeSnZ1COUVg+T7+lSHMBuLcJLQmKiQ1tq9FitHDBFLIMe72Eyb7SpWFhGR52bhv+HdRmAzCefxMEfRZoBXUs+uNWpsBdJ2i/LQ4Ky/gY0h4KyjNDhArq9I2m4TGr73IshnqnNNI008sIQcv0+eP+YP2M25/IjPEDkVxztU3PysvRvLO/Fcywuw/B31xw+W9T4PYaDA0P4rFSDMBgMmRsZ3ZTJZOdCPo8mHuYT382YcH9ExIxERec6PqEXyez6E5YV4LuMv6edtmb2Fhccn0w7sN3cKPq21GGtpTlaYGB4lNvzsDGMr4rrHvL5AgStZx/0A/BDH81Wv8MyiPww3YlhCiX3ci+E3qWdQX9OtQ1h2yFqcsbNWW7OY2VxCZe8w9eHJ3rIInWLliGWE5fwmDfZQBs7QUSMREXmOz6hFs816IYH7cBzDWdS5i/4MZyjkoL5vgvrYxMx3PHqyPUMI+DQlP9DPioMPoG/VEnBmJo8gmcfrFSYJwO9T55vczGau52Ce361X+FfVKzyt9+JYT2A3f06e36RBmt08hP1wVgHCbAXyabNFfXSSDF5XBEpYDJdxFrtwWAUIIiLSi8W1QTsMw4lEXs2deH6fHGtoZ3f0qFWpkyvkSIrFOVP8GUMgxki+XKbY3w/Gd84+z9x8hc5fLTyehAIvJvBmXoPl1TzEn9ACDGfj1Kbxl3glCZtJ2cQbKXEtVVJMlvMQBgmpTqTMSpDgLBM79xFaaa8tTz0lHCnf5iHeztsxnKMAQUREet7aLiLvILAZwwbGsByH4THyGc1Q6J77H9+5l7TZ6rYnnTtCO8U4y8DK/Vh68GpyA8WZn69guvMV2iwnsIkGD3ErbwQim7vD2O5QcfPjbsXyJVI28T+I/AVVAj6D65NCbqDAgOYhzN7CmyQ0KlXa1UbvxcoOg8GT4118Et999aOAW0REet7aLj6d/v2e6zkcy5doUiKlc0Sm5904kFiWHbAG5+bmMQ5rLcYZGpM1qsOjncnNjs7fM7G1mKpXKOAoAPAVLFdxPl8H4AEcLyJgFvFGJ2L4OpZ7cKT8MwmHU82uDmHFIWsxqkOYpS8gGOMY+enPCG3fexahH8cEt3MdZ/E1HEepWFlERDJ5XC1Chz3eGvVe2pxKCYMjZDVDgVZgfOeuWW+N+qT7ihDwbU+hv49lB61hYNUysAbaMxQkTNUrNAmME/AcQ5OvcQsf4Wb+Oy/GY4hsxy3aQPY93c1ei+socDg1fHZ1CCswzilAmLUgPaE2OtbpPNbrZOUES8oeClxBxPAyHTMSEZFsLN6i0W0EbifhfH7AK5ikzGtoZNsatd1qdlqjxrn53J6qV8iVS5SHBogG0nqzM2NhpuoV7OP1CoaEw2hxGq+hzJFs50RqAJyO46FFlFU4FceNeK7mjynxfip4bAbf1TaUVgxQXjKkOoTZixCI7TYTu/b13po4Eihj8VzIZXyNV+I4WEGCiIhktU1b7L//VhxnkLKJ91HkHEYybo26fICBlSvw7fac37xYZwnNFpPDo7Qm6p3/+0w2awp4HI4S4PkJkRs4hg9xeLdewQJvW+CboK3drjQ3cQApD+JZSYvY81TlFJJSjqUHrdU8hFnkcjnGHtvV+X4lPX5XOsXKD5LncLYT+VsdMxIREQUJ2V6DE7H8HZ7L+TR9/GnWMxT6Vi2hvHQJYR4UiU7VK7RqdSp7RzudkGZjvkIO1/0E7sVwJRv4IrDQ6xUM78dRJdDmLnL8HtUMsgjdsGrZwWtxiY4ZzeZ3q9VoMP7ont6Db4OnhCFwNBv4+uN1ViIiIlk9t3QJiBxJJGLI8yZSHqSPhJDRAzeB6u4xmpOVOTND4Sn3k916hVypyLKD1tK/ZnmnU1ObmemENFWv0CZQxwOHY7ibm/kHbuA3F3S9wlYc55DSZCMFfo8qaSbHjAIMrlqOy+cUIMxqCGio7B3ufdWdyiJU+Gs28HW+qQBBRESyp0FWAJ8nMoBlPU1ezl0YjsexlJSA6XEj2v1fNys18n0lknx+zsxQeMrIKURiCORLJUpD/WAN7ZmuVzBY2gTaRHI8n8CpvIblvIqHOZ5JAN62QOoV3oTjejzX8HJyfIgGEZNBIJRCcUkffcuX6ZjRLLJJQm18nOZYj8eMDIEEA4xhOIGvMMkngbvV8lRERLLeiskTHsLxQjzX8tskfJUm/aQZnAeHx1tPLjtoDW4edpaxSYJvtamMjNKaqHW6IM1GvUIZ8OzAcBMF/pJ1NOZ9vcJWLDWgzTLaPEjkYJqEnu87D67gWHbQAYTg1Tl/tr471hK9Z99Pd/S+6k61PK1zIVdyC/fgOEJZBBERmYbnly7Bz3khnodwXMp38LyJIhFHzKw1qo+MP7abSAQ7vy59SFNc4liyen+WHLSapL8w88PYIpFJPG3WYngvde7lBo7lDAJvI7Adx9Z5d08bGljWEWiwlYSDaeJ7DhBiZzM6uGolxKgAYTY/YOeYGB7tZOF6ey3jKeJo8EOex/t5H5Yj1M1IREQUJMxcoLCZhEv4LIF3MoDDZfSmzoFvpIzv3I21dt7lcTr1Cm1yxRzL1q5mcM0KXM7NXL0Cj9crRCp4Iodh+RQ38Hlu4SUchucMAj+cR/UKD2M5l5TreSf9HNctmu/9GGAK/SuXkRQKqkOYzQU2SWjV6p3sW6+Zt4AhB+Q5n7fQ5LsYFP6JiMi0bbrkl9tMwtmkXMtN9HFh1q1Ri0v7GFy1At+evycFbJJA9NTGJ6kOj0MaZ/oIUmf324cl0Cbhg6TcwEZ+CsCPcfzqHD6KcQKOf8BzHb+N5Vu0SAhYMqhDyA+VWLJ65by+vxZKkDD6yA7Serv3lqf9OAKfYAOv4zQcH9IxIxERmT4qXH4y24g8gOMkvsiR/L/08Vs0SDMZtuYgrbaJFor9ffOikPmX6Qxjg2K5TGGon0gkrbW69RczEuJ2/uoMY0so8CIib+I1OH6H73ICTcDwDhz3z7E3rhHDfwCH00/CNiKraULPx4wC2LxjyZr9CZHOUSOZtQChPj5OY7QKuZ4+00gCGCrkOYEvMsrbgc8qiyAiIgoSZscLMByDYYQ7yXE0BQ6mld1U5nalicsn5EoFYpi/z/sQAsZaSgN95PtLeJ8SGunURn4mggWLIdLAE+kn4ZVYjuePGeVuvsf9RE7D8XbMnNlY/T6ONxB4BbdT5FVUu4XZvQUeEGFo7f44l8zb4HNBMAZiYOyxvdlMVu7D0WITG/gU9+D4U9UiiIjIdG+v5Kn9BZa3E7iG/XHcQ+R5NDIYcAWdN+4Rlhy0ilyxQEjn/+mBqWFszckaleFRfD3thKKOmTk9PTWMrfD45OavYrmKC/gaAD/E8fypKz9LTsHxUTzXcDJlPs44vtvutDdTg/uWLFG701nmcgmTe/ZSH6n2eswoUMAQ+D/sx2GMUGcjcYEOExQREQUJ88xbcXwYz3X8JpavkbKEVoatUU23NeoCmobbqVcINCYqTO4b7dQrzGQ5ceiGC/1YApDno7S5jg38GzBVrzDzwcJmLGcTuJHn4XmQlCHa9H7MKIXcQJGla1bhFSDMeqCcttqMPrKz91xtxDOAI+U4NvAp7sVxuGoRRERkBp5nugTPwIfxbCXhEr5P5CTyeBxk2xp1F5HY6Xq0AIQ0JfhAcWiQFYespbR8sLMdT2fszrY4LFU8DSKeU/A8xLu5hqtZ2i1ojpwxo0fuDL+N4as46nwYy1LaGQSbAcgbhvZfgdcRo1lnrGVyeKT38DPgKeNI+QIb+BRvVYAgIiIzRzUJz9SdBLaQcB7/xsvZSR9/Sjuj+cMWYjvSajUpDQ0S48LZ6MUQwFoK/WVK/WV8SPGNdCaLmy0GQxMPFHG8DMOJvIYay3mYj+LZiuVPMXxumrMKW0g4Fs9BXEEfp1DNoGPWVB3CmpUk+ZzqEGb7rUuS0JysUB+epOdP1gKOJgkn8E/s5XAsD+qYkYiIzAwdN3q2bifhTFI2cQ0FLmU029ao+SVllqzaH5+2F94GylqMc7RqVSb3juJr7U6YOnPJk0ggkMNRAOBeHFdxAXcD8ACOFxGm5bz3G3F8HM8NvIzIV2g+vg00vd4z5f0G6VuxjNDWMaPZvcHBGsfwT39GaPve7uuAZxBHjZu5govU8lRERBQkzIdrdiuO80i5ho9R4M1MZhsolPYbYmDFUvwC3fRZa8Fa6hMVasMjhGaY6eLmTi1CCdd9W/uPwFVcyA+Azhv/s7rHkbKwFUsNyDPEKA8Cz6NB6P7Xn7sUknKBZQeuVh3CXLivk4TqyCi1PeO9tjztFCtHfkbCb9HPBOdM5YxERERm6LmmS/CsRdbhuRFLjtNJ+Qr9JISMTtsnUN83Tm18slP8uwCFEAhpSmmwn2UHr6W831AnXG3P2F3fqVioE6gS8LyONg9wM7fyblZxFikQ2Z7JcTxDA8s6AsPcRo7n0cT3HCCETlA1uHqFjhjNjciXkLapjYxncYgzUsJg2MgGxjhSk5VFRERBwvxgiOSB9bRIOJGUf6FIQsjgOIDpbP4qu4Zp12sLNlCATnFzDJG+5UtZdshaikNl8MxscbPtFje3KBJZR5MHuYF3cgtFDsMTMWzu4XuyHcu5pGzirfRxMpXs2ucOrlqBS3ILpiPWfOac63Tx6rVKKeC7MxG+xqV8nM1YDtMxIxERUZAwf6wjcCqOi9gHHIthH3lc9yhL74GCgbEde0hbrQXT8eipggVjHYOr92fJQatIyrlOoBBm7FvgiEQm8aSsxfBeWtzLNRyLIXI2ge047niW35e/6G7wbuHXgD+nmlG5dhtKywYoDgxoHsJcWEStpVWr05qo9V6s3Dly16bE+UDs6diSiIhIj9tR6cVWEs4g5RpeTp67aZCQdnvq9CoFm3csO2g1GLsojpVYazHW0pisUB0exTf8zBY3Tw1jK+K6G7QvYLmSC7gfgA+QcPozqlcwPIjlL7EcyNdx/C71DLIIHlwhYdlBawnB6xDKXLhnk4TRR3eQ1tu9HTUKpPSTUGEz1/B27sFxhLIIIiIyS883XYIenUHKVhIu459pczplbHd8V+/btwRCyzP22G6MMWAX/scVQsCnKYWBfpYddAB9q5aAMTN5BMlgcTQITBII/D4tvsnN3M67OYTTn2G9wm04XoRnLVdT5Hepk/YcIMROWD+4aj+IUQHCHAkQ6hOTpNV2FpOVLSm7yXEFEcO39QmLiMjs0ZyELNxJ4EJyXMt3OQrPAK+gjsdkM0MhNANpu0V5qJ8YFse+IYZAjJF8uUxxqB9CJK23ZnK+QuevFh5PQo4XkfImXoPjVWzneJpEDI/i+O7/tZm7B8fxeK7n9yiwhRoek8G86RT6Vi2j0NdH8HrBPBcYIhM79xCJvX26kUAJR5v1XM7XeDWO16NiExERUZAw791D5D4cJ/M1XsZaSryYZvfgUQaBgq+lBCLFgf5F1c0mhoCxluJgH/n+Et6nhEY6tUObiV2gxRBpEoj0k/BKWhzPKxjlZXyP7xI5Dcc5GD5D5BYsryfSZj9ybMOzhHb35/QYIOSGigyuWKE6hDnCJgmVkTHaE41eaxE8BRxtHqbIWfwJkdcpiyAiIgoSFo47gIdxHMLnafMSivwqDTw2g0DBQVppYnOWXLm0uNpexkgMAZcklAb7SYp50laL2AyPF3lPe6hgsHgiLTwJKylwHK/k5fwhP+UG/oPPEHk/CRHDEQRezse6x4x6r0MIQM6wbM2aTslE1P5x1gMEa0nbKZO79tFjjigSgBKGhFO4iB+zGst3lUUQEZHZpcLlrL0Xy7kErmIpOf4Z+K1MNoqd7QQEGDpgP/LlvkX7RtkmCcRAY6LC5MgYtALMZKfY0C1v7scSieT4GIbrOZ8fA7CJs8izmbGMhuylMHTgSvLlsrIIc4TL5Rh9bBftiXqvtQiefhwpn+YSXqvJyiIioiBhIZt60N/E80j5JoHVmUzY7WwqAFh60GqS3OLukW+ThOhTqqMT1EcnOtdmZoOFTlf8fgyBCoYbgW/i2UabUvff9vYda0NpxSAD+y1bsBO45+N916rWGP/Znt5bnloiOZqUeQHv4sd8AMPpyiKIiIiChIUfKFzP4cBX8ORpQSZHjwKQWJYdtAZn7OIepmUt1llCs8Xk8CitiXqnsHkmG0EFPAmOPNDofquyOBHkwRUTlh14AMErQJg7QYJl5JHH8M2098Fpgzjq3MTlXKyWpyIioiBhsZiaobCJE3D8PTU8EZvJdU/BlXIsO3ANIQYIi/ucurUW4wzNWp3q3tFOz/qZDBYCsdv61vF4s9IeA0Fg6cFrSJJEU5XnTICQ0BidYHL3CD0NOgsEchhgBzl+k8AEG4kYFSyLiMjcoMLl6XQngS0krOMHHE2VPl6dZWvU2Aq0Wy3KgwOLq5D5l4gxEkMkKRQoDQ1i84602SS24swUNxvMz4UkmQSB/auXUyiX1e50rjAGYmBs557e76lIoIwF1rGRe7kOy4E6ZiQiIgoSFo9tBDaTcB7f5HBW0M/vZtkaNdRTQvQUFSh09l4hEEMgXypRHuoHa2g3mnTrA+aHFPJDJQZWLCWkChDmzGKZS5jcN0xabfW2cnoCfTha3E+Jc1mJ4VwFCCIioiBhMQYKkS/h+DFfoMULGOD/zSxQcJBWWpBAvq+sQOHng4UIhf4+ioN9+BjxMzmM7bkKYHOWJWtXEYJOnswV1lrSVovKrpEsiuMjBSKWN3Ix/8ELsWxXkCAiInOLahJmbNfavdbn0c8KvoLhRZm2RvUwuHY/Cv19apP5f10b6yzGOVr1OpW9o6TVZic8tnPvz0qAJQeuIlcq6nOcQ1ziGN2xm3alkU3LU8/fsJGTeRjHoSpWFhGRucfqEsxYOBa5AcN7mMRyPIZHKOLwGbxB7J6Gn9i5l3ajiU30sf78tQkh4Nttcvk8yw5czeDaFZicgzbMqfe3KZSXD5ErK0CYU4tkktCo1HoPECCSYEiZxHAJYLhPhcoiIjI36bjRTPoyke04TmCMo/gmOU4iUiAQu4WvvQUKERqVKsXBAZy1RE3m/cUdWndyc65UpDw0iMkZ2vU5Uq+QQtKXY2jVSgUIcy3ONIaJXbuJIfaWew3dWoSUTVzCp9mO4490zEhERObo80+XYBZsJuFsUq7mTynwKeoEPLbnwVvdzWant/7qzu5D59qflE0SfDulNjZOY3Syc9wnmYU/SOh8E5cdtBabONWVzLF7pDoyRm3PWO8tTwsY4D+pcSgrqfIuIiiTICIic/QZqEswC84m5XYSLuczeN7JAA6Hz2TDkIBvpIzt2ou1TmHgU+3b0hTjLIMrl7P04NXkBoqQdgKtmQ4SBvdfjsvnFCDMJcbi05TayBiZVA4lGOASrmeSYzIbuSciIjItdNxottz5eGvU+3gZAwzwUmrZzVAI9ZTg25QGB7XxfMqtW+cIkksSSoP9JMU8abtFbIaZma+QQnGoTN+K5TpmNNcWx5yjune4M5ivl5Uy4OnD0eTrlLmQlVjOUrGyiIgoSJAns43IPTjeyN28lOdT4jczbY1abRNtpNjfp0DhaWOF7jC2YpHSUD8m7zr1CmmcvnxbCqbgWLpmf01UnmOstbRbbSp7RnpfJS2QwxM5ifX8rNvyVFkEERFRkCBP4QNAEYvjTlKOocDBtLLLKLQrTVw+IVcqKlB4JsFCCMQYKZZLlJb0EQ2k01Hc3N0iLlm7P84lKjKfa0GCs0zs2Udopr197lNZhHE+ylVs5j4c71AWQURE5sGzUJdglhkiOWA9NUocT+Qn2bdG3Ue73sAmia73M5SmKRHoX7GM5QevIT9YyrZeIYXyiiFyxaKyCHNtUbSWZq1OezKDlqc5DG3GGOIyIoZ7lUEQEREFCfJMXUDguzjOZyeB43CMUsAQMthQ2E6wMLZjN77Vwlp95M9Y6BQ32yRhyZqVDB20kqSU632+QgrJQIH+ZUtVhzA3A3cq+0Z6Xx0DgRKWwA1czA7uxbJOLU9FRGTePA5lzphqjXoNryHP56hjuwePev+cPLi8Y+lBawCjo0fPJd5KEgyB2kSV2vAooRme/eTmADjDsgPX4pxVFmEOfsa18QmqO0d6b3laxOD5MUVeSIkG56jlqYiIzB+qSZhLthG4nYTz+TGvYBdl/oR2RqfhLcR2pNVqUhrs107lOYghEEMkXypRHurv1Hw823oFD4OrV5ArFQleR9PnVoQAhMD4rj1ZdLYKlLEY/oz1fJ/VWL6rLIKIiChIkOfqTgJbSFjHg7ycHH28nHpGHY8shIYnpG1KQ2qN2lOwEKHQ30dxqI8YImmj1ckSPNWnlEJxSR99y5fpmNFcjBGShOrwKOlkM5uWp54vspFLOB3HB1WsLCIi8+y5qEswB52F51YSNnIpFf6aARJCRiWzCTTGalT2jeByihF7EdopzjoGV+/HkoNWkfQXnrxeIQVTdAysXIFXBmHuLYTWkjbb1EcneztmBBGHIdImz8WADnWKiMi8pF3iXHUXkTKWhDtpchQlfiWz1qiu2xo1l5ArlZRReK7M1HyFQJIklIcGcYUcrWYTWvHx7lJTZ7uWrl2FdVbXey4GCUnC5N69+EbPLU8DfTiqfJBL+QD34jhbWQQREZmHz0Zdgjm7AY3kgYtpAa/D868UcYSMNhwWJnYN01Jr1EyEEPBpSmGgjxUHH0DfqiXgDLSAFPpWLiEpFlSoPCcDBEurVqc1Ue+t5WkgkHRbnvZxNRHDfSr/ERERBQmStXUETsFxOXsJvBbHPvI4QgYFkN3WqOM7dhHUGjW7YCFNCd5TXrKEFQcfSGFJmdxgkb4lQ6pDmLPxuKWybziLY0GRIpbATaxnB19Wy1MREZnPz0eZ+x7GcSieaziGPF+gQdItZc6kNarNW5YdeADGoDfdmYbgFms67WYjgKYqz72PKEloTkwy8dhw7y1PCxgiPyHhMAI1NhIxyiSIiMg8fUbqEswDh+LZTMJlfIUWf9ZtrZjNbt5BaAbGHtvZbfuouDEzIRC8J8aoAGGuip7K8GgW1VmREoaEy9hAhaMxChBERGQ+U+HyfPHEDIXvcgyBfl6RaWvUZiBttygPqjWqLA42SaiOjtOaaPRai+Ap4/B8iw2czxYsxylAEBGRef6c1CWYR87Es4WEDVxDhQ8wQILPrjVqa7xOZd8wLqdCZlnoK5/Ft1NqI+O9BQgdBksgx3og8h0MmqwsIiLznM6WzMfP7P0YqlhabKPAq6jgsRllhdrQv2YppUEV2srC5XIJE7v20hir9p5F6McR+Fs2cNLj9UMiIiLznDIJ808kAdaTUuBEPD+gnGFr1AQqu0Zp1apqjSoLc9GzlnajSWO85wAhkmDwVMlxGajlqYiIKEiQ2XQWgc1YLmaMlGMJ7M6sNarp/D3+2F7SZkuBgiw4xlkq+0ayOBAUKGNpcxsX8O88gOUMtTwVEZEF8rzUJZjHtuM4DM+1vJSEL9KkkFlr1NAZMrXsoDVgNCVYFgabJLQqVcZ37O19cFoOg2EXCb+BY4wy8A4FCSIiskCemboE89hh3daol/JNPKdRwnbzCTGLOyO0AqM7dmOMAQ1bk4UgRirDI5n8pO5chKvYwAhHYxQgiIjIQqIWqPPdVGvU8/gex1Cjj1fTwGfVGjW2plqjDiibIPOaTRJq45M0x2q9DU4DTwFHm+3keDt/ROR41SKIiMgCe27qEiwAZ5LyHhI2cDN1/oIBEkK2rVEn9w7jcjlda5m3ok+pZjE4LWDIAZGNbKDFS9XyVEREFh7VJCyYHRCG38dyOJEmn2aIP2aSFEc2lcdt6Fu9hPLQErVGlXnHJgmV4RHqeyd6yyIEPH04Ap9nI39IzHD6uYiIyFx6duoSLJhwL3IUkcuJNHkjKd+hTJJla9TqrjGaFbVGlfm2yllCq019dKLXwWkRhyHQJs8GAE7VixYREVGQIHPdBgInY3k3k+Q4DvgZBRw+o9aoFiZ27qXdaGIT3ToyPzjnmBwZBU9vudNAoA9Lkw9wAd/jXhwf1eA0ERFRkCDzwd/gOQ3HRTyC47VYauQwmcxQ6N4tYzt249OAVccjmesLnLW06w1aE7XesgiGQIKlxQgp1xA1OE1ERBQkyHzzITxbSbiYB4m8kSIGR4RsWqOSBsZ37Or8OAUKMocZa6mMjPZ+53siZQyeG7iOndyLZZ1qEUREZOFSC9SF6s7HW6P+b17BGGX+gBY+k8Cw2xq1PdUaNWqvJHOPTSytWo3a3oneB6cVsKT8Ozn+F3+E5/XKIoiIyAJ/juoSLGBnkrKZhA28lzrvYSjb1qjtiQYTu/fiErVGlTkoQmXfaCY/h4gh5VI2UONotTwVEZGFT5mEhW4bkXtwvJG7eBm/QZnfoE6KzSBAdJBW20QLxf6yhq3JnGGThPpkheZItffBaWUcdb7J1VxIH5bjdcxIREQWwbNUl2DBixxOYBOWCqfS4l76s22NWts7RmOiotaoMndu+uCpZTE4LXb/2c9GDJFj1PJUREQUJMhCYYiUgE1UyfM6PP9BAZdJoPB4a9R9tBsNtUaV2V/UkoTa2AShGXpb4aYGp8EnWc83OAXHi9TyVEREFCTIQvIuAqfiuJDHgNeSME6Czaw1qplqjerVGlVmMSA2+HZKbWS81yxCZ3BaSpN+LgMMR6sOQUREFCTIQvSRbmvUS3iYwEkUiTgiIavWqJHxHbuIao0qs8QljtpYd3Bab1mEQAlLgzs4hx9xIpZTVYsgIiIKEmShOoOULSRs5AsEzmYAh8OTRbeWBHzDM75zTyebYHR7yQwuZtbSbrZojFV7a3kKkRyGlBGKXEfEcJSyCCIisriou9FitI3AFhLW8QAvp0A/R1HDY7KZoRDqnpC2KA0NqOORzFyQkCRM7NlLaKS9ZxH6cbS4hsu4i4dwvEdZBBERWVzUqWMxf/a34jiPlKv5OAVOpkKKJZsWRW0orxyif/lSfDvV1ZbpDRC6WYSxR3ZlMTjNEPgpOX4TS42LiBhlEkREZHFRz8rFK7IOTxtL5H+RcgD9vIwKHptBhimB2t5xbC6hNDBASBUoyDRGvNZS2TecxQHKSIKlxZVsoMI9OIyyCCIisvjo0Pii3lkRccB6GlhOwPOvlHD4DDZFBnBQ2TlMu17TDAWZvkUsSWhMVkhr7V5XNE8RR4uHKPJx3o/lCAUIIiKiIEEWowsIbMexgT1YjgOGyWfUGtXQbY26h7TVUmtUmR4xMjky2vtqFjDkgBIbuYiU72BAx4xERERBgixWh+HZTMJ6/jeW11OgRQKZtUYNMP7YbiKdYyEimS1gSUJtfJxY91kMTrOk3MWF3M0dWD6iwWkiIqIgQRa7s7utUdfzFQJ/RhnbbY3aOweh6RnbsRNjoPMPkV5XL0v0gepwRoPTIil5LgHg22rqICIii5taoMoTnmiN+l2OJtLPK6iRZtYatRlI0zbloX5i0CkO6fGWShzV4VHSarO3lSwQ6MNR4aNcyu3ci+PtyiKIiMgif87qEsgvOIuU95CwkauZ5MMMkBDIpjVRAq2xGpV9o7hcTtdaeli5LKHZpj460esqFkkwpEyS52oihvtUhyAiIqJMgvxXdxH5Po5DuJMmh1Pif9DMKKPgoD3ZxOQM+XJZw9bkud1GiWVieBhfb2eTRahzC1fySY7E8UZ1NBIREVEmQf4rQ+QoIhfhKfAGUn5IiYSQ0RGMBCq7R2nV1BpVnsOiZS3teovWWK33wWk5LC0eY5B3EzF8X1kEERERBQny5M4lcAqWixgh8FoMuyhkOEPBwPhje0ibao0qz/L2cYbKyGgWzUljd7ry9VzAKPdiWacsgoiIyNR2TeTJPYzjUDzXcySGu2lR6B486v3eCUBiWXbQGpyxBB09kqdhraXVbDD+yJ5e58V78lhSfkSe36ZIi3VENBdBRESk88zVJZCndGh3hsJGvoHlNEpYLCGTzZQFWoHxHbs7bVHVGlWePkqgui+DwWkRyGMocBkbaHC0BqeJiIgoSJBnZ2qGwsX8LSkXM4TLsj7B19uM79qDc6qjl6dYrBJLs1IhrbZ7H5xWwtHm66znU2zBcpiOGYmIiChIkGfvLDxbSdjITaRsYVnGrVHH60zuG1ZrVHly0VAdHut9cBoYDJE8G4HIdmURREREFCTIc99cnYHnTTgu4u002UZ/hoFCDur7JqmOjarjkfzXhSpJqE9M4Otpr1mEQB+WyD9wAd/idBxbNThNREREQYL0Fij8VfdYxjJOos32rFujVneP0axUFSjIL954wVMbGe89i+AwBBrkuIJORkFEREQUJEjPDJE/x/BnTGJ5LYafUcARsmuNOrFzL2mzqUBBOotUklAbmyA0QzZZhDof4gL+hXux3KEsgoiIiIIEyca7CGzHsZH/JPI6EqokGEJGHY+A0Z/tJvpUMxQW+wJlLdGn1EYmeh2cFslhaDJGjk1EDPepDkFERERBgmTrsMdbo34bz1soQjefkE2gkAZGd+zqtEVVoLA4RcBZqqPj4GOvU10CJSyB93E5P+PrGpwmIiLyVNRzUp67bQRuJ+E8fsgxjNPH79PAYzIIPi3EVqDdalIeHCBq0NqiY50lbadM7trX20oVCOSweHaR5xReQ5MxYJsyCSIiIk+xFRPpwZmk3YzCn1Pnz7NujdqeaDCxd69aoy5CxjmqI6Nk8L4/UsAQ2cR6RihgOVtZBBERkad8DusSSCb30T1YjsBzOf9IH8dTIcWSTeVxG/pWLqG8fAmhnepqLwLWWdrNNmOP7Ow9i1DAEPg3cryAMg3OJaK5CCIiIk/9LNYlkAxEHiASMeR5Cy2+TV/GrVH3jtGcqKjj0WKJOq3pZBFiBvdmgsFyFRuocZQGp4mIiDyjZ7EugWTmPd1i0JtYQ5t7iBxME4/NoPalezhkyYGryBXzhFSnRRYqmyS0azXGHt3Ta0cjTxFHyoOUOJwigXcqiyAiIvKMnse6BJKZdQROxXERj+E5loQJEmwmMxS6d+rYjt341Ks16gJmgMrwaDY/KgcUuIyLSHlIWQQREREFCTI7PoJnKwmXs53AyRQIuIze3logjYzv2NX5cQoUFt6ClCQ0Jiuk1XbvWYQ+LJ5/Yj13cQeWj2hwmoiIiIIEmT1nkLKFhI1so805DOAw2dUn+IZnbOeeTjZBB+YWlhiZHBnttTlzxGGIBBIuA+BbulNERESeDc1JkOmxjcAtJGzgfo6iyAAvo0aa1QyFUPekvkV5cFAzFBYImyQ0JiZpjtZ6zSIE+nAE/p71vJeI5bVqeSoiIqIgQeaGu4kcieNU/omX8muUOJQGKTaDQMGBr7aJJlIc6FOgsCBExnbtBXqarhxxgKFJiZO4m2HWYLhTtQgiIiLPhvpJyvQxRCKBTVgMp9HiAPo5kmpGMxRyUNs3TpLPURjoJ6SaoTBf2STptDxtBuhlbt5UFsHzIdbxL5yK4wzVIoiIiDzrZ7MugUx7oFAE1tOgwOvw/BtFEnxGxz8sTOzcR6te0wyFebsKWbxPqY1O9HrMKJLD0GKMPq4HDC9VBkFERERBgsxN6wicguMi9mA5FsMohQxboxqY2LGHtNVWa9R5GSNY6qMT0I69FqIHSlgC7+Od7OAeLKerFkFEROS5UMcPmTkP4zgUzzW8khxfoIntljL3fh+mYAuOZQeuBmNVozCPAgQfPCP/uaO3FSkQyGGAXeT4DQJjbCRilEkQERF5Ts9oXQKZMYfi2UzCZXyJyJ9RxuLwZDFDIYHQ9Iw9thtjAKNbez4wzlIbHgNPr68sIgUMgU1sYIQXYhQgiIiIKEiQ+eJsUjaTsIEPk3IVgySE7GYopLU247v34RI17przi4+1tBstGhPV3lueFrA0+TGWO4gYXqNjRiIiIr3QTkpm3jYi9+A4ma/yPzmEPl6YWWtUC77WJpig1qhzPUhIEib37sM30l5fV3SChMg7uYzvcgyOQxQkiIiI9EI1CTI7IoYbMFgSUraR8EomSXEZtEaNgIeB1cspDg6oNeqcDBAs7UabsUd29vqqwlPE0eYBchxBkcA6BQgiIiI9P6t1CWSWwtNOa9SLaZHyBlJ+SF9GR48M4GBy1zCtqlqjzs2P33XmIvRaNRA7nzWGy9lAytF68SEiIqIgQea3qdaoVzBMk+Mw7CGPy6Q1qun8Pb5zD2mrhU10q8+ZRSdJaNXrtCcbvdYieMo4PF/iMu7idiyHaXCaiIiIggSZ/z6KZzuOq/kxnteRo0lCZzBWFnd3hPEdu4g+YDRDYU4wQGVkOJsfZQgkXAnAd5RFEBERUZAgC8dheB7CcQnfwHMaJSyWANkECqEVGNuxG2MMnf6oMmsLTpLQrFRJK+3eswj9WAKf5UK+xVYsdyiLICIioiBBFpYXdmcoXML/h2cjQ7hMW6PW24zv2qPWqHNAdWSs1x8RcRgCKTmu0komIiKSPe2YZO7YRuAWEi7m67yE1fTzEprdmcwZhMO+lhLwFAcG1Bp1FtgkoTlZoT5SgVwPPygQ6MMR+DgXs4XtOP5QHY1EREQyfW7rEsiccj6er+I4hrdT5/P0kxDIpodpDurDk1RHxtXxaDbEyOTIaK+vJiIOi6dOnmsBw7c1WVlERERBgixshsgDRI4msJyTabOdUoZTmR1U947QqlQVKMzkQpMk1CcrxLrvbdUJBMoYLB/iAn7MdixnKIsgIiKiIEEWvosI/AWGcxjHcxywgwIOn2Vr1L2kzSbW6SswI2Kk1msWIRBJMLQYp8SNKIsgIiKiIEEWmbcT2I7jCv6DlNeRUCXBEDMIFLp3/ehju4hpwKo16vQuMklCbWyc0Ai9HjUKFLEEtnAuj3KPsggiIiIKEmTxOazb8ehy7qPFKZQwWGJmMxRakdGduzptUa1ao06XGALVsfFOgPBcP7lAJIelxV5yvIfrsByuAEFERERBgixOZ5NyOwmX8wk86xjEAZ4sZigk4OspYzv34KwafU3LAtPNItCMva42gRKGNreygT28FIPRUSMREZHpop2RzH13EtlKwjru4SiWMcjh1PBZtUYN9RQfPaVBtUbNNkKwEDzju/Z2AoTnmqwJBHJYPD/D8la+QotDFCCIiIhM62Ncl0DmgcgZeI7DsZFzGeHT9GXbGrUxPElldBSXU8ej7GIES3V0HNo9BAhTn38RA9zI5UzwIBYUJIiIiEwnHcSW+RQqdO7X2ygzzlewvIQ6HptBRiwCHgbXrqDQ309IU13vHgMEHzwj/7mjt5UmEChgCPw7kxzKWpq8i6ggQUREZJqf5boEMo9C2sgNGM6hiuN4Ij+lmGFrVAsTO/fRbjQ0Q6HXy+ksteGxTvVIr1kEhwGu4WbqvByjAEFERGQmtl0i8812HIfhuZoXkuerNBnAEzOpUQiANSw7eA3OOoJqFJ79mwdrSdOU0Z8+1vvgtCKGlIf5f3gJO/CcpyyCiIjIjDzPdQlk3nmiNepDeN5IkZBpa9Q0Mr5jF5HYKb6VZ/fmwTmqI6P0nN8xRCwGy1WcSJujlUUQERFRkCDyVKZao17CnXjeyQAOh8/kZyfgG57xnXs7g9aUb3vmC4q1tOsNWhN16OXEVsBTwtHmWxT4LO/D8oKMPl8RERF5WmqBKvPXnQQ2k3Ae3+bllOjnZdRIs2yNGnyb0uCgWqM+08uWWCb2DhOaaW+vICJQAAynsZ6fsD+W7coiiIiIzBS9I5X5fw/fg+UIPNfxN+R4A5OkWLKpPG5D38ohysuXEtrqePSUAYK1tFstxh7Z1dvrh4CnD0fkbjbwGu7A8jZNVxYREZnR57ougcxzkcMJbMKS8FaafKs7QyGzo0fVveM0JyrqePR00Zo1VEfGeq0a6HQzigSKXAnAg3qZISIioiBB5FnvTokUgYtpEDkez/+hiMskUPj51qh1tUZ90oXEWlr1Ju1Ko9dahEARC3ySddzHaTi2qhZBREREQYLIc7GOwFtxXM5uEo7FMkIeR8jgmIrtBAtjO3fjW+1OMbP8YixlLZWR0WyyCJ42Ja4GDEeoDkFERERBgkgvPoxnKwkX8wNSTiJPmwQya43ajow91mmNqkDh5y6NtTRrNdJKs/csQhmL4a95F99nO5bTVYsgIiKiIEGkV2d0W6NexhdpcRZlbJatUUPTM7ZjV+cYktFR+W6UQHV0tNefEnFY2tTJsQkwPKgsgoiIiIIEkaycScqtJFzGB0m5miESAtm0JkogrbWZ3L0Pp/oEbGJpVaqklXbvWYQ+DIaPcCH/xluVRRAREVGQIJK1dXhuJWEDVzDJR+knwWcXKDTGalT2DeNyiz1QsFRGxnr7EYFIgqHNJCVuAgxHKYsgIiIym/QqVBYmQwQ878OyirfxIw6kn2Oo4rEZDBHMQW3fBDaXozQ4QEgX3wwFmyQ0J6v4WhtyPYYJJRyeD7KO/+RUHKeqo5GIiMisPud1CWQBiyTA62mT5wQ8P6KUUWvU7rensmuYVr2GTdxivLpUR0Z7HZwWSbC0GKePWwDDS5VFEBERUZAgMp3OJrAFywZGcBwL7M26NerEjr2krfaimqFgk4RmpYKvp72uIoEShhabeSc7+JpqEURERBQkiMyEswhsIWE9PwZOIKGZZWvU6CPjO3YRfYpZLK1RY2QyqyxCk2H6eS8Rw8sUIIiIiChIEJm5QCFlCwkb+RrwvyhhsYRMAoUEQiswtmM3xhgwC/trZZOE+kSFWPe9ZxHKGNq8j4vYzV3Ybi2JiIiIKEgQmeFAYQMfJ+UShnCQ3QyFtN5mfNceXLLAv1YhUhvtOYsQyGFpsJuE24gY/kUBgoiIyFzhdAlkUdlG4BYSLuZr/E9W08dLaJJiMgiYLfhaSsBTHBgghoV3csYmCbWJSVpjtd56o0UC/VhqXMtVfIljcLxBR41EREQUJIjMlruJHIPjMb5AmxdT4tdokGIzCBQcpNUWJJZ8ubTgAgVDZGL3XmKMnanTz8VUFiGygxqn8U1aHKIsgoiIyFyi40ay+Bgi3yGygZQCJ5PyMH0kmbVGdVDdNUKrUl1QHY+msgihEXpdOSIFDA1u4UbG+TIWFCSIiIgoSBCZbecROAXHRYzheS2RnRQyao1qOt+s8Z17SZtN7AKpUYghUB0d770WoYClxX+wHx8gYniFjhmJiIgoSBCZKz6K52Ecl/MfpByHpY7DZDZDARh9bDfRe+w8b41qk4Ta2Dg0Y+9ZhByGJjfyTqp8WR2NREREFCSIzDWH4tlMwuXcR+QtlDE4ImQzQ4FWYOxnu8EYmK+BgrVEn1Ibncgmi9Dgx6ziY0QMr1QWQUREREGCyFx0Nim3k3AJ/0iLCxjEZVafkEDaSBnbtRs3T4MEay21sQlo95hFiEAOg2cT76DOPapFEBERmavU3UgE4M5ua9T1fIsXs4wSh9POrjVqqKX46CkNzrPWqNZC8Izv2teptTDP+Sd5ClhS/jeWd/AVAgcpQBAREZmzWwBdApGu8/F8Asc1vItRPkM/CZ40k5+dg8bwJNXRUWxunnQ8ihlmETwGh8FwLZfT4kFlEURERBQkiMwHhshx3TPyTd5MygP0Z9gaNYHq7jGak5X50RrV0qlFGJvsvRahhKXFd/lVPsGtWF6sWgQREREFCSLzKVC4BcP7mQSOJ/AIBRw+u9aoEzv30W40sW5uf/2sSzpZhLTHLIIh4oEW13AibY7GoCyCiIiIggSReeVCAm/FcQmP4ngtCZPkMMTsWqOO7diFn8utUQ3EdrvT0aiXP2LAU8TR5ttcy6e5FcsLMsrMiIiIiIIEkRn1YTxbSbiYh2jxRgqAJRIyao2aRsZ37CYYA9bMvYUhSaiOT4LveZUwJECJqzBEHsLo5hIREVGQIDJ/ndFtjXo5nyNyDgPdUWtZHJVJwDdSJh7bjbOOObV1NuDTlPpYBlmEMpaUr7ORL7AFy18riyAiIqIgQWS+O5OUzSRs4C9ocAtLs52h0J5sMLF7Ly7JzZlf2bmE+uhENlkECxS4BoCXKYsgIiKiIEFkoTgbzz04LuFCJvl7BjJujTpapTI8gpsLrVEtpO2Msgh9WAJf4Xy+xG1Ynq8sgoiIiIIEkYUjcjiBTVgKvJUW38q6NWpt7ziNidlvjepsQn18PJssggFKXAvQ7WgkIiIiChJEFhBDJAdcTI2UE/D8H4oZHT0ygOu2Rq03Zi9QsJY0bVMfm+x1ZZjKInyJd/FVZRFEREQUJIgsXBcQ2I7jSnYCr8UxQh5LyGiGgoGxnbvxrfastEZ1zmZTixAwRJRFEBERUZAgskgchmczCZfwfdqcRB6Pgyxbo449totIxMxkoGAtvt3NIvSWyOhkESJf5F18jS3KIoiIiChIEFkMzu62Rr2MLxI4kwEsLqONsIPQ9Izt3IMx0PnH9HPOUh2doOeciMEQieS6WQTNRRAREZmXnC6ByHNwJ4EtJKzjOxxNQj8vp0babfrZc+geGp6QtikN9RNDnN7fxVqiT5ncta+3FaEzF8EBd3Exm9iC5eoMjmKJiIjIjFMmQeS5OgvPrSRs4DIq/BUDJISMWqMm0BirUdk3Ou2tUTtZhLHeswium0XId7MImosgIiIybymTINKLu4iUsVjupMmRlHkezewyCu1qE5Nz5EslYsj+pby1ltBOmdw9nFUW4fNcxI3chuWPlEUQERGZr5RJEOmFIeKA9bTIcQIpP6KY0QyFbmvUyq5hWtXatLRGNVllETpXIlDsZhGOURZBREREQYLIYnYBgYdwbGAfluOw7CWPw2fXGnVi5x7SVivT1qjWWtrNFo2Jam8djQIpfVhgG+u4j1Nx/Lo6GomIiChIEFnsXojvBgr/QsrrydMkR2atUWOE8R27iDFkFigYZ6mPjmeRRbBE/OO1CAXdDiIiIgoSROSJQGEzCZfxz8DplB5vjZpJoBBagbHHdnfaovbYGjXDLILvZhE+x/nczxYsW5VFEBERUZAgIk84m5TNJFzMX9PmUgYzqk8ASCCttRnftQfneus5YJyjnl1HI09ZHY1EREQWEnU3EsnaNiIP4ziBr/Fy1lLmxTSy63jkaymBQHGg/zl1PLLWkrbbVPaMZNXR6FNcwHvZguUP1NFIRERkIVAmQSR7kUMJvAnHg5xFk7vpzzCjkIP6vgmqY2PPqeNRxlmElCLXA4btyiKIiIgoSBCRpw4UjiLyCTw53kDK9ynhMgsUHFR3j9KqPbvWqNZafLNJY6LWey1CCQt8inU8xOmqRRAREVGQICJP720ENmO5iDEsxwI7KWTUGtUCBsZ37CFtNp95oOAs1d47GkUMFk9KwnUA5PRxi4iIKEgQkWfmbALbcWzgJ1heR45alq1RAcYf20306dO2RrXWEprtLDoaBfowGP6RC3mY03BsURZBREREQYKIPHOHdVujruce2ry12xo1ZBUohFZgdMeubmvUJ/9K/8J05V6qBxyGQIscmwDDERn8HiIiIjKnqLuRyEzYRuB2Es7j+xxNlT5eTQOfVcej2Aq0W03KQwPE+EvOEllLbKdM7hl+fIrzc9KZi+AI/C0Xs4XtOHU0EhERWXiUSRCZKWeSspWEjdxCjdsYICGQZvKzE2hPNJjYO4xL/muBgPv5LMJz/9ZP1SK0KHEDAA8qiyAiIqIgQUR6cwae43CknMs4n6WfBJ9RoJCDxvAk1dFRbO6JogNrLb7dpjGZyXRlA/wj6/gBp+E4XVkEERERBQki0qvIkUSuIJLwRlK+Q1+2U5mru8doTlaxiSPG2O1oNAGeXmoRIgaHp02eG0AzEURERBYyPehFZsN7sKwjcB0HYriHwAE0CLgMAvfuu/0lB+5PoVQibbYZ/umO3l4LBDz93VqEDZxExGKURRAREVmolEkQmQ3rCLwVxyU8SsprSaiQwxAzmqEAjO3YQ4iR2uRENrUIgZT+bi3CX+ojFBERUZAgItn7MJ6tJFzOgwTeRAGwxMxmKPjIyCOP0Zio9PZNf6IW4ZOc052LcIayCCIiIgoSRGR6nEHKFhI28hnanMsgDpdRfYID32wTQ+ztm96pRfAUuREwFPSxiYiILHSakyAy27YRuIWEDdzHkQwwwEupkWYyQ6HXqqNOLYIFPsmFvI/TcNyh6coiIiILnTIJInPB+Xi+hONSLmCSf2Agw9aoz10ELAFPuVuL8FLNRRAREVkM1N1IZK6IGG7AYCmS8k84jqCKx85Sxi+Q0kdC5FNs4Dg2YzlbtQgiIiKLgTIJInMnZI8UgYupEXkdgZ9QxOFnZWMecd0y6hybAHi5XiqIiIgoSBCRmbeOwCk4LmUnnmNxjFHAEmf4mE8gUMQCn+MCHuCtOH5dtQgiIiIKEkRkdnwUz8M4LuP7BN5AQsisNeozE6E7s2Eqi1DUxyIiIqIgQURm16F4NpOwkbuBMxnAdlujTn+gEAj0YYFtXMC3eSuOLcoiiIiIKEgQkdl3dneGwgbuoM11DJEQpn2z3skiBCJJN4twoT4KERERBQkiMnechedWEtZzKRP89bS3Rp3KIhg+z4XcyxasahFEREQWn0SXQGROi6zD08KyitP5GQcwwMupTEtr1IjDEH8ui/CQOhqJiIgsRsokiMx1hkgBOJUmgdfj+ReKuGk4ejTV0ehuLuRb3IHVdGUREREFCSIyV60j8FYcl7OXyHE49pHHETKcoWAezyJcD8D9yiKIiIgoSBCRue3DeLaScAk/osnrydMigUxaowY8RSyRf+JCvsEWZRFEREQUJIjI/HAGKbeTcAVfJXA65cxao3ayBrluFuFlyiKIiIgoSBCR+ePMx1uj/hUplzPYc2tUTx+WwJe4kK9xG5bnK4sgIiKiIEFE5pezSLmVhA1cwyQfZICE8BxbowYMhieyCEcriyAiIrLYaTMgMl9FDDdhiFg820h4FRVS7LNobRzw9OEIfIWNvILbsLwjw2JoERERmZeUSRCZvyF+pw5hPSmRk0j5PuVnffRIWQQRERH5JRsEEZnfIhZD4HL+O0W+QWAVDQLuaV8CeEo4Av/MRo7hNoyyCCIiIgLKJIgshFA/8ACOq/l34HXkaJDDPG1r1IDBMpVFiMoiiIiIiIIEkYXkxXg2k7CRb5FyCkUCjsiTtUYN3Y5Gnq9yIV9SRyMRERFRkCCyEJ1NyhUkXMLfU+FKEizxSYIEg6FNpMXNQGSlsggiIiLyhP8fG3b5nADsbxsAAAAASUVORK5CYII="