from tkinter import *
from tkinter import ttk
from tkinter import filedialog, messagebox, simpledialog
from openpyxl import load_workbook
from math import ceil
from copy import deepcopy


class RegisterTab:
    def __init__(self, parent):
        self.p = parent
        self.reg = self.p.reg
        self.eep = self.p.eep
        self.spl = self.p.spl
        self.mer = self.p.mer
        self.adr = self.p.adr

        self.reg_note = ttk.Notebook(self.p.reg_frm)
        self.reg_note.pack(fill=BOTH, expand=YES)

        name_frm = ttk.Frame(self.reg_note)
        addr_frm = ttk.Frame(self.reg_note)
        self.hid0_frm = ttk.Frame(self.reg_note)
        self.hid1_frm = ttk.Frame(self.reg_note)
        self.reg_note.add(name_frm, text=" Name ")
        self.reg_note.add(self.hid0_frm, text=" Name (hidden) ")
        self.reg_note.add(addr_frm, text=" Address ")
        self.reg_note.add(self.hid1_frm, text=" Address (hidden) ")
        self.reg_note.hide(self.hid0_frm)
        self.reg_note.hide(self.hid1_frm)

        # Name Tab
        search_frm = ttk.Frame(name_frm)
        search_frm.place(relx=0, rely=0, relwidth=0.8, relheight=0.03)
        name_tree_frm = ttk.Frame(name_frm)
        name_tree_frm.place(relx=0, rely=0.03, relwidth=0.8, relheight=0.97)
        name_ctrl_frm = ttk.Frame(name_frm)
        name_ctrl_frm.place(relx=0.8, rely=0, relwidth=0.2, relheight=1)

        self.search_entry = ttk.Entry(search_frm, textvariable=self.reg.v.search, width=40)
        self.search_entry.bind("<Return>", self.search_down_key)
        self.search_entry.bind("<Shift-Return>", self.search_up_key)
        self.search_entry.bind("<Escape>", self.search_reset)
        self.search_entry.pack(side=LEFT)
        ttk.Button(search_frm, text="\u2227", style="Bigger.TButton", width=3, command=self.search_up).pack(side=LEFT)
        ttk.Button(search_frm, text="\u2228", style="Bigger.TButton", width=3, command=self.search_down).pack(side=LEFT)

        self.name_tree = ttk.Treeview(name_tree_frm, columns=("col1", "col2", "col3", "col4", "col5", "col6", "col7"))
        name_scrollbar = ttk.Scrollbar(name_tree_frm, orient=VERTICAL, command=self.name_tree.yview)
        name_scrollbar.pack(side=RIGHT, fill=Y)
        self.name_tree["yscrollcommand"] = name_scrollbar.set
        self.name_tree.heading("#0", text="Register Name")
        self.name_tree.heading("col1", text="Width")
        self.name_tree.heading("col2", text="BIN")
        self.name_tree.heading("col3", text="DEC")
        self.name_tree.heading("col4", text="HEX")
        self.name_tree.column("col1", width=60)
        self.name_tree.column("col2", width=150)
        self.name_tree.column("col3", width=60)
        self.name_tree.column("col4", width=150)
        self.name_tree.column("col5", width=20)
        self.name_tree.column("col6", width=20)
        self.name_tree.column("col7", width=150)
        self.name_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        self.name_tree.bind("<Double-1>", lambda event, click_type=1: self.popup0(event, click_type))
        self.name_tree.bind("<Button-3>", lambda event, click_type=3: self.popup0(event, click_type))
        self.name_tree.bind("<<TreeviewSelect>>", self.tree_sel)
        self.name_tree.tag_configure("modified", background="yellow")
        self.name_tree.tag_configure("eeprom", background="DarkOliveGreen3")
        self.name_tree.tag_configure("gpio", background="orange")

        ttk.Label(name_ctrl_frm, text="MAIN CONTROL\n", anchor=CENTER).pack(side=TOP, fill=X, pady=(25, 0))
        ttk.Button(name_ctrl_frm, text="Reset Register Lists", width=23, command=self.reset_reg).pack(side=TOP, anchor=CENTER, pady=(0, 13), ipadx=3)
        ttk.Button(name_ctrl_frm, text="Import Modified Register List", width=23, command=self.p.imp_reg).pack(side=TOP, anchor=CENTER, pady=(0, 3), ipadx=3)
        ttk.Button(name_ctrl_frm, text="Export Modified Register List", width=23, command=self.p.exp_reg).pack(side=TOP, anchor=CENTER, pady=(0, 13), ipadx=3)
        ttk.Button(name_ctrl_frm, text="Import Modified Byte List", width=23, command=self.p.imp_byte).pack(side=TOP, anchor=CENTER, pady=(0, 3), ipadx=3)
        ttk.Button(name_ctrl_frm, text="Export Modified Byte List", width=23, command=self.p.exp_byte).pack(side=TOP, anchor=CENTER, pady=(0, 13), ipadx=3)
        ttk.Button(name_ctrl_frm, text="Check Modified Pages", width=23, command=self.p.check_mod).pack(side=TOP, anchor=CENTER, pady=(0, 3), ipadx=3)
        ttk.Entry(name_ctrl_frm, textvariable=self.eep.v.check_mod, width=23, state="readonly").pack(side=TOP, anchor=CENTER, ipadx=4)
        ttk.Label(name_ctrl_frm, text="\n").pack(side=TOP, fill=X)
        ttk.Checkbutton(name_ctrl_frm, text="Show hidden tab", variable=self.reg.v.check0, command=lambda choice="name": self.show_hide(choice)).pack(side=TOP)

        # Address Tab
        addr_tree_frm = ttk.Frame(addr_frm)
        addr_tree_frm.place(relx=0, rely=0, relwidth=0.8, relheight=1)
        addr_ctrl_frm = ttk.Frame(addr_frm)
        addr_ctrl_frm.place(relx=0.8, rely=0, relwidth=0.2, relheight=1)

        self.addr_tree = ttk.Treeview(addr_tree_frm, columns=tuple(self.reg.l.hex_col + ["last"]))
        addr_scrollbar = ttk.Scrollbar(addr_tree_frm, orient=VERTICAL, command=self.addr_tree.yview)
        addr_scrollbar.pack(side=RIGHT, fill=Y)
        self.addr_tree["yscrollcommand"] = addr_scrollbar.set
        self.addr_tree.heading("#0", text="Address")
        self.addr_tree.column("#0", width=120)
        for (i, item) in list(enumerate(self.reg.l.hex_col)):
            if "blank" not in item:
                self.addr_tree.heading(item, text=item)
                self.addr_tree.column(item, width=20, anchor=CENTER)
            if "blank" in item:
                self.addr_tree.heading(item, text="\u250a")
                self.addr_tree.column(item, width=20, stretch=NO)
        self.addr_tree.column("last", width=1, stretch=NO)
        self.addr_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        self.addr_tree.bind("<Double-1>", lambda event, click_type=1: self.popup1(event, click_type))
        self.addr_tree.bind("<Button-3>", lambda event, click_type=3: self.popup1(event, click_type))
        self.addr_tree.tag_configure("modified", background="yellow")
        self.addr_tree.tag_configure("pg_modified", background="yellow")
        self.addr_tree.tag_configure("eeprom", background="DarkOliveGreen3")
        self.addr_tree.tag_configure("gpio", background="orange")

        ttk.Label(addr_ctrl_frm, text="MAIN CONTROL\n", anchor=CENTER).pack(side=TOP, fill=X, pady=(25, 0))
        ttk.Button(addr_ctrl_frm, text="Reset Register Lists", width=23, command=self.reset_reg).pack(side=TOP, anchor=CENTER, pady=(0, 13), ipadx=3)
        ttk.Button(addr_ctrl_frm, text="Import Modified Register List", width=23, command=self.p.imp_reg).pack(side=TOP, anchor=CENTER, pady=(0, 3), ipadx=3)
        ttk.Button(addr_ctrl_frm, text="Export Modified Register List", width=23, command=self.p.exp_reg).pack(side=TOP, anchor=CENTER, pady=(0, 13), ipadx=3)
        ttk.Button(addr_ctrl_frm, text="Import Modified Byte List", width=23, command=self.p.imp_byte).pack(side=TOP, anchor=CENTER, pady=(0, 3), ipadx=3)
        ttk.Button(addr_ctrl_frm, text="Export Modified Byte List", width=23, command=self.p.exp_byte).pack(side=TOP, anchor=CENTER, pady=(0, 13), ipadx=3)
        ttk.Button(addr_ctrl_frm, text="Check Modified Pages", width=23, command=self.p.check_mod).pack(side=TOP, anchor=CENTER, pady=(0, 3), ipadx=3)
        ttk.Entry(addr_ctrl_frm, textvariable=self.eep.v.check_mod, width=23, state="readonly").pack(side=TOP, anchor=CENTER, ipadx=4)
        ttk.Label(addr_ctrl_frm, text="\n").pack(side=TOP, fill=X)
        ttk.Checkbutton(addr_ctrl_frm, text="Show hidden tab", variable=self.reg.v.check1, command=lambda choice="addr": self.show_hide(choice)).pack(side=TOP)

        # Hidden Tab (Split Name)
        hid0_tree_frm = ttk.Frame(self.hid0_frm)
        hid0_tree_frm.place(relx=0, rely=0, relwidth=0.8, relheight=1)
        hid0_ctrl_frm = ttk.Frame(self.hid0_frm)
        hid0_ctrl_frm.place(relx=0.8, rely=0, relwidth=0.2, relheight=1)

        self.hid0_tree = ttk.Treeview(hid0_tree_frm, columns=("col1", "col2", "col3", "col4", "col5", "col6", "col7"))
        hid0_scrollbar = ttk.Scrollbar(hid0_tree_frm, orient=VERTICAL, command=self.hid0_tree.yview)
        hid0_scrollbar.pack(side=RIGHT, fill=Y)
        self.hid0_tree["yscrollcommand"] = hid0_scrollbar.set
        self.hid0_tree.heading("#0", text="Register Name")
        self.hid0_tree.heading("col1", text="Width")
        self.hid0_tree.heading("col2", text="BIN")
        self.hid0_tree.heading("col3", text="DEC")
        self.hid0_tree.heading("col4", text="HEX")
        self.hid0_tree.column("col1", width=60)
        self.hid0_tree.column("col2", width=150)
        self.hid0_tree.column("col3", width=60)
        self.hid0_tree.column("col4", width=150)
        self.hid0_tree.column("col5", width=20)
        self.hid0_tree.column("col6", width=20)
        self.hid0_tree.column("col7", width=150)
        self.hid0_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        self.hid0_tree.bind("<Button-3>", self.popup_hid0)
        self.hid0_tree.tag_configure("modified", background="yellow")
        self.hid0_tree.tag_configure("eeprom", background="DarkOliveGreen3")
        self.hid0_tree.tag_configure("gpio", background="orange")

        # Hidden Tab (Address)
        hid1_tree_frm = ttk.Frame(self.hid1_frm)
        hid1_tree_frm.place(relx=0, rely=0, relwidth=0.8, relheight=1)
        hid1_ctrl_frm = ttk.Frame(self.hid1_frm)
        hid1_ctrl_frm.place(relx=0.8, rely=0, relwidth=0.2, relheight=1)

        self.hid1_tree = ttk.Treeview(hid1_tree_frm, columns=("col1", "col2", "col3", "col4"))
        hid1_scrollbar = ttk.Scrollbar(hid1_tree_frm, orient=VERTICAL, command=self.hid1_tree.yview)
        hid1_scrollbar.pack(side=RIGHT, fill=Y)
        self.hid1_tree["yscrollcommand"] = hid1_scrollbar.set
        self.hid1_tree.heading("#0", text="Address")
        self.hid1_tree.heading("col1", text="BIN")
        self.hid1_tree.heading("col2", text="HEX")
        self.hid1_tree.column("#0", width=120)
        self.hid1_tree.column("col1", width=150)
        self.hid1_tree.column("col2", width=150)
        self.hid1_tree.column("col3", width=20)
        self.hid1_tree.column("col4", width=350)
        self.hid1_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        self.hid1_tree.bind("<Button-3>", self.popup_hid1)
        self.hid1_tree.tag_configure("modified", background="yellow")
        self.hid1_tree.tag_configure("eeprom", background="DarkOliveGreen3")
        self.hid1_tree.tag_configure("gpio", background="orange")

    def reg_add_xlsx(self):
        filename = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All Files", "*.*")))
        if filename == "":
            return False
        self.reg.l.xlsx_list.append(filename)
        self.reg.v.xlsx_names.set("\n".join(self.reg.l.xlsx_list))
        self.read_xlsx(filename, "desc")

        self.spl.l.bina_default = self.spl.l.bina.copy()
        self.spl.l.deci_default = self.spl.l.deci.copy()
        self.spl.l.hexa_default = self.spl.l.hexa.copy()
        self.mer.l.bina_default = self.mer.l.bina.copy()
        self.mer.l.deci_default = self.mer.l.deci.copy()
        self.mer.l.hexa_default = self.mer.l.hexa.copy()
        self.adr.l.bina_default = self.adr.l.bina.copy()
        self.adr.l.hexa_default = self.adr.l.hexa.copy()
        self.adr.l.hexa_per_pg_default = deepcopy(self.adr.l.hexa_per_pg)
        
        return True

    def read_xlsx(self, file, sheet):
        wb = load_workbook(filename=file)
        ws = wb[sheet]

        row = 1
        while True:
            row += 1

            page = "A" + str(row)
            grp = "B" + str(row)
            name = "C" + str(row)
            addr = "D" + str(row)
            default = "E" + str(row)
            desc = "G" + str(row)
            merge = "H" + str(row)
            ronly = "K" + str(row)

            if ws[page].value == "end":
                break

            self.spl.l.pages.append(ws[page].value)
            self.spl.l.groups.append(ws[grp].value)
            self.spl.l.names_w_width.append(ws[name].value)
            self.spl.l.addr_raw.append(ws[addr].value)
            self.spl.l.addr_total.append(str(ws[page].value) + "-" + ws[addr].value)
            self.spl.l.defaults.append(ws[default].value)
            self.spl.l.desc.append(ws[desc].value)
            self.spl.l.merge_check.append(ws[merge].value)
            self.spl.l.ronly_check.append(ws[ronly].value)

        for item in self.spl.l.pages:
            if item not in self.spl.l.pages_unique:
                self.spl.l.pages_unique.append(item)

        for item in self.spl.l.groups:
            if item not in self.spl.l.groups_unique:
                self.spl.l.groups_unique.append(item)

        self.spl.l.names_wo_width = []
        self.spl.l.widths = []
        for item in self.spl.l.names_w_width:
            if "[" not in item:
                raw_name = item
                width = "-"
            elif "[" in item:
                raw_name, width = item.split("[")
                width = width.strip("]")
            self.spl.l.names_wo_width.append(raw_name)
            self.spl.l.widths.append(width)

        for i in range(len(self.spl.l.addr_num), len(self.spl.l.addr_raw)):
            addr0, addr1 = self.spl.l.addr_raw[i].strip(")").split("(")
            self.spl.l.addr_num.append(addr0)
            self.spl.l.addr_bit.append(addr1)

        self.spl.l.bina = []
        for (i, item) in list(enumerate(self.spl.l.defaults)):
            if item[0] == "b":
                bin = item.replace("b", "")
            elif ("R" not in item) & (item[0] != "b"):
                bin = item
                print("[Warning - Excel Reading] Wrong Default Format: " + self.spl.l.names_w_width[i] + ", " + item + " (row " + str(i + 2) + ", column E)")
            else:
                bin = "-"
            self.spl.l.bina.append(bin)

        self.spl.l.desc_icon = []
        for item in self.spl.l.desc:
            if item is None:
                desc_icon = ""
            else:
                desc_icon = "ⓘ"
            self.spl.l.desc_icon.append(desc_icon)

        self.spl.l.r_icon = []
        for item in self.spl.l.ronly_check:
            if item is None:
                r_icon = ""
            else:
                r_icon = "®"
            self.spl.l.r_icon.append(r_icon)

        self.merge()
        self.bin_to_dec()
        self.dec_to_hex()

        self.spl.l.bina_mod = self.spl.l.bina.copy()
        self.mer.l.bina_mod = self.mer.l.bina.copy()
        self.spl.l.deci_mod = self.spl.l.deci.copy()
        self.mer.l.deci_mod = self.mer.l.deci.copy()
        self.spl.l.hexa_mod = self.spl.l.hexa.copy()
        self.mer.l.hexa_mod = self.mer.l.hexa.copy()

        self.name_to_addr()
        self.adr.l.bina_mod = self.adr.l.bina.copy()
        self.adr.l.hexa_mod = self.adr.l.hexa.copy()
        self.adr.l.hexa_per_pg_mod = deepcopy(self.adr.l.hexa_per_pg)

    def reg_add_treeview(self):
        self.name_tree.delete(*self.name_tree.get_children())
        self.addr_tree.delete(*self.addr_tree.get_children())
        self.hid0_tree.delete(*self.hid0_tree.get_children())
        self.hid1_tree.delete(*self.hid1_tree.get_children())

        for (i, ancestor) in list(enumerate(self.spl.l.groups_unique)):
            if ancestor != "EDID":
                self.name_tree.insert("", i, iid=ancestor, text=ancestor)
        for (i, item) in list(enumerate(self.spl.l.groups)):
            try:
                condition = (self.mer.l.names[i] != "MERGED") & (i < self.spl.l.groups.index("EDID"))
            except ValueError:
                condition = self.mer.l.names[i] != "MERGED"
            if condition:
                self.name_tree.insert(item, i, iid=str(i), text=self.mer.l.names[i], values=(self.mer.l.widths[i], self.mer.l.bina[i], self.mer.l.deci[i], self.mer.l.hexa[i], self.mer.l.desc_icon[i], self.spl.l.r_icon[i], self.mer.l.addr[i]))

        for (i, pg) in list(enumerate(self.spl.l.pages_unique)):
            if pg == 237:
                continue
            hex_val_list = []
            self.addr_tree.insert("", i, iid=str(pg), text="Page " + f"{pg:02X}")
            for (j, item) in list(enumerate(self.adr.l.num_range[i])):
                hex_val = []
                dec_from, dec_to = item.split("~")
                hex_from = f"{int(dec_from):02X}"
                hex_to = f"{int(dec_to):02X}"
                for k in range(16):
                    hex_val.append(self.adr.l.hexa_per_pg[i][j * 16 + k])
                for x in (4, 9, 14):
                    hex_val.insert(x, "\u250a")
                hex_val_list.append(hex_val)
                self.addr_tree.insert(str(pg), j, iid=str(pg) + " " + item, text=hex_from + "~" + hex_to, values=tuple(hex_val_list[j]))

        for (i, ancestor) in list(enumerate(self.spl.l.groups_unique)):
            if ancestor != "EDID":
                self.hid0_tree.insert("", i, iid=ancestor + "-split", text=ancestor)
        for (i, item) in list(enumerate(self.spl.l.groups)):
            try:
                condition = i < self.spl.l.groups.index("EDID")
            except ValueError:
                condition = True
            if condition:
                self.hid0_tree.insert(item + "-split", i, iid=str(i) + "-split", text=self.spl.l.names_wo_width[i], values=(self.spl.l.widths[i], self.spl.l.bina[i], self.spl.l.deci[i], self.spl.l.hexa[i], self.spl.l.desc_icon[i], self.spl.l.r_icon[i], self.spl.l.addr_total[i]))

        for (i, pg) in list(enumerate(self.spl.l.pages_unique)):
            if pg != 237:
                self.hid1_tree.insert("", i, iid=str(pg), text="Page " + f"{pg:02X}")
        for (j, item) in list(enumerate(self.adr.l.pg_and_num)):
            if item.split("-")[0] != "237":
                self.hid1_tree.insert(item.split("-")[0], j, iid=item, text=item.split("-")[1], values=(self.adr.l.bina[j], self.adr.l.hexa[j], self.adr.l.desc_icon[j]))

        self.bin_over_20b()
        self.get_all_children()

    def merge(self):
        warn_tups = []
        self.mer.l.names = []
        self.mer.l.widths = self.spl.l.widths.copy()
        self.mer.l.desc = self.spl.l.desc.copy()
        self.mer.l.desc_icon = self.spl.l.desc_icon.copy()
        self.mer.l.addr = self.spl.l.addr_total.copy()
        self.mer.l.bina = self.spl.l.bina.copy()

        for (i, item) in list(enumerate(self.spl.l.merge_check)):
            if item is None:
                self.mer.l.names.append("")
            elif item is not None:
                self.mer.l.names.append(self.spl.l.names_wo_width[i])

        for (i0, item0) in list(enumerate(self.mer.l.names)):
            if item0 == "":
                continue
            for (i1, item1) in list(enumerate(self.spl.l.names_wo_width)):
                if item0 == item1:
                    self.mer.l.names[i1] = item0

        for (i, item0) in enumerate(self.mer.l.names):
            sorts = []
            if (item0 == "") | (item0 == "MERGED"):
                continue
            sorts.append(self.spl.l.names_w_width[i])
            for (j, item1) in list(enumerate(self.mer.l.names)):
                if (item1 == "") | (i == j):
                    continue
                if item0 == item1:
                    sorts.append(self.spl.l.names_w_width[j])
            sorts.sort(key=lambda txt: self.sort_key(txt))

            for (k, item) in enumerate(sorts):
                if k == 0:
                    continue

                base_index = self.spl.l.names_w_width.index(sorts[0])
                prev_index = self.spl.l.names_w_width.index(sorts[k - 1])
                curr_index = self.spl.l.names_w_width.index(item)
                base_width = self.mer.l.widths[base_index]
                prev_width = self.mer.l.widths[prev_index]
                curr_width = self.mer.l.widths[curr_index]
                base_desc = self.mer.l.desc[base_index]
                prev_desc = self.mer.l.desc[prev_index]
                curr_desc = self.mer.l.desc[curr_index]

                if (":" in base_width) & (":" in curr_width):
                    if int(curr_width.split(":")[0]) > int(base_width.split(":")[0]):
                        new_width = curr_width.split(":")[0] + ":" + base_width.split(":")[1]
                        diff = int(curr_width.split(":")[1]) - int(base_width.split(":")[0])
                        step = int(curr_width.split(":")[1]) - int(prev_width.split(":")[0])
                    elif int(curr_width.split(":")[0]) < int(base_width.split(":")[0]):
                        new_width = base_width.split(":")[0] + ":" + curr_width.split(":")[1]
                        diff = int(base_width.split(":")[1]) - int(curr_width.split(":")[0])
                        step = int(prev_width.split(":")[1]) - int(curr_width.split(":")[0])
                    else:
                        diff = 0
                        step = 0
                elif (":" in base_width) & (":" not in curr_width):
                    if int(curr_width) > int(base_width.split(":")[0]):
                        new_width = curr_width + ":" + base_width.split(":")[1]
                        diff = int(curr_width) - int(base_width.split(":")[0])
                        step = int(curr_width) - int(prev_width.split(":")[0])
                    elif int(curr_width) < int(base_width.split(":")[0]):
                        new_width = base_width.split(":")[0] + ":" + curr_width
                        diff = int(base_width.split(":")[1]) - int(curr_width)
                        step = int(prev_width.split(":")[1]) - int(curr_width)
                    else:
                        diff = 0
                        step = 0
                elif (":" not in base_width) & (":" in curr_width):
                    if int(curr_width.split(":")[0]) > int(base_width):
                        new_width = curr_width.split(":")[0] + ":" + base_width
                        diff = int(curr_width.split(":")[1]) - int(base_width)
                        step = int(curr_width.split(":")[1]) - int(prev_width)
                    elif int(curr_width.split(":")[0]) < int(base_width):
                        new_width = base_width + ":" + curr_width.split(":")[1]
                        diff = int(base_width) - int(curr_width.split(":")[0])
                        step = int(prev_width) - int(curr_width.split(":")[0])
                    else:
                        diff = 0
                        step = 0
                elif (":" not in base_width) & (":" not in curr_width):
                    if int(curr_width) > int(base_width):
                        new_width = curr_width + ":" + base_width
                        diff = int(curr_width) - int(base_width)
                        step = int(curr_width) - int(prev_width)
                    elif int(curr_width) < int(base_width):
                        new_width = base_width + ":" + curr_width
                        diff = int(base_width) - int(curr_width)
                        step = int(prev_width) - int(curr_width)
                    else:
                        diff = 0
                        step = 0

                if (base_desc is None) & (curr_desc is None):
                    new_desc = None
                elif (base_desc is not None) & (curr_desc is None):
                    new_desc = base_desc
                elif (base_desc is None) & (curr_desc is not None):
                    new_desc = curr_desc
                elif (base_desc is not None) & (curr_desc is not None) & (base_desc == curr_desc):
                    new_desc = curr_desc
                elif (base_desc is not None) & (curr_desc is not None) & (base_desc != curr_desc) & ("\n\n" in base_desc):
                    new_desc = "- [" + curr_width + "]\n" + curr_desc + "\n\n" + base_desc
                elif (base_desc is not None) & (curr_desc is not None) & (base_desc != curr_desc) & ("\n\n" not in base_desc):
                    new_desc = "\u2022 [" + curr_width + "]\n" + curr_desc + "\n\n" + "\u2022 [" + prev_width + "]\n" + base_desc

                if diff == 1:
                    self.mer.l.widths[base_index] = new_width
                    self.mer.l.names[curr_index] = "MERGED"
                    self.mer.l.addr[base_index] = self.mer.l.addr[base_index] + ", " + self.spl.l.addr_total[curr_index]
                    self.mer.l.desc[base_index] = new_desc
                    if (self.spl.l.desc_icon[base_index] == "") & (self.spl.l.desc_icon[curr_index] != ""):
                        self.mer.l.desc_icon[base_index] = "ⓘ"
                    if self.spl.l.bina[base_index] != "-":
                        self.mer.l.bina[base_index] = self.spl.l.bina[curr_index] + self.mer.l.bina[base_index]
                elif (diff != 1) & (step != 1):
                    if ((self.spl.l.names_w_width[prev_index], self.spl.l.names_w_width[curr_index]) not in warn_tups) & ((self.spl.l.names_w_width[curr_index], self.spl.l.names_w_width[prev_index]) not in warn_tups):
                        warn_tups.append((self.spl.l.names_w_width[prev_index], self.spl.l.names_w_width[curr_index]))
                        print("[Warning - Merge] Bits Not in Sequence: " + self.spl.l.names_w_width[prev_index] + ", " + self.spl.l.names_w_width[curr_index])

        for (i, item) in list(enumerate(self.mer.l.names)):
            if item == "":
                self.mer.l.names[i] = self.spl.l.names_wo_width[i]

    def sort_key(self, txt):
        if ("[" in txt) & (":" in txt):
            return int(txt[txt.index("[") + 1:txt.index(":")])
        elif ("[" in txt) & (":" not in txt):
            return int(txt[txt.index("[") + 1:txt.index("]")])
        elif "[" not in txt:
            return txt

    def bin_to_dec(self):
        self.spl.l.deci = []
        for (i, item) in list(enumerate(self.spl.l.bina)):
            try:
                dec = int(item, 2)
            except ValueError:
                dec = "-"
            self.spl.l.deci.append(dec)

        self.mer.l.deci = []
        for (i, item) in list(enumerate(self.mer.l.bina)):
            try:
                dec = int(item, 2)
            except ValueError:
                dec = "-"
            self.mer.l.deci.append(dec)

    def dec_to_hex(self):
        for (i, item) in list(enumerate(self.spl.l.deci)):
            if item == "-":
                hex = "-"
            else:
                bits = int(ceil(len(self.spl.l.bina[i]) / 4))
                hex = f"{int(item):X}".zfill(bits)
            self.spl.l.hexa.append(hex)

        for (i, item) in list(enumerate(self.mer.l.deci)):
            if item == "-":
                hex = "-"
            else:
                bits = ceil(len(self.mer.l.bina[i]) / 4)
                hex = f"{int(item):X}".zfill(bits)
            self.mer.l.hexa.append(hex)

    def bin_over_20b(self, mod=False):
        if mod:
            for (i, item) in list(enumerate(self.mer.l.bina_mod)):
                try:
                    condition = (len(item) > 20) & (i < self.spl.l.groups.index("EDID"))
                except ValueError:
                    condition = len(item) > 20
                if condition:
                    self.name_tree.set(str(i), column="#2", value="-")
                    self.name_tree.set(str(i), column="#3", value="-")
        else:
            for (i, item) in list(enumerate(self.mer.l.bina)):
                try:
                    condition = (len(item) > 20) & (i < self.spl.l.groups.index("EDID"))
                except ValueError:
                    condition = len(item) > 20
                if condition:
                    self.name_tree.set(str(i), column="#2", value="-")
                    self.name_tree.set(str(i), column="#3", value="-")

    def name_to_addr(self):
        self.adr.l.num_range = []
        self.adr.l.pg_and_num = []
        self.adr.l.bina = []
        self.adr.l.hexa = []
        self.adr.l.desc = []
        self.adr.l.desc_icon = []
        self.adr.l.num_per_pg = []
        self.adr.l.hexa_per_pg = []
        prev_last_index = 0
        for (i, pg) in list(enumerate(self.spl.l.pages_unique)):
            pg_addr_range = []
            pg_bin = []
            pg_desc = []
            curr_last_index = len(self.spl.l.pages) - self.spl.l.pages[::-1].index(pg) - 1

            sorted_addr = sorted(self.spl.l.addr_num[prev_last_index:curr_last_index + 1], key=int)
            sorted_addr_unique = []
            for j in range(int(sorted_addr[-1]) + 1):
                if j not in sorted_addr_unique:
                    sorted_addr_unique.append(j)
                if str(j) not in self.spl.l.addr_num[prev_last_index:curr_last_index + 1]:
                    pg_bin.append("-")
                    pg_desc.append(None)
                elif str(j) in self.spl.l.addr_num[prev_last_index:curr_last_index + 1]:
                    pg_bin.append("00000000")
                    pg_desc.append("")

            for k in range(int(sorted_addr_unique[-1]) + 1):
                if k != int(sorted_addr_unique[-1]) - (int(sorted_addr_unique[-1]) % 16):
                    if k % 16 == 0:
                        pg_addr_range.append(str(k) + "~" + str(k + 15))
                elif k == int(sorted_addr_unique[-1]) - (int(sorted_addr_unique[-1]) % 16):
                    pg_addr_range.append(str(k) + "~" + str(sorted_addr_unique[-1]))
            self.adr.l.num_range.append(pg_addr_range)

            for (n, num) in list(enumerate(self.spl.l.addr_num))[prev_last_index:curr_last_index + 1]:
                if (":" in self.spl.l.addr_bit[n]) & (self.spl.l.bina[n] != "-"):
                    pg_bin[int(num)] = self.p.replace_text(int(self.spl.l.addr_bit[n].split(":")[1]), int(self.spl.l.addr_bit[n].split(":")[0]), self.spl.l.bina[n], text=pg_bin[int(num)])
                elif (":" in self.spl.l.addr_bit[n]) & (self.spl.l.bina[n] == "-"):
                    pg_bin[int(num)] = self.p.replace_text(int(self.spl.l.addr_bit[n].split(":")[1]), int(self.spl.l.addr_bit[n].split(":")[0]), "0" * (int(self.spl.l.addr_bit[n].split(":")[0]) - int(self.spl.l.addr_bit[n].split(":")[1]) + 1), text=pg_bin[int(num)])
                elif (":" not in self.spl.l.addr_bit[n]) & (self.spl.l.bina[n] != "-"):
                    pg_bin[int(num)] = self.p.replace_text(int(self.spl.l.addr_bit[n]), None, self.spl.l.bina[n], text=pg_bin[int(num)])
                elif (":" not in self.spl.l.addr_bit[n]) & (self.spl.l.bina[n] == "-"):
                    pg_bin[int(num)] = self.p.replace_text(int(self.spl.l.addr_bit[n]), None, "0", text=pg_bin[int(num)])
                if pg_desc[int(num)] is not None:
                    pg_desc[int(num)] = pg_desc[int(num)] + "(" + self.spl.l.addr_bit[n] + ")\n" + self.spl.l.names_w_width[n] + "\n\n"

            for x in range(int(sorted_addr[-1]) + 1):
                self.adr.l.pg_and_num.append(str(pg) + "-" + str(x))
                self.adr.l.bina.append(pg_bin[x])
                if pg_bin[x] == "-":
                    self.adr.l.hexa.append("-")
                elif pg_bin[x] != "-":
                    self.adr.l.hexa.append(f"{int(pg_bin[x], 2):02X}")
                if pg_desc[x] is None:
                    self.adr.l.desc.append(None)
                elif pg_desc is not None:
                    self.adr.l.desc.append(pg_desc[x][:-2])
                if pg_bin[x] != "-":
                    self.adr.l.desc_icon.append("ⓘ")
                elif pg_bin[x] == "-":
                    self.adr.l.desc_icon.append("")
            prev_last_index = curr_last_index + 1

        cnt = 0
        for (i, pg) in list(enumerate(self.spl.l.pages_unique)):
            pg_hex_range = []
            pg_addr = []
            for (j, item) in list(enumerate(self.adr.l.pg_and_num)):
                if str(pg) == item.split("-")[0]:
                    pg_addr.append(item.split("-")[1])
            self.adr.l.num_per_pg.append(pg_addr)

            for k in range(len(self.adr.l.num_per_pg[i])):
                pg_hex_range.append(self.adr.l.hexa[cnt + k])
                if k == len(self.adr.l.num_per_pg[i]) - 1:
                    cnt += k + 1
            pg_hex_range.extend([""] * (16 - (len(pg_hex_range) % 16)))
            self.adr.l.hexa_per_pg.append(pg_hex_range)

    # name_tree click event
    def popup0(self, event, click_type):
        col = self.name_tree.identify_column(event.x)
        row = self.name_tree.identify_row(event.y)
        name = self.name_tree.item(row, "text")
        val = self.name_tree.item(row, "value")
        if (row in self.spl.l.groups_unique) | (row == ""):
            return
        if click_type == 1:
            self.popup_entry0(col, row, name, val)
        elif click_type == 3:
            self.popup_desc0(col, row, name, val)
        else:
            raise ValueError("Invalid argument value in popup0(): 'click_type'")

    def popup_entry0(self, col, row, name, val):
        bin_valid = re.compile(r"[^01]")
        dec_valid = re.compile(r"[^0-9]")
        hex_valid = re.compile(r"[^0-9a-fA-F]")

        if (col == "#2") & (val[1] != "-"):
            new_value = simpledialog.askstring(name, "Please Enter a New BINARY Value", initialvalue=val[1])
            if (new_value is None) | (new_value == ""):
                return
            if bin_valid.search(new_value) is not None:
                messagebox.showerror("INPUT ERROR", "Input a Valid BINARY Number\n\"" + new_value + "\" is Not Valid")
                return
            if len(new_value) > len(val[1]):
                new_value = new_value[-len(val[1]):]
            if new_value.zfill(len(val[1])) != val[1]:
                self.modify0(col, row, name, val, new_value)

        if (col == "#3") & (val[2] != "-"):
            new_value = simpledialog.askstring(name, "Please Enter a New DECIMAL Value", initialvalue=val[2])
            if (new_value is None) | (new_value == ""):
                return
            if dec_valid.search(new_value) is not None:
                messagebox.showerror("INPUT ERROR", "Input a Valid DECIMAL Number\n\"" + new_value + "\" is Not Valid")
            elif new_value != val[2]:
                if int(new_value) > int("1" * len(self.mer.l.bina[int(row)]), 2):
                    new_value = str(int("1" * len(self.mer.l.bina[int(row)]), 2))
                    if new_value == val[2]:
                        return
                self.modify0(col, row, name, val, new_value)

        if (col == "#4") & (val[3] != "-"):
            new_value = simpledialog.askstring(name, "Please Enter a New HEXADECIMAL Value", initialvalue=val[3])
            if (new_value is None) | (new_value == ""):
                return
            if hex_valid.search(new_value) is not None:
                messagebox.showerror("INPUT ERROR", "Input a Valid HEXADECIMAL Number\n\"" + new_value + "\" is Not Valid")
                return
            if len(new_value) > len(val[3]):
                new_value = new_value[-len(val[3]):]
            if new_value.zfill(len(val[3])).upper() != val[3]:
                if int(new_value, 16) > int("1" * len(self.mer.l.bina[int(row)]), 2):
                    new_value = f"{int('1' * len(self.mer.l.bina[int(row)]), 2):X}".zfill(len(val[3]))
                    if new_value == val[3]:
                        return
                self.modify0(col, row, name, val, new_value)

    def popup_desc0(self, col, row, name, val):
        if (col == "#5") & (val[4] != ""):
            messagebox.showinfo(name, self.mer.l.desc[int(row)])

    def modify0(self, col, row, name, val, new, mod_tag="modified"):
        if col == "#2":
            if val[1] == new.zfill(len(self.mer.l.bina[int(row)])):
                return
            self.mer.l.bina_mod[int(row)] = new.zfill(len(self.mer.l.bina[int(row)]))
            self.mer.l.deci_mod[int(row)] = int(new, 2)
            self.mer.l.hexa_mod[int(row)] = f"{int(new, 2):X}".zfill(len(val[3]))
        if col == "#3":
            if val[2] == new:
                return
            self.mer.l.bina_mod[int(row)] = f"{int(new):b}".zfill(len(self.mer.l.bina[int(row)]))
            self.mer.l.deci_mod[int(row)] = new
            self.mer.l.hexa_mod[int(row)] = f"{int(new):X}".zfill(len(val[3]))
        if col == "#4":
            if val[3] == new.upper().zfill(len(val[3])):
                return
            self.mer.l.bina_mod[int(row)] = f"{int(new, 16):b}".zfill(len(self.mer.l.bina[int(row)]))
            self.mer.l.deci_mod[int(row)] = int(new, 16)
            self.mer.l.hexa_mod[int(row)] = new.upper().zfill(len(val[3]))

        self.name_tree.set(row, column="#2", value=self.mer.l.bina_mod[int(row)])
        self.name_tree.set(row, column="#3", value=self.mer.l.deci_mod[int(row)])
        self.name_tree.set(row, column="#4", value=self.mer.l.hexa_mod[int(row)])

        self.name_tree.item(row, tag=mod_tag)
        self.name_tree.item(self.name_tree.parent(row), tag="modified")

        self.mod0_mg2sp(row, name, mod_tag)
        self.mod0_sp2ad(name, mod_tag)
        self.modified_pages()
        self.bin_over_20b(mod=True)

    def mod0_mg2sp(self, row, name, mod_tag):
        spl_iid = []
        if self.spl.l.names_wo_width.count(name) == 1:
            spl_iid.append(int(row))
            self.spl.l.bina_mod[int(row)] = self.mer.l.bina_mod[int(row)]
            self.spl.l.deci_mod[int(row)] = int(self.spl.l.bina_mod[int(row)], 2)
            self.spl.l.hexa_mod[int(row)] = f"{int(self.spl.l.bina_mod[int(row)], 2):X}".zfill(len(self.spl.l.hexa_mod[int(row)]))
            self.hid0_tree.item(row + "-split", tag=mod_tag)
            self.hid0_tree.item(self.hid0_tree.parent(row + "-split"), tag="modified")
        elif self.spl.l.names_wo_width.count(name) > 1:
            for (i, item) in list(enumerate(self.spl.l.names_wo_width)):
                if name == item:
                    spl_iid.append(i)
                    if ":" in self.spl.l.widths[i]:
                        large, small = [int(num) for num in self.spl.l.widths[i].split(":")]
                        if self.spl.l.bina_mod[i] != self.mer.l.bina_mod[int(row)][len(self.mer.l.bina_mod[int(row)]) - 1 - large:len(self.mer.l.bina_mod[int(row)]) - small]:
                            self.spl.l.bina_mod[i] = self.mer.l.bina_mod[int(row)][len(self.mer.l.bina_mod[int(row)]) - 1 - large:len(self.mer.l.bina_mod[int(row)]) - small]
                            self.hid0_tree.item(str(i) + "-split", tag=mod_tag)
                            self.hid0_tree.item(self.hid0_tree.parent(str(i) + "-split"), tag="modified")
                    elif ":" not in self.spl.l.widths[i]:
                        if self.spl.l.bina_mod[i] != self.mer.l.bina_mod[int(row)][len(self.mer.l.bina_mod[int(row)]) - 1 - int(self.spl.l.widths[i])]:
                            self.spl.l.bina_mod[i] = self.mer.l.bina_mod[int(row)][len(self.mer.l.bina_mod[int(row)]) - 1 - int(self.spl.l.widths[i])]
                            self.hid0_tree.item(str(i) + "-split", tag=mod_tag)
                            self.hid0_tree.item(self.hid0_tree.parent(str(i) + "-split"), tag="modified")
                    self.spl.l.deci_mod[i] = int(self.spl.l.bina_mod[i], 2)
                    self.spl.l.hexa_mod[i] = f"{int(self.spl.l.bina_mod[i], 2):X}".zfill(len(self.spl.l.hexa_mod[i]))

        for iid in spl_iid:
            self.hid0_tree.set(str(iid) + "-split", column="#2", value=self.spl.l.bina_mod[iid])
            self.hid0_tree.set(str(iid) + "-split", column="#3", value=self.spl.l.deci_mod[iid])
            self.hid0_tree.set(str(iid) + "-split", column="#4", value=self.spl.l.hexa_mod[iid])

    def mod0_sp2ad(self, name, mod_tag):
        adr_iid = []
        for (i, item) in list(enumerate(self.spl.l.names_wo_width)):
            if name == item:
                page_num, bit = self.spl.l.addr_total[i].strip(")").split("(")
                page, num = page_num.split("-")
                adr_iid.append(page_num)

                if ":" in bit:
                    if self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)] != self.p.replace_text(int(bit.split(":")[1]), int(bit.split(":")[0]), self.spl.l.bina_mod[i], text=self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)]):
                        self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)] = self.p.replace_text(int(bit.split(":")[1]), int(bit.split(":")[0]), self.spl.l.bina_mod[i], text=self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)])
                        self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(page_num)] = f"{int(self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)], 2):02X}"
                        self.adr.l.hexa_per_pg_mod[self.spl.l.pages_unique.index(int(page))][int(num)] = self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(page_num)]
                        self.hid1_tree.item(page_num, tag=mod_tag)
                        self.hid1_tree.item(self.hid1_tree.parent(page_num), tag="modified")
                        self.addr_tree.item((page + " " + self.adr.l.num_range[self.spl.l.pages_unique.index(int(page))][int(num) // 16]), tag=mod_tag)
                        self.addr_tree.item(self.addr_tree.parent(page + " " + self.adr.l.num_range[self.spl.l.pages_unique.index(int(page))][int(num) // 16]), tag="pg_modified")
                elif ":" not in bit:
                    if self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)] != self.p.replace_text(int(bit), None, self.spl.l.bina_mod[i], text=self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)]):
                        self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)] = self.p.replace_text(int(bit), None, self.spl.l.bina_mod[i], text=self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)])
                        self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(page_num)] = f"{int(self.adr.l.bina_mod[self.adr.l.pg_and_num.index(page_num)], 2):02X}"
                        self.adr.l.hexa_per_pg_mod[self.spl.l.pages_unique.index(int(page))][int(num)] = self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(page_num)]
                        self.hid1_tree.item(page_num, tag=mod_tag)
                        self.hid1_tree.item(self.hid1_tree.parent(page_num), tag="modified")
                        self.addr_tree.item((page + " " + self.adr.l.num_range[self.spl.l.pages_unique.index(int(page))][int(num) // 16]), tag=mod_tag)
                        self.addr_tree.item(self.addr_tree.parent(page + " " + self.adr.l.num_range[self.spl.l.pages_unique.index(int(page))][int(num) // 16]), tag="pg_modified")

        for iid in adr_iid:
            self.hid1_tree.set(iid, column="#1", value=self.adr.l.bina_mod[self.adr.l.pg_and_num.index(iid)])
            self.hid1_tree.set(iid, column="#2", value=self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(iid)])

            pg, num = iid.split("-")
            pg_idx = self.spl.l.pages_unique.index(int(pg))
            row = pg + " " + self.adr.l.num_range[pg_idx][int(num) // 16]
            num_l, num_r = [int(x) for x in self.adr.l.num_range[pg_idx][int(num) // 16].split("~")]
            row_hexa = self.adr.l.hexa_per_pg_mod[pg_idx][num_l:num_r + 1]
            for i in (4, 9, 14):
                row_hexa.insert(i, "\u250a")
            for (i, hex) in list(enumerate(row_hexa, 1)):
                self.addr_tree.set(row, column="#" + str(i), value=hex)

    # addr_tree click event
    def popup1(self, event, click_type):
        col = self.addr_tree.identify_column(event.x)
        row = self.addr_tree.identify_row(event.y)
        name = self.addr_tree.item(row, "text")
        val = self.addr_tree.item(row, "value")
        if (row in [str(item) for item in self.spl.l.pages_unique]) | (row == "") | (col == "") | (col == "#20"):
            return
        if val[int(col.strip("#")) - 1] in ("-", "", "\u250a"):
            return
        pg, rng = row.split()
        c = int(col.strip("#")) - 1
        num = self.adr.l.num_range[self.spl.l.pages_unique.index(int(pg))].index(rng) * 16 + int(self.reg.l.hex_col[c], 16)
        if click_type == 1:
            self.popup_entry1(row, name, val, pg, c, num)
        elif click_type == 3:
            self.popup_desc1(pg, num)
        else:
            raise ValueError("Invalid argument value in popup1(): 'click_type'")

    def popup_entry1(self, row, name, val, pg, c, num):
        hex_valid = re.compile(r"[^0-9a-fA-F]")

        new_value = simpledialog.askstring(name, "Please Enter a New HEXADECIMAL Value", initialvalue=val[c])
        if (new_value is None) | (new_value == ""):
            return
        if hex_valid.search(new_value) is not None:
            messagebox.showerror("INPUT ERROR", "Input a Valid HEXADECIMAL Number\n\"" + new_value + "\" is Not Valid")
            return

        if len(new_value) > 2:
            new_value = new_value[-2:]
        if new_value.upper().zfill(2) != val[c]:
            self.modify1(row, pg, num, new_value, mod_tag="modified")

    def popup_desc1(self, pg, num):
        messagebox.showinfo(pg + "-" + str(num), self.adr.l.desc[self.adr.l.pg_and_num.index(pg + "-" + str(num))])

    def modify1(self, row, pg, num, new, mod_tag="modified"):
        pg_idx = self.spl.l.pages_unique.index(int(pg))
        if new.upper().zfill(2) == self.adr.l.hexa_per_pg_mod[pg_idx][num]:
            return

        self.adr.l.hexa_per_pg_mod[pg_idx][num] = new.upper().zfill(2)

        num_l, num_r = [int(x) for x in self.adr.l.num_range[pg_idx][int(num) // 16].split("~")]
        row_hexa = self.adr.l.hexa_per_pg_mod[pg_idx][num_l:num_r + 1]
        for i in (4, 9, 14):
            row_hexa.insert(i, "\u250a")
        for (i, hex) in list(enumerate(row_hexa, 1)):
            self.addr_tree.set(row, column="#" + str(i), value=hex)

        self.addr_tree.item(row, tag=mod_tag)
        self.addr_tree.item(self.addr_tree.parent(row), tag="pg_modified")

        self.mod1_ad2sp(pg, num, mod_tag)
        self.mod1_sp2mg(pg, num, mod_tag)
        self.modified_pages()
        self.bin_over_20b(mod=True)

    def mod1_ad2sp(self, pg, num, mod_tag):
        self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(pg + "-" + str(num))] = self.adr.l.hexa_per_pg_mod[self.spl.l.pages_unique.index(int(pg))][num]
        self.adr.l.bina_mod[self.adr.l.pg_and_num.index(pg + "-" + str(num))] = f"{int(self.adr.l.hexa_per_pg_mod[self.spl.l.pages_unique.index(int(pg))][num], 16):08b}"

        spl_idx = []
        for (idx, addr) in list(enumerate([x.split("(")[0] for x in self.spl.l.addr_total])):
            if addr == pg + "-" + str(num):
                spl_idx.append(idx)

        for i in spl_idx:
            if ":" in self.spl.l.addr_bit[i]:
                self.spl.l.bina_mod[i] = self.adr.l.bina_mod[self.adr.l.pg_and_num.index(pg + "-" + str(num))][7 - int(self.spl.l.addr_bit[i].split(":")[0]):8 - int(self.spl.l.addr_bit[i].split(":")[1])]
            elif ":" not in self.spl.l.addr_bit[i]:
                self.spl.l.bina_mod[i] = self.adr.l.bina_mod[self.adr.l.pg_and_num.index(pg + "-" + str(num))][7 - int(self.spl.l.addr_bit[i])]
            self.spl.l.deci_mod[i] = int(self.spl.l.bina_mod[i], 2)
            self.spl.l.hexa_mod[i] = f"{int(self.spl.l.bina_mod[i], 2):X}".zfill(len(self.spl.l.hexa[i]))

        self.hid1_tree.set(pg + "-" + str(num), column="#1", value=self.adr.l.bina_mod[self.adr.l.pg_and_num.index(pg + "-" + str(num))])
        self.hid1_tree.set(pg + "-" + str(num), column="#2", value=self.adr.l.hexa_mod[self.adr.l.pg_and_num.index(pg + "-" + str(num))])

        for j in spl_idx:
            self.hid0_tree.set(str(j) + "-split", column="#2", value=self.spl.l.bina_mod[j])
            self.hid0_tree.set(str(j) + "-split", column="#3", value=self.spl.l.deci_mod[j])
            self.hid0_tree.set(str(j) + "-split", column="#4", value=self.spl.l.hexa_mod[j])

        self.hid1_tree.item(pg + "-" + str(num), tag=mod_tag)
        self.hid1_tree.item(self.hid1_tree.parent(pg + "-" + str(num)), tag="modified")
        for k in spl_idx:
            if self.spl.l.bina_mod[k] != self.spl.l.bina[k]:
                self.hid0_tree.item(str(k) + "-split", tag=mod_tag)
                self.hid0_tree.item(self.hid0_tree.parent(str(k) + "-split"), tag="modified")

    def mod1_sp2mg(self, pg, num, mod_tag):
        spl_idx = []
        for (idx, addr) in list(enumerate([x.split("(")[0] for x in self.spl.l.addr_total])):
            if addr == pg + "-" + str(num):
                spl_idx.append(idx)

        for i in spl_idx:
            mer_idx = self.mer.l.names.index(self.spl.l.names_wo_width[i])
            if self.spl.l.names_wo_width.count(self.spl.l.names_wo_width[i]) == 1:
                self.mer.l.bina_mod[mer_idx] = self.spl.l.bina_mod[i]
                self.mer.l.deci_mod[mer_idx] = self.spl.l.deci_mod[i]
                self.mer.l.hexa_mod[mer_idx] = self.spl.l.hexa_mod[i]
            elif self.spl.l.names_wo_width.count(self.spl.l.names_wo_width[i]) > 1:
                if ":" in self.spl.l.widths[i]:
                    self.mer.l.bina_mod[mer_idx] = self.p.replace_text(int(self.spl.l.widths[i].split(":")[1]), int(self.spl.l.widths[i].split(":")[0]), self.spl.l.bina_mod[i], self.mer.l.bina_mod[mer_idx])
                elif ":" not in self.spl.l.widths[i]:
                    self.mer.l.bina_mod[mer_idx] = self.p.replace_text(int(self.spl.l.widths[i]), None, self.spl.l.bina_mod[i], self.mer.l.bina_mod[mer_idx])
                self.mer.l.deci_mod[mer_idx] = int(self.mer.l.bina_mod[mer_idx], 2)
                self.mer.l.hexa_mod[mer_idx] = f"{self.mer.l.deci_mod[mer_idx]:X}".zfill(len(self.mer.l.hexa[mer_idx]))

        for j in spl_idx:
            mer_idx = self.mer.l.names.index(self.spl.l.names_wo_width[j])
            self.name_tree.set(str(mer_idx), column="#2", value=self.mer.l.bina_mod[mer_idx])
            self.name_tree.set(str(mer_idx), column="#3", value=self.mer.l.deci_mod[mer_idx])
            self.name_tree.set(str(mer_idx), column="#4", value=self.mer.l.hexa_mod[mer_idx])
            if self.mer.l.bina_mod[mer_idx] != self.mer.l.bina[mer_idx]:
                self.name_tree.item(mer_idx, tag=mod_tag)
                self.name_tree.item(self.name_tree.parent(mer_idx), tag="modified")

    # hid0_tree (split name) click event
    def popup_hid0(self, event):
        col = self.hid0_tree.identify_column(event.x)
        row = self.hid0_tree.identify_row(event.y)
        name = self.hid0_tree.item(row, "text")
        val = self.hid0_tree.item(row, "value")
        if (row in self.spl.l.groups_unique) | (row == ""):
            return
        if (col == "#5") & (val[4] != ""):
            messagebox.showinfo(name, self.spl.l.desc[int(row.split("-")[0])])

    # hid1_tree (address) click event
    def popup_hid1(self, event):
        col = self.hid1_tree.identify_column(event.x)
        row = self.hid1_tree.identify_row(event.y)
        name = self.hid1_tree.item(row, "text")
        val = self.hid1_tree.item(row, "value")
        if (row in [str(x) for x in self.spl.l.pages_unique]) | (row == ""):
            return
        if (col == "#3") & (val[2] != ""):
            messagebox.showinfo(row, self.adr.l.desc[self.adr.l.pg_and_num.index(row)])

    # Treeview etc
    def get_all_children(self):
        for (i, grp) in list(enumerate(self.name_tree.get_children())):
            self.reg.l.all_children.append(grp)
            for iid in list(self.name_tree.get_children(grp)):
                self.reg.l.all_children.append(self.spl.l.names_wo_width[int(iid)])
                self.reg.l.children.append((iid, self.spl.l.names_wo_width[int(iid)]))

    def tree_sel(self, event):
        for item in self.name_tree.selection():
            sel_item = self.name_tree.item(item, "text")
            break

        try:
            self.reg.v.curr_row.set(self.reg.l.all_children.index(sel_item))
        except UnboundLocalError:
            self.reg.v.curr_row.set(0)
        self.children_down = self.reg.l.all_children[self.reg.v.curr_row.get() + 1:]
        self.children_up = list(reversed(self.reg.l.all_children[:self.reg.v.curr_row.get()]))

    def search_reset(self, event):
        self.name_tree.selection_clear()
        self.name_tree.yview_moveto(0)
        self.search_entry.delete(0, END)
        for grp_iid in self.spl.l.groups_unique:
            if grp_iid != "EDID":
                self.name_tree.item(grp_iid, open=False)

    def search_down(self):
        if len(self.name_tree.selection()) == 0:
            self.reg.v.curr_row.set(-1)
            self.children_down = self.reg.l.all_children
        for child in self.children_down:
            if child not in self.spl.l.groups_unique:
                child_iid = str(self.mer.l.names.index(child))
            else:
                child_iid = child
            parent_iid = self.name_tree.parent(child_iid)
            for grp_iid in self.spl.l.groups_unique:
                if grp_iid != "EDID":
                    self.name_tree.item(grp_iid, open=False)
            if self.reg.v.search.get().lower() in child.lower():
                self.name_tree.selection_set(child_iid)
                self.name_tree.item(parent_iid, open=True)
                self.move_yview()
                break
            if (child == self.children_down[-1]) & (self.reg.v.search.get().lower() not in child):
                self.name_tree.selection_clear()
                self.reg.v.curr_row.set(-1)
                self.children_down = self.reg.l.all_children
                self.name_tree.yview_moveto(0)

    def search_down_key(self, event):
        if len(self.name_tree.selection()) == 0:
            self.reg.v.curr_row.set(-1)
            self.children_down = self.reg.l.all_children
        for child in self.children_down:
            if child not in self.spl.l.groups_unique:
                child_iid = str(self.mer.l.names.index(child))
            else:
                child_iid = child
            parent_iid = self.name_tree.parent(child_iid)
            for grp_iid in self.spl.l.groups_unique:
                if grp_iid != "EDID":
                    self.name_tree.item(grp_iid, open=False)
            if self.reg.v.search.get().lower() in child.lower():
                self.name_tree.selection_set(child_iid)
                self.name_tree.item(parent_iid, open=True)
                self.move_yview()
                break
            if (child == self.children_down[-1]) & (self.reg.v.search.get().lower() not in child):
                self.name_tree.selection_clear()
                self.reg.v.curr_row.set(-1)
                self.children_down = self.reg.l.all_children
                self.name_tree.yview_moveto(0)

    def search_up(self):
        if len(self.name_tree.selection()) == 0:
            self.reg.v.curr_row.set(None)
            self.children_up = list(reversed(self.reg.l.all_children))
        for child in self.children_up:
            if child not in self.spl.l.groups_unique:
                child_iid = str(self.mer.l.names.index(child))
            else:
                child_iid = child
            parent_iid = self.name_tree.parent(child_iid)
            for grp_iid in self.spl.l.groups_unique:
                if grp_iid != "EDID":
                    self.name_tree.item(grp_iid, open=False)
            if self.reg.v.search.get().lower() in child.lower():
                self.name_tree.selection_set(child_iid)
                self.name_tree.item(parent_iid, open=True)
                self.move_yview()
                break
            if (child == self.children_up[-1]) & (self.reg.v.search.get().lower() not in child):
                self.name_tree.selection_clear()
                self.reg.v.curr_row.set(None)
                self.children_up = list(reversed(self.reg.l.all_children))
                self.name_tree.yview_moveto(1)

    def search_up_key(self, event):
        if len(self.name_tree.selection()) == 0:
            self.reg.v.curr_row.set(None)
            self.children_up = list(reversed(self.reg.l.all_children))
        for child in self.children_up:
            if child not in self.spl.l.groups_unique:
                child_iid = str(self.mer.l.names.index(child))
            else:
                child_iid = child
            parent_iid = self.name_tree.parent(child_iid)
            for grp_iid in self.spl.l.groups_unique:
                if grp_iid != "EDID":
                    self.name_tree.item(grp_iid, open=False)
            if self.reg.v.search.get().lower() in child.lower():
                self.name_tree.selection_set(child_iid)
                self.name_tree.item(parent_iid, open=True)
                self.move_yview()
                break
            if (child == self.children_up[-1]) & (self.reg.v.search.get().lower() not in child):
                self.name_tree.selection_clear()
                self.reg.v.curr_row.set(None)
                self.children_up = list(reversed(self.reg.l.all_children))
                self.name_tree.yview_moveto(1)

    def move_yview(self):
        for item in self.name_tree.selection():
            item_text = self.name_tree.item(item, "text")
            break

        if item_text in self.spl.l.groups_unique:
            fraction = self.spl.l.groups_unique.index(item_text) / len(self.spl.l.groups_unique)
        elif item_text not in self.spl.l.groups_unique:
            child_iid = str(self.mer.l.names.index(item_text))
            parent_iid = self.name_tree.parent(child_iid)
            numerator = self.spl.l.groups_unique.index(parent_iid) + list(self.name_tree.get_children(parent_iid)).index(child_iid) + 1
            denominator = len(self.spl.l.groups_unique) + len(list(self.name_tree.get_children(parent_iid)))
            fraction = numerator / denominator

        self.name_tree.yview_moveto(fraction)

    def modified_pages(self):
        for iid in self.addr_tree.tag_has("pg_modified"):
            if iid not in self.reg.l.modified_pg:
                self.reg.l.modified_pg.append(iid)

    def reset_reg(self):
        for (i, item) in list(enumerate(self.mer.l.names)):
            if (item != "MERGED") & (i < self.spl.l.groups.index("EDID")):
                self.name_tree.set(str(i), column="#2", value=self.mer.l.bina[i])
                self.name_tree.set(str(i), column="#3", value=self.mer.l.deci[i])
                self.name_tree.set(str(i), column="#4", value=self.mer.l.hexa[i])

        for j in range(len(self.spl.l.names_w_width)):
            if j < self.spl.l.groups.index("EDID"):
                self.hid0_tree.set(str(j) + "-split", column="#2", value=self.spl.l.bina[j])
                self.hid0_tree.set(str(j) + "-split", column="#3", value=self.spl.l.deci[j])
                self.hid0_tree.set(str(j) + "-split", column="#4", value=self.spl.l.hexa[j])

        for (k, item) in list(enumerate(self.adr.l.pg_and_num)):
            if item.split("-")[0] != "237":
                self.hid1_tree.set(item, column="#1", value=self.adr.l.bina[k])
                self.hid1_tree.set(item, column="#2", value=self.adr.l.hexa[k])

        for (x, pg) in list(enumerate(self.spl.l.pages_unique)):
            if pg == 237:
                continue
            hex_val_list = []
            for (y, item) in list(enumerate(self.adr.l.num_range[x])):
                hex_val = []
                for z in range(16):
                    hex_val.append(self.adr.l.hexa_per_pg[x][y * 16 + z])
                for w in (4, 9, 14):
                    hex_val.insert(w, "\u250a")
                hex_val_list.append(hex_val)
            for (y, item) in list(enumerate(self.adr.l.num_range[x])):
                for col in range(19):
                    self.addr_tree.set(str(pg) + " " + item, column="#" + str(col + 1), value=hex_val_list[y][col])

        self.mer.l.bina_mod[:self.spl.l.pages.index(237)] = self.mer.l.bina[:self.spl.l.pages.index(237)]
        self.mer.l.deci_mod[:self.spl.l.pages.index(237)] = self.mer.l.deci[:self.spl.l.pages.index(237)]
        self.mer.l.hexa_mod[:self.spl.l.pages.index(237)] = self.mer.l.hexa[:self.spl.l.pages.index(237)]
        self.spl.l.bina_mod[:self.spl.l.pages.index(237)] = self.spl.l.bina[:self.spl.l.pages.index(237)]
        self.spl.l.deci_mod[:self.spl.l.pages.index(237)] = self.spl.l.deci[:self.spl.l.pages.index(237)]
        self.spl.l.hexa_mod[:self.spl.l.pages.index(237)] = self.spl.l.hexa[:self.spl.l.pages.index(237)]
        self.adr.l.bina_mod[:self.spl.l.pages.index(237)] = self.adr.l.bina[:self.spl.l.pages.index(237)]
        self.adr.l.hexa_mod[:self.spl.l.pages.index(237)] = self.adr.l.hexa[:self.spl.l.pages.index(237)]
        self.adr.l.hexa_per_pg_mod[:self.spl.l.pages_unique.index(237)] = deepcopy(self.adr.l.hexa_per_pg[:self.spl.l.pages_unique.index(237)])
        self.reg.l.modified_pg = list(set(self.reg.l.modified_pg) - (set(self.reg.l.modified_pg) - {"237"}))
        self.eep.l.mod_by_eep_pg = list(set(self.eep.l.mod_by_eep_pg) - (set(self.eep.l.mod_by_eep_pg) - {"237"}))

        for iid in self.name_tree.tag_has("modified") + self.name_tree.tag_has("eeprom") + self.name_tree.tag_has("gpio"):
            self.name_tree.item(iid, tag="")
        for iid in self.addr_tree.tag_has("modified") + self.addr_tree.tag_has("pg_modified") + self.addr_tree.tag_has("eeprom") + self.addr_tree.tag_has("gpio"):
            self.addr_tree.item(iid, tag="")
        for iid in self.hid0_tree.tag_has("modified") + self.hid0_tree.tag_has("eeprom") + self.hid0_tree.tag_has("gpio"):
            self.hid0_tree.item(iid, tag="")
        for iid in self.hid1_tree.tag_has("modified") + self.hid1_tree.tag_has("eeprom") + self.hid1_tree.tag_has("gpio"):
            self.hid1_tree.item(iid, tag="")

        if len(self.eep.l.mod_by_eep_pg) == 0:
            for file in self.eep.l.imported_txt:
                if file in self.reg.l.xlsx_list:
                    self.reg.l.xlsx_list.remove(file)
            self.reg.v.xlsx_names.set("\n".join(self.reg.l.xlsx_list))

        self.bin_over_20b()
        self.p.check_mod()

    def show_hide(self, choice):
        if choice.lower() == "name":
            if self.reg.v.check0.get() == 0:
                self.reg_note.hide(self.hid0_frm)
            elif self.reg.v.check0.get() == 1:
                self.reg_note.add(self.hid0_frm)
        elif choice.lower() == "addr":
            if self.reg.v.check1.get() == 0:
                self.reg_note.hide(self.hid1_frm)
            elif self.reg.v.check1.get() == 1:
                self.reg_note.add(self.hid1_frm)

    def parent_tag_check(self):
        for parent in self.name_tree.get_children():
            if not self.name_tree.tag_has("modified", parent):
                continue
            cnt = 0
            for child in self.name_tree.get_children(parent):
                if self.name_tree.tag_has("modified", child) | self.name_tree.tag_has("eeprom", child) | self.name_tree.tag_has("gpio", child):
                    cnt += 1
            if cnt == 0:
                self.name_tree.item(parent, tag="")
        
        for parent in self.addr_tree.get_children():
            if not self.addr_tree.tag_has("pg_modified", parent):
                continue
            cnt = 0
            for child in self.addr_tree.get_children(parent):
                if self.addr_tree.tag_has("modified", child) | self.addr_tree.tag_has("eeprom", child) | self.addr_tree.tag_has("gpio", child):
                    cnt += 1
            if cnt == 0:
                self.addr_tree.item(parent, tag="")
        
        for parent in self.hid0_tree.get_children():
            if not self.hid0_tree.tag_has("modified", parent):
                continue
            cnt = 0
            for child in self.hid0_tree.get_children(parent):
                if self.hid0_tree.tag_has("modified", child) | self.hid0_tree.tag_has("eeprom", child) | self.hid0_tree.tag_has("gpio", child):
                    cnt += 1
            if cnt == 0:
                self.hid0_tree.item(parent, tag="")
        
        for parent in self.hid1_tree.get_children():
            if not self.hid1_tree.tag_has("modified", parent):
                continue
            cnt = 0
            for child in self.hid1_tree.get_children(parent):
                if self.hid1_tree.tag_has("modified", child) | self.hid1_tree.tag_has("eeprom", child) | self.hid1_tree.tag_has("gpio", child):
                    cnt += 1
            if cnt == 0:
                self.hid1_tree.item(parent, tag="")